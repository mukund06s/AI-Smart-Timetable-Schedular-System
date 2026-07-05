# app.py - Part 1: Imports, Constants, Firebase Initialization
# UPDATED VERSION with dynamic time slots and custom lunch/break support
import warnings
warnings.filterwarnings('ignore')

from utils.logging_config import configure_logging, get_logger

configure_logging()
logger = get_logger("app")

import streamlit as st

st.set_page_config(
    page_title="Smart Classroom & Timetable Scheduler",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date, time
import random
from collections import defaultdict, deque
import plotly.graph_objects as go
import plotly.express as px
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier, GradientBoostingRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import json
import io
import copy
import warnings
from scipy.optimize import linear_sum_assignment
import networkx as nx
from genetic_algorithm import GeneticAlgorithm, create_constraints
import firebase_admin
from firebase_admin import credentials, firestore, auth
from google.cloud.firestore_v1 import FieldFilter
import time as time_module
from typing import Dict, List, Any, Optional, Tuple
import hashlib
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

@st.cache_resource
def initialize_firebase():
    if not firebase_admin._apps:
        cred = credentials.Certificate(dict(st.secrets["firebase"]))
        firebase_admin.initialize_app(cred)
    return firestore.client()

db = initialize_firebase()


# ==================== CHANGE 1: UPDATED CONFIGURATION WITH DYNAMIC TIME SUPPORT ====================

# CHANGE 1: Default configurations
DEFAULT_LECTURE_DURATION = 60  # minutes for theory
DEFAULT_LAB_DURATION = 120  # minutes for lab (2 hours)
DEFAULT_LUNCH_DURATION = 50  # minutes
DEFAULT_BREAK_DURATION = 10  # minutes
DEFAULT_DAY_START = "09:00"  # 9:00 AM
DEFAULT_DAY_END = "16:00"  # 4:00 PM

# CHANGE 1: Default lunch start times by school
DEFAULT_LUNCH_START_TIMES = {
    'STME': '13:00',  # 1:00 PM
    'SOC': '11:00',   # 11:00 AM
    'SOL': '12:00'    # 12:00 PM
}

# CHANGE 3: Faculty morning constraint
FACULTY_MORNING_LIMIT = 2  # Max 2 lectures at 9 AM per faculty per week
MORNING_SLOT_START = "09:00"

# Keep legacy SCHOOL_LUNCH_TIMES for backward compatibility
SCHOOL_LUNCH_TIMES = {
    'STME': '13:00-13:50',
    'SOC': '11:00-11:50',
    'SOL': '12:00-12:50'
}

# CHANGE 1: Updated program configuration with default lunch settings
PROGRAM_CONFIG = {
    'BTECH': {
        'school': 'STME',
        'name': 'Bachelor of Technology',
        'semesters': 8,
        'default_lunch_start': '13:00',
        'default_lunch_duration': 50
    },
    'BTECH_AIDS': {
        'school': 'STME',
        'name': 'Bachelor of Technology - AIDS',
        'semesters': 8,
        'default_lunch_start': '13:00',
        'default_lunch_duration': 50
    },
    'MBATECH': {
        'school': 'STME',
        'name': 'MBA in Technology',
        'semesters': 10,
        'default_lunch_start': '13:00',
        'default_lunch_duration': 50
    },
    'BBA': {
        'school': 'SOC',
        'name': 'Bachelor of Business Administration',
        'semesters': 6,
        'default_lunch_start': '11:00',
        'default_lunch_duration': 50
    },
    'BCOM': {
        'school': 'SOC',
        'name': 'Bachelor of Commerce',
        'semesters': 6,
        'default_lunch_start': '11:00',
        'default_lunch_duration': 50
    },
    'LAW': {
        'school': 'SOL',
        'name': 'Bachelor of Law',
        'semesters': 10,
        'default_lunch_start': '12:00',
        'default_lunch_duration': 50
    }
}

SCHOOL_CONFIG = {
    'STME': {
        'name': 'School of Technology, Management and Engineering',
        'programs': ['BTECH', 'BTECH_AIDS', 'MBATECH'],
        'default_lunch_start': '13:00',
        'default_lunch_duration': 50
    },
    'SOC': {
        'name': 'School of Commerce',
        'programs': ['BBA', 'BCOM'],
        'default_lunch_start': '11:00',
        'default_lunch_duration': 50
    },
    'SOL': {
        'name': 'School of Law',
        'programs': ['LAW'],
        'default_lunch_start': '12:00',
        'default_lunch_duration': 50
    }
}

# CHANGE 1: Info Dataset column definitions with tooltips
INFO_DATASET_COLUMNS = {
    'S.No.': 'Serial number of the record',
    'Program': 'Academic program (e.g., B.Tech)',
    'Sem': 'Semester number of the program',
    'Section': 'Section or class group (A, B, etc.)',
    'Batch': 'Batch number or student batch identifier',
    'Module Name': 'Name of the subject/course',
    'Theory Hrs/Week': 'Weekly hours allocated for theory lectures',
    'Practical Hrs/Week': 'Weekly hours allocated for lab/practical sessions',
    'Tutorial Hrs/Week': 'Weekly hours allocated for tutorials (if any)',
    'Theory Load': 'Teaching load generated from theory hours',
    'Practical Load': 'Teaching load generated from practical hours',
    'Total Load': 'Sum of theory and practical teaching load',
    'Faculty': 'Faculty member assigned to the module'
}

# CHANGE 1: Room Dataset column definitions
ROOM_DATASET_COLUMNS = {
    'Subject': 'Module name (must match Info Dataset Module Name)',
    'Class Type': 'Type of class: theory, lab, or tutorial',
    'Room No.': 'Room identifier (e.g., Room-101)'
}


# ==================== CHANGE 1: DYNAMIC TIME SLOT UTILITIES ====================

class TimeSlotManager:
    """
    CHANGE 1: Manages dynamic time slots based on lunch/break configurations
    Replaces fixed 1-hour slots with flexible slot generation
    """
    
    @staticmethod
    def time_to_minutes(time_str: str) -> int:
        """Convert time string (HH:MM) to minutes from midnight"""
        if isinstance(time_str, time):
            return time_str.hour * 60 + time_str.minute
        parts = time_str.split(':')
        return int(parts[0]) * 60 + int(parts[1])
    
    @staticmethod
    def minutes_to_time(minutes: int) -> str:
        """Convert minutes from midnight to time string (HH:MM)"""
        hours = minutes // 60
        mins = minutes % 60
        return f"{hours:02d}:{mins:02d}"
    
    @staticmethod
    def format_time_12hr(time_str: str) -> str:
        """Convert 24hr time to 12hr format for display"""
        parts = time_str.split(':')
        hour = int(parts[0])
        minute = int(parts[1])
        
        if hour == 0:
            return f"12:{minute:02d} AM"
        elif hour < 12:
            return f"{hour}:{minute:02d} AM"
        elif hour == 12:
            return f"12:{minute:02d} PM"
        else:
            return f"{hour-12}:{minute:02d} PM"
    
    @staticmethod
    def generate_dynamic_slots(
        day_start: str = "09:00",
        day_end: str = "16:00",
        lunch_start: str = "13:00",
        lunch_duration: int = 50,
        breaks: List[dict] = None,
        lecture_duration: int = 60
    ) -> List[dict]:
        """
        CHANGE 1: Generate dynamic time slots based on configuration
        
        Returns list of slot dictionaries:
        [
            {'start': '09:00', 'end': '10:00', 'type': 'lecture', 'index': 1},
            {'start': '13:00', 'end': '13:50', 'type': 'lunch', 'index': None},
            {'start': '14:00', 'end': '14:10', 'type': 'break', 'index': None},
            ...
        ]
        """
        slots = []
        breaks = breaks or []
        
        start_minutes = TimeSlotManager.time_to_minutes(day_start)
        end_minutes = TimeSlotManager.time_to_minutes(day_end)
        lunch_start_minutes = TimeSlotManager.time_to_minutes(lunch_start)
        lunch_end_minutes = lunch_start_minutes + lunch_duration
        
        # Sort breaks by placement (after which lecture)
        break_after_lectures = {}
        for brk in breaks:
            placement = brk.get('placement', [])
            duration = brk.get('duration', 10)
            for p in placement:
                break_after_lectures[p] = duration
        
        current_time = start_minutes
        lecture_index = 1
        
        while current_time < end_minutes:
            # Check if we're at lunch time
            if current_time == lunch_start_minutes:
                slots.append({
                    'start': TimeSlotManager.minutes_to_time(current_time),
                    'end': TimeSlotManager.minutes_to_time(lunch_end_minutes),
                    'type': 'lunch',
                    'index': None,
                    'duration': lunch_duration
                })
                current_time = lunch_end_minutes
                continue
            
            # Check if lunch falls within this lecture slot
            lecture_end = current_time + lecture_duration
            if current_time < lunch_start_minutes < lecture_end:
                # Lecture before lunch
                if current_time < lunch_start_minutes:
                    slots.append({
                        'start': TimeSlotManager.minutes_to_time(current_time),
                        'end': TimeSlotManager.minutes_to_time(lunch_start_minutes),
                        'type': 'lecture',
                        'index': lecture_index,
                        'duration': lunch_start_minutes - current_time
                    })
                    lecture_index += 1
                
                # Add lunch
                slots.append({
                    'start': TimeSlotManager.minutes_to_time(lunch_start_minutes),
                    'end': TimeSlotManager.minutes_to_time(lunch_end_minutes),
                    'type': 'lunch',
                    'index': None,
                    'duration': lunch_duration
                })
                current_time = lunch_end_minutes
                continue
            
            # Regular lecture slot
            if current_time + lecture_duration <= end_minutes:
                slots.append({
                    'start': TimeSlotManager.minutes_to_time(current_time),
                    'end': TimeSlotManager.minutes_to_time(current_time + lecture_duration),
                    'type': 'lecture',
                    'index': lecture_index,
                    'duration': lecture_duration
                })
                current_time += lecture_duration
                
                # Check if break after this lecture
                if lecture_index in break_after_lectures:
                    break_duration = break_after_lectures[lecture_index]
                    if current_time + break_duration <= end_minutes:
                        slots.append({
                            'start': TimeSlotManager.minutes_to_time(current_time),
                            'end': TimeSlotManager.minutes_to_time(current_time + break_duration),
                            'type': 'break',
                            'index': None,
                            'duration': break_duration
                        })
                        current_time += break_duration
                
                lecture_index += 1
            else:
                break
        
        return slots
    
    @staticmethod
    def generate_semester_slots(
        program: str,
        semester: int,
        lunch_config: dict = None,
        break_config: dict = None
    ) -> List[dict]:
        """
        CHANGE 1: Generate time slots for a specific semester
        Uses custom config if available, otherwise defaults
        """
        program_info = PROGRAM_CONFIG.get(program.upper(), {})
        
        # Get lunch configuration
        if lunch_config and lunch_config.get('custom', False):
            lunch_start = lunch_config.get('start', program_info.get('default_lunch_start', '13:00'))
            lunch_duration = lunch_config.get('duration', DEFAULT_LUNCH_DURATION)
        else:
            lunch_start = program_info.get('default_lunch_start', '13:00')
            lunch_duration = DEFAULT_LUNCH_DURATION
        
        # Get break configuration
        breaks = []
        if break_config and break_config.get('enabled', False):
            breaks = [{
                'duration': break_config.get('duration', DEFAULT_BREAK_DURATION),
                'placement': break_config.get('placements', [])
            }]
        
        return TimeSlotManager.generate_dynamic_slots(
            day_start=DEFAULT_DAY_START,
            day_end=DEFAULT_DAY_END,
            lunch_start=lunch_start,
            lunch_duration=lunch_duration,
            breaks=breaks
        )
    
    @staticmethod
    def get_slot_key(slot: dict) -> str:
        """Get unique key for a time slot"""
        return f"{slot['start']}-{slot['end']}"
    
    @staticmethod
    def get_lecture_slots_only(slots: List[dict]) -> List[dict]:
        """Filter to get only lecture slots"""
        return [s for s in slots if s['type'] == 'lecture']
    
    @staticmethod
    def compute_faculty_lunch_union(
        faculty_name: str,
        semester_lunch_configs: Dict[int, dict]
    ) -> List[Tuple[str, str]]:
        """
        CHANGE 4: Compute union of lunch intervals for faculty teaching multiple semesters
        Returns list of (start, end) tuples representing unavailable times
        """
        intervals = []
        
        for sem, config in semester_lunch_configs.items():
            if config:
                start = config.get('start', '13:00')
                duration = config.get('duration', 50)
                end_minutes = TimeSlotManager.time_to_minutes(start) + duration
                end = TimeSlotManager.minutes_to_time(end_minutes)
                intervals.append((start, end))
        
        if not intervals:
            return []
        
        # Sort intervals by start time
        intervals.sort(key=lambda x: TimeSlotManager.time_to_minutes(x[0]))
        
        # Merge overlapping intervals
        merged = [intervals[0]]
        for start, end in intervals[1:]:
            last_start, last_end = merged[-1]
            last_end_min = TimeSlotManager.time_to_minutes(last_end)
            start_min = TimeSlotManager.time_to_minutes(start)
            
            if start_min <= last_end_min:
                # Overlapping or adjacent - merge
                end_min = TimeSlotManager.time_to_minutes(end)
                new_end = TimeSlotManager.minutes_to_time(max(last_end_min, end_min))
                merged[-1] = (last_start, new_end)
            else:
                merged.append((start, end))
        
        return merged

def reset_dataset_session_state():
    """Nuclear reset of all session data to prevent cross-semester leakage."""
    keys_to_reset = [
        'info_dataset', 'room_dataset', 'info_df', 'room_df',
        'subjects', 'faculties', 'rooms', 'room_allocations', 
        'best_schedule', 'section_batch_schedules', 'current_schedule',
        'edited_schedule', 'detected_clashes', 'elective_groups',
        'current_semester_config', 'schools_data', 'sections'
    ]
    for key in keys_to_reset:
        if key in st.session_state:
            if key in ['subjects', 'faculties', 'rooms', 'info_dataset', 'room_dataset', 'detected_clashes', 'elective_groups']:
                st.session_state[key] = []
            elif key in ['room_allocations', 'section_batch_schedules', 'schools_data', 'sections']:
                st.session_state[key] = {}
            else:
                st.session_state[key] = None
    
    # Reset file uploader states by clearing their widget states if possible
    # (Though Streamlit doesn't allow direct widget state modification easily, 
    # clearing the parsed data forces a refresh)
    st.toast("♻️ Session state wiped and reset to clean state.")


# ==================== ENHANCED CSS ====================
st.markdown("""
<style>
    .main-header {
        font-size: 3rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: bold;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .firebase-status {
        position: fixed;
        top: 10px;
        right: 10px;
        padding: 5px 10px;
        border-radius: 5px;
        font-size: 0.8rem;
        z-index: 1000;
    }
    .firebase-connected {
        background: #d4edda;
        color: #155724;
        border: 1px solid #c3e6cb;
    }
    .firebase-disconnected {
        background: #f8d7da;
        color: #721c24;
        border: 1px solid #f5c6cb;
    }
    .portal-button {
        padding: 2rem;
        border-radius: 1rem;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        text-align: center;
        font-size: 1.5rem;
        font-weight: bold;
        cursor: pointer;
        transition: transform 0.3s;
        margin: 1rem;
    }
    .batch-popup {
        background: #f0f2f6;
        padding: 2rem;
        border-radius: 10px;
        border: 2px solid #667eea;
    }
    .edit-mode {
        background: #fff3cd;
        padding: 1rem;
        border-radius: 10px;
        border: 2px solid #ffc107;
        margin: 1rem 0;
    }
    .clash-detected {
        background: #f8d7da;
        padding: 1rem;
        border-radius: 10px;
        border: 2px solid #dc3545;
        margin: 1rem 0;
    }
    .no-clash {
        background: #d4edda;
        padding: 1rem;
        border-radius: 10px;
        border: 2px solid #28a745;
        margin: 1rem 0;
    }
    .lunch-break {
        background: #ffeaa7;
        text-align: center;
        font-weight: bold;
    }
    .break-slot {
        background: #dfe6e9;
        text-align: center;
        font-style: italic;
    }
    .genetic-progress {
        background: #e3f2fd;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
    .firebase-sync {
        background: #e8f5e9;
        padding: 0.5rem;
        border-radius: 5px;
        margin: 0.5rem 0;
        font-size: 0.9rem;
    }
    .tooltip-text {
        font-size: 0.85rem;
        color: #666;
        font-style: italic;
    }
    .column-info {
        background: #f8f9fa;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
        border-left: 3px solid #667eea;
    }
    .config-locked {
        background: #e8f5e9;
        border: 2px solid #28a745;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
    .config-unlocked {
        background: #fff3cd;
        border: 2px solid #ffc107;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
    .morning-limit-badge {
        background: #3498db;
        color: white;
        padding: 5px 10px;
        border-radius: 15px;
        font-size: 0.85rem;
    }
</style>
""", unsafe_allow_html=True)


# ==================== FIREBASE DATA MANAGER (UPDATED) ====================
# CHANGE 1, 2, 3, 4: Added new collections for semester configs, faculty constraints

class FirebaseManager:
    """Manage all Firebase database operations - Updated with dynamic config support"""
    
    def __init__(self, db):
        self.db = db
        # Updated collections including new ones for CHANGE 1, 2, 3, 4
        self.collections = {
            'timetables': 'timetables',
            'info_dataset': 'info_dataset',
            'room_dataset': 'room_dataset',
            'room_allocations': 'room_allocations',
            'batches': 'batches',
            'users': 'users',
            'logs': 'logs',
            'conflicts': 'conflicts',
            'archives': 'archives',
            # CHANGE 1: Semester lunch configurations
            'sem_lunch_configs': 'sem_lunch_configs',
            # CHANGE 2: Semester break configurations
            'sem_break_configs': 'sem_break_configs',
            # CHANGE 3: Faculty morning constraints tracking
            'faculty_constraints': 'faculty_constraints',
            # CHANGE 4: Faculty lunch unions
            'faculty_lunch_unions': 'faculty_lunch_unions',
            'agent_sessions': 'agent_sessions',
            'repair_history': 'repair_history',
            'agent_config': 'agent_config',
        }
    
    # ========== TIMETABLE OPERATIONS ==========
    def save_timetable(self, year: str, timetable_data: dict, batch_info: dict = None, 
                       semester_config: dict = None):
        """Save or update timetable in Firebase with semester config"""
        try:
            doc_ref = self.db.collection(self.collections['timetables']).document(year)
            
            firebase_data = {
                'year': year,
                'schedule': timetable_data,
                'created_at': firestore.SERVER_TIMESTAMP,
                'updated_at': firestore.SERVER_TIMESTAMP,
                'batch_info': batch_info or {},
                'semester_config': semester_config or {},  # CHANGE 1, 2: Store config
                'status': 'active',
                'clash_count': 0
            }
            
            if batch_info:
                if 'start_date' in batch_info and 'duration_days' in batch_info:
                    start_date = datetime.strptime(batch_info['start_date'], '%Y-%m-%d')
                    end_date = start_date + timedelta(days=batch_info['duration_days'])
                    firebase_data['batch_info']['end_date'] = end_date.strftime('%Y-%m-%d')
            
            doc_ref.set(firebase_data)
            self.log_operation('timetable_saved', {'year': year})
            
            return True, "Timetable successfully saved to database"
        except Exception as e:
            return False, f"Error saving timetable: {str(e)}"
    
    def load_timetable(self, year: str):
        """Load timetable from Firebase"""
        try:
            doc_ref = self.db.collection(self.collections['timetables']).document(year)
            doc = doc_ref.get()
            
            if doc.exists:
                return doc.to_dict()
            return None
        except Exception as e:
            st.error(f"Error loading timetable: {str(e)}")
            return None
    
    def delete_timetable(self, year: str, archive: bool = True):
        """Delete timetable from Firebase and all its associated data"""
        try:
            if archive:
                timetable = self.load_timetable(year)
                if timetable:
                    archive_ref = self.db.collection(self.collections['archives']).document(f"{year}_{int(time_module.time())}")
                    timetable['archived_at'] = firestore.SERVER_TIMESTAMP
                    archive_ref.set(timetable)
            
            # 1. Delete associated datasets (Info and Room)
            self.db.collection(self.collections['info_dataset']).document(year).delete()
            self.db.collection(self.collections['room_dataset']).document(f"{year}_rooms").delete()
            
            # 2. Delete configurations and allocations
            self.db.collection(self.collections['sem_lunch_configs']).document(f"{year}_lunch").delete()
            self.db.collection(self.collections['sem_break_configs']).document(f"{year}_break").delete()
            self.db.collection(self.collections['room_allocations']).document(f"{year}_allocations").delete()
            
            # 3. Delete constraints and the timetable itself
            self.db.collection(self.collections['faculty_constraints']).document(year).delete()
            doc_ref = self.db.collection(self.collections['timetables']).document(year)
            doc_ref.delete()
            
            self.log_operation('timetable_deleted', {'year': year, 'archived': archive, 'fully_wiped': True})
            return True, "Timetable and all associated configuration data permanently deleted"
        except Exception as e:
            return False, f"Error deleting timetable: {str(e)}"

    def wipe_entire_database(self):
        """
        DANGEROUS: Deletes all documents in all timetable-related collections.
        Iterates through collections and document references to wipe the DB.
        """
        try:
            # We exclude 'users' to keep admin accounts safe
            collections_to_wipe = [
                'timetables', 'info_dataset', 'room_dataset', 'room_allocations',
                'batches', 'conflicts', 'archives', 'sem_lunch_configs',
                'sem_break_configs', 'faculty_constraints', 'faculty_lunch_unions',
                'logs'
            ]
            
            total_deleted = 0
            for col_key in collections_to_wipe:
                col_name = self.collections.get(col_key)
                if not col_name:
                    continue
                
                # Fetch all documents in collection
                docs = self.db.collection(col_name).stream()
                batch = self.db.batch()
                count = 0
                
                for doc in docs:
                    batch.delete(doc.reference)
                    count += 1
                    total_deleted += 1
                    # Firebase batch limit is 500
                    if count >= 400:
                        batch.commit()
                        batch = self.db.batch()
                        count = 0
                
                if count > 0:
                    batch.commit()
            
            self.log_operation('database_wipe', {'total_documents_deleted': total_deleted})
            return True, f"Success: Database completely emptied ({total_deleted} documents deleted)."
        except Exception as e:
            return False, f"Critical Error during wipe: {str(e)}"
    
    def get_all_timetables(self):
        """Get all active timetables"""
        try:
            timetables = []
            docs = self.db.collection(self.collections['timetables']).where('status', '==', 'active').stream()
            
            for doc in docs:
                data = doc.to_dict()
                data['id'] = doc.id
                timetables.append(data)
            
            return timetables
        except Exception as e:
            st.error(f"Error fetching timetables: {str(e)}")
            return []
    
    # ========== CHANGE 1: SEMESTER LUNCH CONFIG OPERATIONS ==========
    def save_semester_lunch_config(self, program: str, semester: int, config: dict):
        """
        CHANGE 1: Save semester-specific lunch configuration
        Config: {custom: bool, start: "HH:MM", duration: int, end: "HH:MM", locked: bool}
        """
        try:
            doc_id = f"{program}_Sem{semester}_lunch"
            doc_ref = self.db.collection(self.collections['sem_lunch_configs']).document(doc_id)
            
            firebase_data = {
                'program': program,
                'semester': semester,
                'custom': config.get('custom', False),
                'start': config.get('start', '13:00'),
                'duration': config.get('duration', DEFAULT_LUNCH_DURATION),
                'end': config.get('end', '13:50'),
                'locked': config.get('locked', False),
                'updated_at': firestore.SERVER_TIMESTAMP
            }
            
            doc_ref.set(firebase_data)
            self.log_operation('lunch_config_saved', {'program': program, 'semester': semester})
            
            return True, f"Lunch config saved for {program} Semester {semester}"
        except Exception as e:
            return False, f"Error saving lunch config: {str(e)}"
    
    def get_semester_lunch_config(self, program: str, semester: int) -> Optional[dict]:
        """CHANGE 1: Get lunch configuration for a specific semester"""
        try:
            doc_id = f"{program}_Sem{semester}_lunch"
            doc = self.db.collection(self.collections['sem_lunch_configs']).document(doc_id).get()
            
            if doc.exists:
                return doc.to_dict()
            return None
        except Exception as e:
            return None
    
    def get_all_lunch_configs(self, program: str = None) -> List[dict]:
        """CHANGE 1: Get all lunch configurations, optionally filtered by program"""
        try:
            query = self.db.collection(self.collections['sem_lunch_configs'])
            
            if program:
                query = query.where('program', '==', program)
            
            configs = []
            docs = query.stream()
            
            for doc in docs:
                data = doc.to_dict()
                data['id'] = doc.id
                configs.append(data)
            
            return configs
        except Exception as e:
            return []
    
    # ========== CHANGE 2: SEMESTER BREAK CONFIG OPERATIONS ==========
    def save_semester_break_config(self, program: str, semester: int, config: dict):
        """
        CHANGE 2: Save semester-specific break configuration
        Config: {enabled: bool, duration: int, frequency: int, placements: [int]}
        """
        try:
            doc_id = f"{program}_Sem{semester}_break"
            doc_ref = self.db.collection(self.collections['sem_break_configs']).document(doc_id)
            
            firebase_data = {
                'program': program,
                'semester': semester,
                'enabled': config.get('enabled', False),
                'duration': config.get('duration', DEFAULT_BREAK_DURATION),
                'frequency': config.get('frequency', 1),
                'placements': config.get('placements', []),
                'updated_at': firestore.SERVER_TIMESTAMP
            }
            
            doc_ref.set(firebase_data)
            self.log_operation('break_config_saved', {'program': program, 'semester': semester})
            
            return True, f"Break config saved for {program} Semester {semester}"
        except Exception as e:
            return False, f"Error saving break config: {str(e)}"
    
    def get_semester_break_config(self, program: str, semester: int) -> Optional[dict]:
        """CHANGE 2: Get break configuration for a specific semester"""
        try:
            doc_id = f"{program}_Sem{semester}_break"
            doc = self.db.collection(self.collections['sem_break_configs']).document(doc_id).get()
            
            if doc.exists:
                return doc.to_dict()
            return None
        except Exception as e:
            return None
    
    def get_all_break_configs(self, program: str = None) -> List[dict]:
        """CHANGE 2: Get all break configurations"""
        try:
            query = self.db.collection(self.collections['sem_break_configs'])
            
            if program:
                query = query.where('program', '==', program)
            
            configs = []
            docs = query.stream()
            
            for doc in docs:
                data = doc.to_dict()
                data['id'] = doc.id
                configs.append(data)
            
            return configs
        except Exception as e:
            return []
    
    # ========== CHANGE 3: FACULTY MORNING CONSTRAINT OPERATIONS ==========
    def save_faculty_morning_counts(self, timetable_id: str, counts: dict):
        """
        CHANGE 3: Save faculty morning lecture counts after generation
        counts: {faculty_name: int (count of 9AM lectures)}
        """
        try:
            doc_ref = self.db.collection(self.collections['faculty_constraints']).document(timetable_id)
            
            firebase_data = {
                'timetable_id': timetable_id,
                'morning_counts': counts,
                'limit': FACULTY_MORNING_LIMIT,
                'updated_at': firestore.SERVER_TIMESTAMP
            }
            
            doc_ref.set(firebase_data)
            return True, "Faculty morning counts saved"
        except Exception as e:
            return False, str(e)
    
    def get_faculty_morning_counts(self, timetable_id: str = None) -> dict:
        """CHANGE 3: Get faculty morning counts for validation"""
        try:
            if timetable_id:
                doc = self.db.collection(self.collections['faculty_constraints']).document(timetable_id).get()
                if doc.exists:
                    return doc.to_dict().get('morning_counts', {})
                return {}
            else:
                # Aggregate across all timetables
                all_counts = defaultdict(int)
                docs = self.db.collection(self.collections['faculty_constraints']).stream()
                for doc in docs:
                    counts = doc.to_dict().get('morning_counts', {})
                    for faculty, count in counts.items():
                        all_counts[faculty] += count
                return dict(all_counts)
        except Exception as e:
            return {}
    
    # ========== CHANGE 4: FACULTY LUNCH UNION OPERATIONS ==========
    def save_faculty_lunch_unions(self, unions: dict):
        """
        CHANGE 4: Save computed faculty lunch unions
        unions: {faculty_name: [(start, end), ...]}
        """
        try:
            for faculty, intervals in unions.items():
                doc_id = faculty.replace(' ', '_').replace('.', '')
                doc_ref = self.db.collection(self.collections['faculty_lunch_unions']).document(doc_id)
                
                firebase_data = {
                    'faculty': faculty,
                    'unavailable_intervals': [{'start': s, 'end': e} for s, e in intervals],
                    'updated_at': firestore.SERVER_TIMESTAMP
                }
                
                doc_ref.set(firebase_data)
            
            return True, "Faculty lunch unions saved"
        except Exception as e:
            return False, str(e)
    
    def get_faculty_lunch_union(self, faculty_name: str) -> List[Tuple[str, str]]:
        """CHANGE 4: Get lunch union for a specific faculty"""
        try:
            doc_id = faculty_name.replace(' ', '_').replace('.', '')
            doc = self.db.collection(self.collections['faculty_lunch_unions']).document(doc_id).get()
            
            if doc.exists:
                intervals = doc.to_dict().get('unavailable_intervals', [])
                return [(i['start'], i['end']) for i in intervals]
            return []
        except Exception as e:
            return []
    
    def get_all_faculty_lunch_unions(self) -> dict:
        """CHANGE 4: Get all faculty lunch unions"""
        try:
            unions = {}
            docs = self.db.collection(self.collections['faculty_lunch_unions']).stream()
            
            for doc in docs:
                data = doc.to_dict()
                faculty = data.get('faculty', '')
                intervals = data.get('unavailable_intervals', [])
                unions[faculty] = [(i['start'], i['end']) for i in intervals]
            
            return unions
        except Exception as e:
            return {}
    
    # ========== INFO DATASET OPERATIONS ==========
    def save_info_dataset(self, program: str, semester: int, data: list):
        """Save Info Dataset to Firebase"""
        try:
            doc_id = f"{program}_Sem{semester}"
            doc_ref = self.db.collection(self.collections['info_dataset']).document(doc_id)
            
            firebase_data = {
                'program': program,
                'semester': semester,
                'data': data,
                'created_at': firestore.SERVER_TIMESTAMP,
                'updated_at': firestore.SERVER_TIMESTAMP,
                'record_count': len(data)
            }
            
            doc_ref.set(firebase_data)
            self.log_operation('info_dataset_saved', {'program': program, 'semester': semester})
            
            return True, f"Info Dataset saved for {program} Semester {semester}"
        except Exception as e:
            return False, f"Error saving Info Dataset: {str(e)}"
    
    def get_info_dataset(self, program: str = None, semester: int = None):
        """Get Info Dataset from Firebase"""
        try:
            if program and semester:
                doc_id = f"{program}_Sem{semester}"
                doc = self.db.collection(self.collections['info_dataset']).document(doc_id).get()
                if doc.exists:
                    return doc.to_dict()
                return None
            else:
                datasets = []
                docs = self.db.collection(self.collections['info_dataset']).stream()
                for doc in docs:
                    data = doc.to_dict()
                    data['id'] = doc.id
                    datasets.append(data)
                return datasets
        except Exception as e:
            st.error(f"Error fetching Info Dataset: {str(e)}")
            return [] if not (program and semester) else None
    
    def get_subjects_from_info_dataset(self, program: str = None, semester: int = None):
        """Extract subjects list from Info Dataset for timetable generation"""
        try:
            info_data = self.get_info_dataset(program, semester)
            subjects = []
            
            if info_data and 'data' in info_data:
                for record in info_data['data']:
                    # Theory subject
                    if record.get('Theory Hrs/Week', 0) > 0:
                        subjects.append({
                            'name': record['Module Name'],
                            'code': f"{record['Module Name'][:3].upper()}{record.get('S.No.', '')}",
                            'type': 'Theory',
                            'weekly_hours': record['Theory Hrs/Week'],
                            'duration': DEFAULT_LECTURE_DURATION,  # CHANGE 1: Use constant
                            'school': PROGRAM_CONFIG.get(record['Program'], {}).get('school', 'STME'),
                            'program': record['Program'],
                            'semester': record['Sem'],
                            'section': record['Section'],
                            'batch': record['Batch'],
                            'faculty': record.get('Faculty', 'TBD'),
                            'load': record.get('Theory Load', 0)
                        })
                    
                    # Lab/Practical subject
                    if record.get('Practical Hrs/Week', 0) > 0:
                        subjects.append({
                            'name': f"{record['Module Name']} Lab",
                            'code': f"{record['Module Name'][:3].upper()}{record.get('S.No.', '')}L",
                            'type': 'Lab',
                            'weekly_hours': record['Practical Hrs/Week'],
                            'duration': DEFAULT_LAB_DURATION,  # CHANGE 1: 2 hours for labs
                            'school': PROGRAM_CONFIG.get(record['Program'], {}).get('school', 'STME'),
                            'program': record['Program'],
                            'semester': record['Sem'],
                            'section': record['Section'],
                            'batch': record['Batch'],
                            'faculty': record.get('Faculty', 'TBD'),
                            'load': record.get('Practical Load', 0)
                        })
                    
                    # Tutorial subject
                    if record.get('Tutorial Hrs/Week', 0) > 0:
                        subjects.append({
                            'name': f"{record['Module Name']} Tutorial",
                            'code': f"{record['Module Name'][:3].upper()}{record.get('S.No.', '')}T",
                            'type': 'Tutorial',
                            'weekly_hours': record['Tutorial Hrs/Week'],
                            'duration': DEFAULT_LECTURE_DURATION,
                            'school': PROGRAM_CONFIG.get(record['Program'], {}).get('school', 'STME'),
                            'program': record['Program'],
                            'semester': record['Sem'],
                            'section': record['Section'],
                            'batch': record['Batch'],
                            'faculty': record.get('Faculty', 'TBD'),
                            'load': 0
                        })
            
            return subjects
        except Exception as e:
            st.error(f"Error extracting subjects: {str(e)}")
            return []
    
    def get_faculty_from_info_dataset(self):
        """Extract unique faculty list from all Info Datasets"""
        try:
            all_datasets = self.get_info_dataset()
            faculty_dict = {}
            
            for dataset in all_datasets:
                if 'data' in dataset:
                    for record in dataset['data']:
                        faculty_name = record.get('Faculty', '')
                        if faculty_name and faculty_name != 'TBD':
                            if faculty_name not in faculty_dict:
                                program = record.get('Program', '').upper()
                                school = PROGRAM_CONFIG.get(program, {}).get('school', 'General')
                                
                                faculty_dict[faculty_name] = {
                                    'name': faculty_name,
                                    'id': f"F{len(faculty_dict)+1:03d}",
                                    'department': school,
                                    'subjects': [],
                                    'max_hours': 20,
                                    'semesters': set()  # CHANGE 4: Track semesters
                                }
                            
                            # Add subject and semester
                            module_name = record.get('Module Name', '')
                            if module_name and module_name not in faculty_dict[faculty_name]['subjects']:
                                faculty_dict[faculty_name]['subjects'].append(module_name)
                            
                            sem = record.get('Sem', 1)
                            faculty_dict[faculty_name]['semesters'].add(sem)
            
            # Convert semesters set to list for JSON serialization
            result = []
            for faculty in faculty_dict.values():
                faculty['semesters'] = list(faculty['semesters'])
                result.append(faculty)
            
            return result
        except Exception as e:
            st.error(f"Error extracting faculty: {str(e)}")
            return []
    
    # ========== ROOM DATASET OPERATIONS ==========
    def save_room_dataset(self, program: str, semester: int, data: list):
        """Save Room Dataset to Firebase"""
        try:
            doc_id = f"{program}_Sem{semester}_rooms"
            doc_ref = self.db.collection(self.collections['room_dataset']).document(doc_id)
            
            firebase_data = {
                'program': program,
                'semester': semester,
                'data': data,
                'created_at': firestore.SERVER_TIMESTAMP,
                'updated_at': firestore.SERVER_TIMESTAMP,
                'record_count': len(data)
            }
            
            doc_ref.set(firebase_data)
            self.log_operation('room_dataset_saved', {'program': program, 'semester': semester})
            
            return True, f"Room Dataset saved for {program} Semester {semester}"
        except Exception as e:
            return False, f"Error saving Room Dataset: {str(e)}"
    
    def get_room_dataset(self, program: str = None, semester: int = None):
        """Get Room Dataset from Firebase"""
        try:
            if program and semester:
                doc_id = f"{program}_Sem{semester}_rooms"
                doc = self.db.collection(self.collections['room_dataset']).document(doc_id).get()
                if doc.exists:
                    return doc.to_dict()
                return None
            else:
                datasets = []
                docs = self.db.collection(self.collections['room_dataset']).stream()
                for doc in docs:
                    data = doc.to_dict()
                    data['id'] = doc.id
                    datasets.append(data)
                return datasets
        except Exception as e:
            st.error(f"Error fetching Room Dataset: {str(e)}")
            return [] if not (program and semester) else None
    
    def get_rooms_list(self):
        """Get unique rooms from all Room Datasets"""
        try:
            all_datasets = self.get_room_dataset()
            rooms_set = set()
            rooms_list = []
            
            for dataset in all_datasets:
                if 'data' in dataset:
                    for record in dataset['data']:
                        room_no = record.get('Room No.', '')
                        if room_no and room_no not in rooms_set:
                            rooms_set.add(room_no)
                            class_type = record.get('Class Type', 'theory').lower()
                            rooms_list.append({
                                'room_id': room_no,
                                'name': room_no,
                                'capacity': 60,
                                'building': 'Main',
                                'type': 'Lab' if class_type == 'lab' else 'Classroom',
                                'equipment': ['Projector', 'Whiteboard'] if class_type != 'lab' else ['Computers', 'Projector']
                            })
            
            return rooms_list
        except Exception as e:
            st.error(f"Error extracting rooms: {str(e)}")
            return []
    
    # ========== ROOM ALLOCATION OPERATIONS ==========
    def save_room_allocation(self, program: str, semester: int, allocations: dict):
        """Save room allocations to Firebase"""
        try:
            doc_id = f"{program}_Sem{semester}_allocations"
            doc_ref = self.db.collection(self.collections['room_allocations']).document(doc_id)
            
            firebase_data = {
                'program': program,
                'semester': semester,
                'allocations': allocations,
                'created_at': firestore.SERVER_TIMESTAMP,
                'updated_at': firestore.SERVER_TIMESTAMP
            }
            
            doc_ref.set(firebase_data)
            self.log_operation('room_allocation_saved', {'program': program, 'semester': semester})
            
            return True, "Room allocations saved"
        except Exception as e:
            return False, f"Error saving room allocations: {str(e)}"
    
    def get_room_allocation(self, program: str, semester: int):
        """Get room allocations from Firebase"""
        try:
            doc_id = f"{program}_Sem{semester}_allocations"
            doc = self.db.collection(self.collections['room_allocations']).document(doc_id).get()
            if doc.exists:
                return doc.to_dict()
            return None
        except Exception as e:
            return None
    
    # ========== BATCH OPERATIONS ==========
    def save_batch(self, batch_data: dict):
        """Save batch information"""
        try:
            batch_id = f"{batch_data['program']}_{batch_data['semester']}_{batch_data.get('section', 'A')}"
            doc_ref = self.db.collection(self.collections['batches']).document(batch_id)
            
            if 'start_date' in batch_data and 'duration_days' in batch_data:
                start = datetime.strptime(batch_data['start_date'], '%Y-%m-%d')
                end = start + timedelta(days=batch_data['duration_days'])
                batch_data['end_date'] = end.strftime('%Y-%m-%d')
            
            batch_data['updated_at'] = firestore.SERVER_TIMESTAMP
            doc_ref.set(batch_data, merge=True)
            
            return True, batch_id
        except Exception as e:
            return False, str(e)
    
    def get_batches(self, program: str = None):
        """Get batch information"""
        try:
            query = self.db.collection(self.collections['batches'])
            
            if program:
                query = query.where('program', '==', program)
            
            batches = []
            docs = query.stream()
            
            for doc in docs:
                data = doc.to_dict()
                data['id'] = doc.id
                batches.append(data)
            
            return batches
        except Exception as e:
            st.error(f"Error fetching batches: {str(e)}")
            return []
    
    # ========== CONFLICT/CLASH OPERATIONS ==========
    def save_clash(self, clash_data: dict):
        """Save detected clash for tracking"""
        try:
            clash_id = f"clash_{int(time_module.time())}_{random.randint(1000, 9999)}"
            doc_ref = self.db.collection(self.collections['conflicts']).document(clash_id)
            
            clash_data['detected_at'] = firestore.SERVER_TIMESTAMP
            clash_data['resolved'] = False
            doc_ref.set(clash_data)
            
            return True, clash_id
        except Exception as e:
            return False, str(e)
    
    def get_unresolved_clashes(self, year: str = None):
        """Get unresolved clashes"""
        try:
            query = self.db.collection(self.collections['conflicts']).where('resolved', '==', False)
            
            if year:
                query = query.where('year', '==', year)
            
            clashes = []
            docs = query.stream()
            
            for doc in docs:
                data = doc.to_dict()
                data['id'] = doc.id
                clashes.append(data)
            
            return clashes
        except Exception as e:
            return []
    
    # ========== LOGGING OPERATIONS ==========
    def log_operation(self, operation_type: str, details: dict):
        """Log operations for auditing"""
        try:
            log_entry = {
                'operation': operation_type,
                'details': details,
                'timestamp': firestore.SERVER_TIMESTAMP,
                'user': st.session_state.get('user_email', 'system')
            }
            
            self.db.collection(self.collections['logs']).add(log_entry)
        except Exception as e:
            print(f"Error logging operation: {str(e)}")
    
    # ========== USER OPERATIONS ==========
    def save_user(self, user_data: dict):
        """Save user information"""
        try:
            user_id = user_data.get('email', '').replace('@', '_').replace('.', '_')
            doc_ref = self.db.collection(self.collections['users']).document(user_id)
            
            user_data['updated_at'] = firestore.SERVER_TIMESTAMP
            doc_ref.set(user_data, merge=True)
            
            return True, user_id
        except Exception as e:
            return False, str(e)
    
    def get_user(self, email: str):
        """Get user information"""
        try:
            user_id = email.replace('@', '_').replace('.', '_')
            doc = self.db.collection(self.collections['users']).document(user_id).get()
            
            if doc.exists:
                return doc.to_dict()
            return None
        except Exception as e:
            return None
    
    # ========== AGENTIC AI OPERATIONS ==========
    def save_agent_session(self, session_id: str, session_data: dict):
        """Save agent repair session to /agent_sessions/{session_id}."""
        try:
            doc_ref = self.db.collection(self.collections['agent_sessions']).document(session_id)
            session_data['updated_at'] = firestore.SERVER_TIMESTAMP
            doc_ref.set(session_data)
            self.log_operation('agent_session_saved', {'session_id': session_id})
            return True, session_id
        except Exception as e:
            logger.exception("Failed to save agent session %s", session_id)
            return False, str(e)
        """Save a single repair action to /repair_history/{repair_id}."""
        try:
            doc_ref = self.db.collection(self.collections['repair_history']).document(repair_id)
            repair_data['timestamp_server'] = firestore.SERVER_TIMESTAMP
            doc_ref.set(repair_data)
            self.log_operation('repair_history_saved', {'repair_id': repair_id})
            return True, repair_id
        except Exception as e:
            logger.exception("Failed to save repair history %s", repair_id)
            return False, str(e)
        """Load agent configuration from /agent_config/{config_id}."""
        try:
            doc = self.db.collection(self.collections['agent_config']).document(config_id).get()
            if doc.exists:
                return doc.to_dict()
            return {
                'max_turns': 10,
                'llm_model': 'claude-sonnet-4-5',
                'enabled': True,
                'fallback_to_random_repair': True,
            }
        except Exception as e:
            return {
                'max_turns': 10,
                'llm_model': 'claude-sonnet-4-5',
                'enabled': True,
                'fallback_to_random_repair': True,
            }

    def save_agent_config(self, config_id: str, config_data: dict):
        """Save agent configuration to /agent_config/{config_id}."""
        try:
            doc_ref = self.db.collection(self.collections['agent_config']).document(config_id)
            config_data['updated_at'] = firestore.SERVER_TIMESTAMP
            doc_ref.set(config_data, merge=True)
            return True, config_id
        except Exception as e:
            return False, str(e)

    def get_repair_history(self, session_id: str = None, limit: int = 50):
        """Load repair history entries, optionally filtered by session."""
        try:
            query = self.db.collection(self.collections['repair_history'])
            if session_id:
                query = query.where('session_id', '==', session_id)
            docs = query.limit(limit).stream()
            history = []
            for doc in docs:
                data = doc.to_dict()
                data['id'] = doc.id
                history.append(data)
            return history
        except Exception as e:
            return []

    def get_agent_sessions(self, limit: int = 50):
        """Load agent repair sessions from /agent_sessions."""
        try:
            docs = (
                self.db.collection(self.collections['agent_sessions'])
                .limit(limit)
                .stream()
            )
            sessions = []
            for doc in docs:
                data = doc.to_dict()
                data['id'] = doc.id
                sessions.append(data)
            return sessions
        except Exception as e:
            return []

    def health_check(self) -> dict:
        """Lightweight connectivity check for monitoring and deployment."""
        try:
            list(self.db.collection(self.collections['logs']).limit(1).stream())
            return {"status": "ok", "firebase": "connected"}
        except Exception as exc:
            logger.exception("Firebase health check failed")
            return {"status": "degraded", "firebase": "error", "message": str(exc)}

    # ========== TIMETABLE OPTIMIZATION QUERIES ==========
    def get_all_faculty_schedules(self, exclude_timetable_key: str = None):
        """Get all faculty schedules from existing timetables for optimization.

        Args:
            exclude_timetable_key: Firestore document ID to skip (e.g. the timetable
                being regenerated right now). Prevents the old copy of the SAME
                semester from blocking all its own slots.
        """
        try:
            faculty_schedules = defaultdict(lambda: defaultdict(list))
            timetables = self.get_all_timetables()
            
            for timetable in timetables:
                # Skip the timetable we are currently regenerating so it does
                # not falsely block every slot that was in the previous run.
                tt_id = timetable.get('id', '')
                if exclude_timetable_key and tt_id == exclude_timetable_key:
                    continue

                schedule = timetable.get('schedule', {})
                for school in schedule:
                    for batch in schedule[school]:
                        for day in schedule[school][batch]:
                            for slot, slot_content in schedule[school][batch][day].items():
                                entries = slot_content if isinstance(slot_content, list) else [slot_content]
                                for class_info in entries:
                                    if (class_info and
                                            class_info.get('faculty') and
                                            class_info.get('faculty') not in ('TBD', '') and
                                            class_info.get('type') not in ['LUNCH', 'BREAK']):
                                        faculty_name = class_info['faculty']
                                        faculty_schedules[faculty_name][f"{day}_{slot}"].append({
                                            'school': school,
                                            'batch': batch,
                                            'subject': class_info.get('subject'),
                                            'timetable_id': tt_id
                                        })
            
            return dict(faculty_schedules)
        except Exception as e:
            st.error(f"Error fetching faculty schedules: {str(e)}")
            return {}
    
    def get_all_room_schedules(self, exclude_timetable_key: str = None):
        """Get all room schedules from existing timetables for optimization."""
        try:
            room_schedules = defaultdict(lambda: defaultdict(list))
            timetables = self.get_all_timetables()
            
            for timetable in timetables:
                tt_id = timetable.get('id', '')
                if exclude_timetable_key and tt_id == exclude_timetable_key:
                    continue
                schedule = timetable.get('schedule', {})
                for school in schedule:
                    for batch in schedule[school]:
                        for day in schedule[school][batch]:
                            for slot, slot_content in schedule[school][batch][day].items():
                                entries = slot_content if isinstance(slot_content, list) else [slot_content]
                                for class_info in entries:
                                    if class_info and class_info.get('room') and class_info.get('type') not in ['LUNCH', 'BREAK']:
                                        room_name = class_info['room']
                                        if room_name not in ['TBD', 'Cafeteria', '']:
                                            room_schedules[room_name][f"{day}_{slot}"].append({
                                                'school': school,
                                                'batch': batch,
                                                'subject': class_info.get('subject'),
                                                'timetable_id': tt_id
                                            })
            
            return dict(room_schedules)
        except Exception as e:
            st.error(f"Error fetching room schedules: {str(e)}")
            return {}


# Initialize Firebase Manager
if db:
    firebase_manager = FirebaseManager(db)
else:
    firebase_manager = None 

# app.py - Part 2: Sidebar Configuration, Algorithms, Core Classes
# Continuation from Part 1

# ==================== CHANGE 1, 2: SIDEBAR CONFIGURATION COMPONENT ====================

def render_semester_config_sidebar(firebase_mgr, selected_program: str, selected_semester: int):
    """
    CHANGE 1, 2: Render sidebar configuration for lunch and breaks per semester
    """
    if not firebase_mgr or not selected_program:
        return
    
    program_info = PROGRAM_CONFIG.get(selected_program.upper(), {})
    max_semesters = program_info.get('semesters', 8)
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⏰ Semester Time Configuration")
    
    # Initialize session state for configs if not exists
    config_key = f"sem_configs_{selected_program}"
    if config_key not in st.session_state:
        st.session_state[config_key] = {}
    
    # CHANGE 1: Per-Semester Lunch Configuration
    with st.sidebar.expander(f"🍴 Lunch Configuration - Sem {selected_semester}", expanded=True):
        # Load existing config from Firebase
        existing_lunch = firebase_mgr.get_semester_lunch_config(selected_program, selected_semester)
        
        # Check if locked
        is_locked = existing_lunch.get('locked', False) if existing_lunch else False
        
        # Custom lunch checkbox
        lunch_key = f"lunch_custom_{selected_program}_{selected_semester}"
        default_custom = existing_lunch.get('custom', False) if existing_lunch else False
        
        use_custom_lunch = st.checkbox(
            f"Set Custom Lunch for Sem {selected_semester}",
            value=default_custom,
            key=lunch_key,
            disabled=is_locked
        )
        
        if use_custom_lunch:
            # Get default values
            default_start = existing_lunch.get('start', program_info.get('default_lunch_start', '13:00')) if existing_lunch else program_info.get('default_lunch_start', '13:00')
            default_duration = existing_lunch.get('duration', DEFAULT_LUNCH_DURATION) if existing_lunch else DEFAULT_LUNCH_DURATION
            
            # Parse default start time
            try:
                start_parts = default_start.split(':')
                default_start_time = time(int(start_parts[0]), int(start_parts[1]))
            except (ValueError, TypeError, IndexError):
                default_start_time = time(13, 0)
            
            col1, col2 = st.columns(2)
            
            with col1:
                lunch_start = st.time_input(
                    "Start Time",
                    value=default_start_time,
                    key=f"lunch_start_{selected_program}_{selected_semester}",
                    disabled=is_locked
                )
            
            with col2:
                lunch_duration = st.number_input(
                    "Duration (min)",
                    min_value=30,
                    max_value=90,
                    value=default_duration,
                    step=5,
                    key=f"lunch_duration_{selected_program}_{selected_semester}",
                    disabled=is_locked
                )
            
            # Compute end time
            lunch_start_str = lunch_start.strftime("%H:%M")
            lunch_end_minutes = TimeSlotManager.time_to_minutes(lunch_start_str) + lunch_duration
            lunch_end_str = TimeSlotManager.minutes_to_time(lunch_end_minutes)
            
            st.info(f"🍴 Lunch: {TimeSlotManager.format_time_12hr(lunch_start_str)} - {TimeSlotManager.format_time_12hr(lunch_end_str)} ({lunch_duration} min)")
            
            # Lock/Save button
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("💾 Save & Lock", key=f"save_lunch_{selected_program}_{selected_semester}", disabled=is_locked):
                    lunch_config = {
                        'custom': True,
                        'start': lunch_start_str,
                        'duration': lunch_duration,
                        'end': lunch_end_str,
                        'locked': True
                    }
                    success, msg = firebase_mgr.save_semester_lunch_config(selected_program, selected_semester, lunch_config)
                    if success:
                        st.success("✅ Lunch config saved & locked!")
                        st.rerun()
                    else:
                        st.error(f"❌ {msg}")
            
            with col2:
                if is_locked:
                    if st.button("🔓 Unlock", key=f"unlock_lunch_{selected_program}_{selected_semester}"):
                        lunch_config = existing_lunch.copy()
                        lunch_config['locked'] = False
                        firebase_mgr.save_semester_lunch_config(selected_program, selected_semester, lunch_config)
                        st.success("🔓 Unlocked!")
                        st.rerun()
            
            if is_locked:
                st.markdown('<div class="config-locked">🔒 Configuration is locked</div>', unsafe_allow_html=True)
        else:
            # Using default lunch
            default_start = program_info.get('default_lunch_start', '13:00')
            default_end_minutes = TimeSlotManager.time_to_minutes(default_start) + DEFAULT_LUNCH_DURATION
            default_end = TimeSlotManager.minutes_to_time(default_end_minutes)
            
            st.info(f"Using default: {TimeSlotManager.format_time_12hr(default_start)} - {TimeSlotManager.format_time_12hr(default_end)} ({DEFAULT_LUNCH_DURATION} min)")
            
            # Save default config
            if st.button("💾 Save Default", key=f"save_default_lunch_{selected_program}_{selected_semester}"):
                lunch_config = {
                    'custom': False,
                    'start': default_start,
                    'duration': DEFAULT_LUNCH_DURATION,
                    'end': default_end,
                    'locked': False
                }
                firebase_mgr.save_semester_lunch_config(selected_program, selected_semester, lunch_config)
                st.success("✅ Default lunch config saved!")
    
    # CHANGE 2: Per-Semester Break Configuration
    with st.sidebar.expander(f"☕ Break Configuration - Sem {selected_semester}", expanded=False):
        # Load existing config
        existing_break = firebase_mgr.get_semester_break_config(selected_program, selected_semester)
        
        break_key = f"break_enabled_{selected_program}_{selected_semester}"
        default_enabled = existing_break.get('enabled', False) if existing_break else False
        
        enable_breaks = st.checkbox(
            f"Add Custom Breaks for Sem {selected_semester}",
            value=default_enabled,
            key=break_key
        )
        
        if enable_breaks:
            # Get defaults
            default_duration = existing_break.get('duration', DEFAULT_BREAK_DURATION) if existing_break else DEFAULT_BREAK_DURATION
            default_frequency = existing_break.get('frequency', 1) if existing_break else 1
            default_placements = existing_break.get('placements', [4]) if existing_break else [4]
            
            break_duration = st.number_input(
                "Break Duration (min)",
                min_value=5,
                max_value=30,
                value=default_duration,
                step=5,
                key=f"break_duration_{selected_program}_{selected_semester}"
            )
            
            break_frequency = st.number_input(
                "Breaks per Day",
                min_value=1,
                max_value=3,
                value=default_frequency,
                key=f"break_freq_{selected_program}_{selected_semester}"
            )
            
            st.markdown("**Place break after lecture #:**")
            
            placements = []
            cols = st.columns(min(break_frequency, 3))
            for i in range(break_frequency):
                with cols[i % 3]:
                    default_val = default_placements[i] if i < len(default_placements) else 4
                    placement = st.number_input(
                        f"Break {i+1}",
                        min_value=1,
                        max_value=8,
                        value=default_val,
                        key=f"break_place_{selected_program}_{selected_semester}_{i}"
                    )
                    placements.append(placement)
            
            st.info(f"☕ {break_frequency} break(s) of {break_duration} min after lecture(s): {placements}")
            
            if st.button("💾 Save Breaks", key=f"save_break_{selected_program}_{selected_semester}"):
                break_config = {
                    'enabled': True,
                    'duration': break_duration,
                    'frequency': break_frequency,
                    'placements': placements
                }
                success, msg = firebase_mgr.save_semester_break_config(selected_program, selected_semester, break_config)
                if success:
                    st.success("✅ Break config saved!")
                else:
                    st.error(f"❌ {msg}")
        else:
            st.info("No additional breaks configured")
            
            # Clear break config if disabled
            if existing_break and existing_break.get('enabled', False):
                if st.button("Clear Break Config", key=f"clear_break_{selected_program}_{selected_semester}"):
                    break_config = {'enabled': False, 'duration': 0, 'frequency': 0, 'placements': []}
                    firebase_mgr.save_semester_break_config(selected_program, selected_semester, break_config)
                    st.success("Break config cleared!")
                    st.rerun()
    
    # CHANGE NEW-SECTION-6: Mentor-Mentee Sessions Configuration
    with st.sidebar.expander("👨‍🎓 Mentor-Mentee Sessions", expanded=False):
        mm_key = f"mm_enabled_{selected_program}_{selected_semester}"
        
        # Load existing config from Firebase (stored inside sem_break_configs)
        existing_break_cfg = firebase_mgr.get_semester_break_config(selected_program, selected_semester)
        existing_mm = {}
        if existing_break_cfg and isinstance(existing_break_cfg, dict):
            existing_mm = existing_break_cfg.get('mentor_mentee', {})
        
        mm_enabled = st.checkbox(
            "Enable Mentor-Mentee Sessions",
            value=existing_mm.get('enabled', False),
            key=mm_key
        )
        
        if mm_enabled:
            col_mm1, col_mm2 = st.columns(2)
            with col_mm1:
                mm_sessions = st.number_input(
                    "Sessions/Week",
                    min_value=1, max_value=5,
                    value=int(existing_mm.get('sessions_per_week', 1)),
                    key=f"mm_sessions_{selected_program}_{selected_semester}"
                )
            with col_mm2:
                mm_duration = st.number_input(
                    "Duration (min)",
                    min_value=15, max_value=120, step=5,
                    value=int(existing_mm.get('duration_min', 60)),
                    key=f"mm_duration_{selected_program}_{selected_semester}"
                )
            
            # Section-batches aware config
            # CHANGE NEW-SECTION-6 fix: selected_school is not in scope here; read from session_state
            _school = st.session_state.get('selected_school', '')
            existing_sections = st.session_state.get('sections', {}).get(
                f"{selected_program}_{_school}", {}
            ).get(selected_semester, ['A'])
            
            mm_batch_wise = False
            section_batches_exist = False
            for sec in existing_sections:
                sec_key = f"{selected_program}_{_school}"
                section_batches = st.session_state.get('schools_data', {}).get(
                    sec_key, {}
                ).get('section_batches', {}).get(selected_semester, {}).get(sec, [])
                if section_batches and len(section_batches) > 1:
                    section_batches_exist = True
                    break
            
            if section_batches_exist:
                mm_batch_wise = st.checkbox(
                    "Assign batch-wise (different mentors per batch)",
                    value=existing_mm.get('batch_wise', False),
                    key=f"mm_batchwise_{selected_program}_{selected_semester}"
                )
            
            st.info(f"🎓 {mm_sessions} session(s)/week × {mm_duration} min")
            
            if st.button("💾 Save Mentor-Mentee Config",
                         key=f"save_mm_{selected_program}_{selected_semester}"):
                # Save inside existing sem_break_configs doc as a sub-key
                base_break = firebase_mgr.get_semester_break_config(
                    selected_program, selected_semester
                ) or {}
                base_break['mentor_mentee'] = {
                    'enabled': True,
                    'sessions_per_week': mm_sessions,
                    'duration_min': mm_duration,
                    'batch_wise': mm_batch_wise
                }
                success, msg = firebase_mgr.save_semester_break_config(
                    selected_program, selected_semester, base_break
                )
                if success:
                    st.success("✅ Mentor-Mentee config saved!")
                else:
                    st.error(f"❌ {msg}")
        else:
            if existing_mm.get('enabled', False):
                if st.button("Clear Mentor-Mentee Config",
                             key=f"clear_mm_{selected_program}_{selected_semester}"):
                    base_break = firebase_mgr.get_semester_break_config(
                        selected_program, selected_semester
                    ) or {}
                    base_break['mentor_mentee'] = {'enabled': False}
                    firebase_mgr.save_semester_break_config(
                        selected_program, selected_semester, base_break
                    )
                    st.success("Mentor-Mentee config cleared!")
                    st.rerun()
    
    # CHANGE 3: Display morning limit info

    st.sidebar.markdown("---")
    st.sidebar.markdown(f'<span class="morning-limit-badge">🌅 Morning Limit: {FACULTY_MORNING_LIMIT} max @ 9AM/faculty/week</span>', unsafe_allow_html=True)
    
    # CHANGE 4: Faculty Lunch Preview (for multi-semester faculty)
    with st.sidebar.expander("👨‍🏫 Faculty Lunch Preview", expanded=False):
        st.markdown("Shows lunch union for faculty teaching multiple semesters")
        
        # Get faculty list
        faculties = firebase_mgr.get_faculty_from_info_dataset()
        
        if faculties:
            # Filter faculty teaching in selected program
            program_faculty = [f for f in faculties if any(
                s in f.get('semesters', []) for s in range(1, max_semesters + 1)
            )]
            
            if program_faculty:
                selected_faculty = st.selectbox(
                    "Select Faculty",
                    [f['name'] for f in program_faculty],
                    key=f"faculty_preview_{selected_program}"
                )
                
                if selected_faculty:
                    # Find faculty info
                    faculty_info = next((f for f in program_faculty if f['name'] == selected_faculty), None)
                    
                    if faculty_info:
                        semesters_taught = faculty_info.get('semesters', [])
                        st.write(f"**Teaches in Sem:** {sorted(semesters_taught)}")
                        
                        # Get lunch configs for each semester
                        sem_lunch_configs = {}
                        for sem in semesters_taught:
                            config = firebase_mgr.get_semester_lunch_config(selected_program, sem)
                            if config:
                                sem_lunch_configs[sem] = config
                            else:
                                # Use default
                                default_start = program_info.get('default_lunch_start', '13:00')
                                sem_lunch_configs[sem] = {
                                    'start': default_start,
                                    'duration': DEFAULT_LUNCH_DURATION
                                }
                        
                        # Compute union
                        if len(sem_lunch_configs) > 1:
                            union = TimeSlotManager.compute_faculty_lunch_union(selected_faculty, sem_lunch_configs)
                            
                            if len(union) > 1:
                                st.warning("⚠️ Faculty has multiple lunch intervals!")
                            
                            st.write("**Unavailable times:**")
                            for start, end in union:
                                st.write(f"  • {TimeSlotManager.format_time_12hr(start)} - {TimeSlotManager.format_time_12hr(end)}")
                        else:
                            st.success("✅ Single lunch interval")
            else:
                st.info("No faculty data available")
        else:
            st.info("Upload Info Dataset to view faculty")


# ==================== ENHANCED ALGORITHMS ====================

class HungarianAlgorithm:
    """Hungarian Algorithm for optimal teacher-course assignment"""
    
    def __init__(self):
        self.cost_matrix = None
        self.assignments = None
    
    def create_cost_matrix(self, faculties: List[dict], courses: List[dict], 
                           morning_counts: dict = None) -> np.ndarray:
        """
        Create cost matrix for assignment problem
        CHANGE 3: Added morning limit consideration
        """
        n_faculties = len(faculties)
        n_courses = len(courses)
        morning_counts = morning_counts or {}
        
        cost_matrix = np.ones((n_faculties, n_courses)) * 1000
        
        for i, faculty in enumerate(faculties):
            faculty_subjects = faculty.get('subjects', [])
            max_hours = faculty.get('max_hours', 20)
            current_load = faculty.get('current_load', 0)
            faculty_name = faculty.get('name', '')
            
            # CHANGE 3: Get current morning count
            current_morning = morning_counts.get(faculty_name, 0)
            
            for j, course in enumerate(courses):
                course_name = course.get('name', '')
                course_hours = course.get('weekly_hours', 3)
                
                can_teach = any(subj.lower() in course_name.lower() for subj in faculty_subjects)
                
                if can_teach and (current_load + course_hours <= max_hours):
                    cost = 10
                    
                    if course.get('preferred_faculty') == faculty['name']:
                        cost -= 5
                    
                    workload_ratio = current_load / max_hours
                    cost += workload_ratio * 5
                    
                    # CHANGE 3: Add penalty if faculty already at morning limit
                    if current_morning >= FACULTY_MORNING_LIMIT:
                        cost += 50  # High penalty for morning slots
                    
                    cost_matrix[i][j] = max(0, cost)
        
        return cost_matrix
    
    def solve(self, faculties: List[dict], courses: List[dict], 
              morning_counts: dict = None) -> Dict[str, str]:
        """Solve assignment problem using Hungarian algorithm"""
        self.cost_matrix = self.create_cost_matrix(faculties, courses, morning_counts)
        
        row_indices, col_indices = linear_sum_assignment(self.cost_matrix)
        
        assignments = {}
        for row, col in zip(row_indices, col_indices):
            if row < len(faculties) and col < len(courses):
                if self.cost_matrix[row][col] < 1000:
                    assignments[courses[col]['name']] = faculties[row]['name']
        
        self.assignments = assignments
        return assignments


class GraphColoringAlgorithm:
    """Graph Coloring for conflict-free slot allocation"""
    
    def __init__(self):
        self.graph = nx.Graph()
        self.colors = {}
    
    def build_conflict_graph(self, classes: List[dict], 
                             faculty_lunch_unions: dict = None) -> nx.Graph:
        """
        Build conflict graph where edges represent conflicts
        CHANGE 4: Consider faculty lunch unions
        """
        self.graph.clear()
        faculty_lunch_unions = faculty_lunch_unions or {}
        
        for i, class_info in enumerate(classes):
            self.graph.add_node(i, **class_info)
        
        for i in range(len(classes)):
            for j in range(i + 1, len(classes)):
                if self._has_conflict(classes[i], classes[j], faculty_lunch_unions):
                    self.graph.add_edge(i, j)
        
        return self.graph
    
    def _has_conflict(self, class1: dict, class2: dict, 
                      faculty_lunch_unions: dict = None) -> bool:
        """Check if two classes have a conflict"""
        # Faculty conflict
        if class1.get('faculty') == class2.get('faculty'):
            return True
        
        # Same batch conflict
        if (class1.get('batch') == class2.get('batch') and 
            class1.get('school') == class2.get('school')):
            
            # Allow parallel labs/tutorials if they belong to different real batches
            type1 = str(class1.get('type', '')).upper()
            type2 = str(class2.get('type', '')).upper()
            # CHANGE PARALLEL-LAB-FIX-3: Fixed type2 definition and ensured parallel batches run DIFFERENT subjects
            if ('LAB' in type1 or 'TUTORIAL' in type1 or 'PRACTICAL' in type1) and ('LAB' in type2 or 'TUTORIAL' in type2 or 'PRACTICAL' in type2):
                if class1.get('real_batch') and class2.get('real_batch') and class1.get('real_batch') != class2.get('real_batch'):
                    # Conflict if they are trying to teach the same subject at the same time
                    if str(class1.get('subject', '')).strip().upper() == str(class2.get('subject', '')).strip().upper():
                        return True # Conflict!
                    return False  # No conflict! (They can run parallel in different rooms/faculties acting on different subjects)
            return True
        
        # Room conflict
        if (class1.get('room') and class2.get('room') and 
            class1['room'] == class2['room']):
            return True
        
        return False
    
    def color_graph(self, classes: List[dict], available_slots: List[tuple]) -> Dict[int, tuple]:
        """Assign time slots (colors) to classes"""
        self.build_conflict_graph(classes)
        
        coloring = nx.greedy_color(self.graph, strategy='largest_first')
        
        slot_assignments = {}
        for node, color in coloring.items():
            if color < len(available_slots):
                slot_assignments[node] = available_slots[color]
            else:
                slot_assignments[node] = available_slots[color % len(available_slots)]
        
        self.colors = slot_assignments
        return slot_assignments


# CHANGE 1: Room Allocation Logic (Updated)
class RoomAllocator:
    """Handle room allocation before timetable generation"""
    
    def __init__(self, firebase_manager=None):
        self.firebase = firebase_manager
        self.allocations = {}
    
    def allocate_rooms(self, program: str, semester: int):
        """
        Allocate rooms to subjects for a given program/semester
        Uses Room Dataset mapping and greedy algorithm for auto-allocation
        """
        if not self.firebase:
            return {}, "Firebase not connected"
        
        # Get Info Dataset subjects
        info_data = self.firebase.get_info_dataset(program, semester)
        if not info_data or 'data' not in info_data:
            return {}, f"No Info Dataset found for {program} Semester {semester}"
        
        # Get Room Dataset
        room_data = self.firebase.get_room_dataset(program, semester)
        
        # CHANGE PARALLEL-LAB-FIX-3: Normalize subject names for Room Dataset lookup
        def _normalize_subj(name: str) -> str:
            n = name.strip()
            for suffix in (' Lab', ' Tutorial', ' Practical'):
                if n.lower().endswith(suffix.lower()):
                    n = n[: -len(suffix)].strip()
                    break
            return n

        room_mapping = {}
        if room_data and 'data' in room_data:
            for record in room_data['data']:
                subject = record.get('Subject', '')
                class_type = record.get('Class Type', 'theory').lower()
                room_no = record.get('Room No.', '')
                section = str(record.get('Section', '')).strip().upper()

                if subject and room_no:
                    # Store under the NORMALIZED (suffix-stripped) key
                    norm_base = f"{_normalize_subj(subject)}_{class_type}"
                    orig_base = f"{subject}_{class_type}"
                    
                    norm_key = f"{norm_base}_{section}" if section else norm_base
                    orig_key = f"{orig_base}_{section}" if section else orig_base
                    
                    if norm_key not in room_mapping:
                        room_mapping[norm_key] = []
                    if room_no not in room_mapping[norm_key]:
                        room_mapping[norm_key].append(room_no)
                    
                    # Also store under the original key as a fallback
                    if orig_key != norm_key:
                        if orig_key not in room_mapping:
                            room_mapping[orig_key] = []
                        if room_no not in room_mapping[orig_key]:
                            room_mapping[orig_key].append(room_no)
        
        # Get all available rooms
        all_rooms = self.firebase.get_rooms_list()
        theory_rooms = [r['name'] for r in all_rooms if r.get('type') != 'Lab']
        lab_rooms = [r['name'] for r in all_rooms if r.get('type') == 'Lab']
        
        if not theory_rooms:
            theory_rooms = [f"Classroom-{i}" for i in range(1, 11)]
        if not lab_rooms:
            lab_rooms = [f"Lab-{i}" for i in range(1, 6)]
        
        # Track room usage for greedy allocation
        room_usage_count = defaultdict(int)
        
        allocations = {}
        for k, v in room_mapping.items():
            allocations[k] = list(v)
            
        unallocated = []
        
        for record in info_data['data']:
            module_name = record.get('Module Name', '')
            stripped_name = _normalize_subj(module_name)
            section = str(record.get('Section', '')).strip().upper()
            
            def _is_allocated(suffix):
                keys = [
                    f"{stripped_name}_{suffix}_{section}" if section else None,
                    f"{module_name}_{suffix}_{section}" if section else None,
                    f"{stripped_name}_{suffix}",
                    f"{module_name}_{suffix}"
                ]
                return any(k in allocations for k in filter(None, keys))
            
            # Theory allocation for unmapped subjects
            if record.get('Theory Hrs/Week', 0) > 0 and not _is_allocated("theory"):
                key_to_set = f"{module_name}_theory_{section}" if section else f"{module_name}_theory"
                available = sorted(theory_rooms, key=lambda r: room_usage_count[r])
                if available:
                    allocations.setdefault(key_to_set, []).append(available[0])
                    room_usage_count[available[0]] += 1
                else:
                    unallocated.append(key_to_set)
            
            # Lab allocation for unmapped subjects
            if record.get('Practical Hrs/Week', 0) > 0 and not _is_allocated("lab"):
                key_to_set = f"{module_name}_lab_{section}" if section else f"{module_name}_lab"
                available = sorted(lab_rooms, key=lambda r: room_usage_count[r])
                if available:
                    allocations.setdefault(key_to_set, []).append(available[0])
                    room_usage_count[available[0]] += 1
                else:
                    unallocated.append(key_to_set)
            
            # Tutorial allocation for unmapped subjects
            if record.get('Tutorial Hrs/Week', 0) > 0 and not _is_allocated("tutorial"):
                key_to_set = f"{module_name}_tutorial_{section}" if section else f"{module_name}_tutorial"
                available = sorted(theory_rooms, key=lambda r: room_usage_count[r])
                if available:
                    allocations.setdefault(key_to_set, []).append(available[0])
                    room_usage_count[available[0]] += 1
                else:
                    unallocated.append(key_to_set)
        
        # Save allocations to Firebase
        self.firebase.save_room_allocation(program, semester, allocations)
        
        self.allocations = allocations
        
        if unallocated:
            return allocations, f"Allocated {len(allocations)} subjects. {len(unallocated)} could not be allocated."
        
        return allocations, f"Successfully allocated rooms for {len(allocations)} subject-types"
    
    def get_room_for_subject(self, subject_name: str, class_type: str):
        """Get allocated room for a subject"""
        key = f"{subject_name}_{class_type.lower()}"
        return self.allocations.get(key, 'TBD')
    
    def validate_room_dataset(self, room_df: pd.DataFrame, info_data: dict) -> tuple:
        """Validate Room Dataset against Info Dataset"""
        errors = []
        warnings = []
        
        if not info_data or 'data' not in info_data:
            return ["No Info Dataset to validate against"], []
        
        info_subjects = set()
        for record in info_data['data']:
            info_subjects.add(record.get('Module Name', ''))
        
        for _, row in room_df.iterrows():
            subject = row.get('Subject', '')
            if subject and subject not in info_subjects:
                warnings.append(f"Subject '{subject}' in Room Dataset not found in Info Dataset")
        
        return errors, warnings


# ==================== CLASH DETECTION SYSTEM ====================

class ClashDetector:
    """Detect and analyze scheduling clashes with Firebase integration"""
    
    def __init__(self, firebase_manager=None):
        self.clashes = []
        self.clash_count = 0
        self.firebase = firebase_manager
        
    def detect_all_clashes(self, schedule, save_to_firebase=False,
                           existing_faculty_schedules=None,
                           existing_room_schedules=None):
        """Detect all types of clashes in the schedule.

        Args:
            schedule: The newly generated schedule dict.
            save_to_firebase: Whether to persist detected clashes.
            existing_faculty_schedules: Loaded from Firebase — other semesters'
                faculty bookings. Used to flag cross-semester clashes.
            existing_room_schedules: Loaded from Firebase — other semesters'
                room bookings.
        """
        self.clashes = []
        self.clash_count = 0
        
        faculty_clashes = self.detect_faculty_clashes(schedule)
        self.clashes.extend(faculty_clashes)
        
        room_clashes = self.detect_room_clashes(schedule)
        self.clashes.extend(room_clashes)

        # Cross-semester faculty clashes
        if existing_faculty_schedules:
            cs_clashes = self._detect_cross_semester_faculty_clashes(
                schedule, existing_faculty_schedules)
            self.clashes.extend(cs_clashes)

        # Cross-semester room clashes
        if existing_room_schedules:
            cs_room_clashes = self._detect_cross_semester_room_clashes(
                schedule, existing_room_schedules)
            self.clashes.extend(cs_room_clashes)
        
        self.clash_count = len(self.clashes)
        
        if save_to_firebase and self.firebase:
            for clash in self.clashes:
                self.firebase.save_clash(clash)
        
        return self.clashes

    def _detect_cross_semester_faculty_clashes(self, schedule, existing_faculty_schedules):
        """Detect cases where a faculty is placed in a slot already used by another semester."""
        clashes = []
        for school in schedule:
            for batch in schedule[school]:
                for day in schedule[school][batch]:
                    for slot, slot_val in schedule[school][batch][day].items():
                        _items = slot_val if isinstance(slot_val, list) else [slot_val]
                        for ci in _items:
                            if not isinstance(ci, dict):
                                continue
                            if ci.get('type') in ('LUNCH', 'BREAK'):
                                continue
                            fac = ci.get('faculty', '')
                            if not fac or fac in ('TBD', ''):
                                continue
                            key = f"{day}_{slot}"
                            if fac in existing_faculty_schedules and key in existing_faculty_schedules[fac]:
                                clashes.append({
                                    'type': 'Cross-Semester Faculty Clash',
                                    'severity': 'Critical',
                                    'faculty': fac,
                                    'time': key.replace('_', ' at '),
                                    'details': (f"{fac} is already teaching another semester's "
                                                f"class at {key.replace('_', ' ')}"),
                                    'locations': [{'school': school, 'batch': batch,
                                                   'subject': ci.get('subject', ''),
                                                   'other_semester': existing_faculty_schedules[fac][key]}]
                                })
        return clashes

    def _detect_cross_semester_room_clashes(self, schedule, existing_room_schedules):
        """Detect cases where a room is used in a slot already booked by another semester."""
        clashes = []
        for school in schedule:
            for batch in schedule[school]:
                for day in schedule[school][batch]:
                    for slot, slot_val in schedule[school][batch][day].items():
                        _items = slot_val if isinstance(slot_val, list) else [slot_val]
                        for ci in _items:
                            if not isinstance(ci, dict):
                                continue
                            if ci.get('type') in ('LUNCH', 'BREAK'):
                                continue
                            room = ci.get('room', '')
                            if not room or room in ('TBD', 'Cafeteria', ''):
                                continue
                            key = f"{day}_{slot}"
                            if room in existing_room_schedules and key in existing_room_schedules[room]:
                                clashes.append({
                                    'type': 'Cross-Semester Room Clash',
                                    'severity': 'Critical',
                                    'room': room,
                                    'time': key.replace('_', ' at '),
                                    'details': (f"Room {room} is already booked by another semester "
                                                f"at {key.replace('_', ' ')}"),
                                    'locations': [{'school': school, 'batch': batch,
                                                   'subject': ci.get('subject', '')}]
                                })
        return clashes
    
    def detect_faculty_clashes(self, schedule):
        """Detect faculty scheduling conflicts"""
        faculty_schedule = defaultdict(lambda: defaultdict(list))
        clashes = []
        
        for school in schedule:
            for batch in schedule[school]:
                for day in schedule[school][batch]:
                    for slot, slot_val in schedule[school][batch][day].items():
                        _slot_items = slot_val if isinstance(slot_val, list) else [slot_val]
                        for class_info in _slot_items:
                            if (class_info and isinstance(class_info, dict)
                                    and 'faculty' in class_info
                                    and class_info.get('type') not in ['LUNCH', 'BREAK']):
                                faculty_name = class_info['faculty']
                                key = f"{day}_{slot}"
                                faculty_schedule[faculty_name][key].append({
                                    'school': school,
                                    'batch': batch,
                                    'subject': class_info.get('subject', 'Unknown'),
                                    'room': class_info.get('room', 'TBD')
                                })
        
        for faculty, slots in faculty_schedule.items():
            for slot_key, assignments in slots.items():
                if len(assignments) > 1:
                    clashes.append({
                        'type': 'Faculty Clash',
                        'severity': 'High',
                        'faculty': faculty,
                        'time': slot_key.replace('_', ' at '),
                        'details': f"{faculty} assigned to {len(assignments)} classes simultaneously",
                        'locations': assignments
                    })
        
        return clashes
    
    def detect_room_clashes(self, schedule):
        """Detect room booking conflicts"""
        room_schedule = defaultdict(lambda: defaultdict(list))
        clashes = []
        
        for school in schedule:
            for batch in schedule[school]:
                for day in schedule[school][batch]:
                    for slot, slot_val in schedule[school][batch][day].items():
                        _slot_items = slot_val if isinstance(slot_val, list) else [slot_val]
                        for class_info in _slot_items:
                            if (class_info and isinstance(class_info, dict)
                                    and 'room' in class_info
                                    and class_info['room'] not in ['TBD', 'Cafeteria', '']
                                    and class_info.get('type') not in ['LUNCH', 'BREAK']):
                                room_name = class_info['room']
                                key = f"{day}_{slot}"
                                room_schedule[room_name][key].append({
                                    'school': school,
                                    'batch': batch,
                                    'subject': class_info.get('subject', 'Unknown'),
                                    'faculty': class_info.get('faculty', 'TBD')
                                })
        
        for room, slots in room_schedule.items():
            for slot_key, bookings in slots.items():
                if len(bookings) > 1:
                    clashes.append({
                        'type': 'Room Clash',
                        'severity': 'High',
                        'room': room,
                        'time': slot_key.replace('_', ' at '),
                        'details': f"{room} booked for {len(bookings)} classes simultaneously",
                        'bookings': bookings
                    })
        
        return clashes


# ==================== TIMETABLE EDITOR ====================

class TimetableEditor:
    """Handle timetable editing with Firebase integration"""
    
    def __init__(self, firebase_manager=None):
        self.edit_history = deque(maxlen=20)
        self.original_schedule = None
        self.clash_detector = ClashDetector(firebase_manager)
        self.firebase = firebase_manager
        
    def enable_edit_mode(self, schedule):
        """Create editable copy of schedule"""
        self.original_schedule = copy.deepcopy(schedule)
        return copy.deepcopy(schedule)
    
    def swap_slots(self, schedule, school, batch, day1, slot1, day2, slot2):
        """Swap two time slots with clash checking"""
        self.edit_history.append(('swap', copy.deepcopy(schedule)))
        temp = schedule[school][batch][day1][slot1]
        schedule[school][batch][day1][slot1] = schedule[school][batch][day2][slot2]
        schedule[school][batch][day2][slot2] = temp
        
        return schedule, True, "Slots swapped successfully"
    
    def update_class_info(self, schedule, school, batch, day, slot, new_info):
        """Update class information"""
        self.edit_history.append(('update', copy.deepcopy(schedule)))
        schedule[school][batch][day][slot] = new_info
        return schedule, True, "Class updated successfully"
    
    def remove_class(self, schedule, school, batch, day, slot):
        """Remove a class from schedule"""
        self.edit_history.append(('remove', copy.deepcopy(schedule)))
        schedule[school][batch][day][slot] = None
        return schedule, True, "Class removed"
    
    def add_class(self, schedule, school, batch, day, slot, class_info):
        """Add a new class to schedule"""
        self.edit_history.append(('add', copy.deepcopy(schedule)))
        schedule[school][batch][day][slot] = class_info
        return schedule, True, "Class added successfully"
    
    def undo_last_change(self, schedule):
        """Undo the last change"""
        if self.edit_history:
            action, previous_state = self.edit_history.pop()
            return previous_state
        return schedule
    
    def reset_to_original(self):
        """Reset to original schedule"""
        return copy.deepcopy(self.original_schedule)
    
    def validate_changes(self, schedule):
        """Validate the edited schedule"""
        clashes = self.clash_detector.detect_all_clashes(schedule)
        return clashes
    
    def save_to_firebase(self, schedule, year, batch_info=None, semester_config=None):
        """Save edited schedule to Firebase"""
        if self.firebase:
            success, msg = self.firebase.save_timetable(year, schedule, batch_info, semester_config)
            return success, msg
        return False, "Firebase not connected"


# ==================== DATASET UPLOAD MANAGER ====================

class DatasetUploadManager:
    """Handle bulk dataset uploads with Firebase integration - Updated structure"""
    
    def __init__(self, firebase_manager=None):
        self.firebase = firebase_manager
    
    # CHANGE NEW-SECTION-5: Universal column name normalization – accept any common variant
    # Maps alternative/variant column names (case-insensitive) to the canonical name used internally.
    COLUMN_ALIASES = {
        # S.No. variants
        's.no': 'S.No.',
        's. no.': 'S.No.',
        's.no.': 'S.No.',
        'serial no': 'S.No.',
        'sr.no': 'S.No.',
        'sr. no.': 'S.No.',
        'sno': 'S.No.',
        # Program variants
        'programme': 'Program',
        # Sem variants
        'semester': 'Sem',
        'sem.': 'Sem',
        # Section variants
        'sec': 'Section',
        'sec.': 'Section',
        'sect': 'Section',
        'sect.': 'Section',
        'section no': 'Section',
        'section no.': 'Section',
        # Batch variants
        'batch no': 'Batch',
        'batch no.': 'Batch',
        # Module Name variants  (professor's Excel uses "Name of the Module" and similar)
        'name of the module': 'Module Name',
        'module': 'Module Name',
        'subject': 'Module Name',
        'subject name': 'Module Name',
        'course name': 'Module Name',
        'course': 'Module Name',
        # Theory Hrs/Week variants
        'theory hrs/week': 'Theory Hrs/Week',
        'theory hours per week': 'Theory Hrs/Week',
        'theory hours/week': 'Theory Hrs/Week',
        'theory hrs per week': 'Theory Hrs/Week',
        'theory hrs': 'Theory Hrs/Week',
        'theory': 'Theory Hrs/Week',
        'theory hours': 'Theory Hrs/Week',
        # Practical Hrs/Week variants
        'practical hrs/week': 'Practical Hrs/Week',
        'practicals hours per week': 'Practical Hrs/Week',
        'practical hours per week': 'Practical Hrs/Week',
        'practical hrs per week': 'Practical Hrs/Week',
        'practicals hrs/week': 'Practical Hrs/Week',
        'practical hours/week': 'Practical Hrs/Week',
        'practicals hrs per week': 'Practical Hrs/Week',
        'practical hrs': 'Practical Hrs/Week',
        'practicals': 'Practical Hrs/Week',
        'lab hrs/week': 'Practical Hrs/Week',
        'lab hours': 'Practical Hrs/Week',
        'practicals hours': 'Practical Hrs/Week',
        # Tutorial Hrs/Week variants
        'tutorial hrs/week': 'Tutorial Hrs/Week',
        'tutorials hours per week': 'Tutorial Hrs/Week',
        'tutorial hours per week': 'Tutorial Hrs/Week',
        'tutorial hrs per week': 'Tutorial Hrs/Week',
        'tutorials hrs/week': 'Tutorial Hrs/Week',
        'tutorial hours/week': 'Tutorial Hrs/Week',
        'tutorials hrs per week': 'Tutorial Hrs/Week',
        'tutorial hrs': 'Tutorial Hrs/Week',
        'tutorials hours': 'Tutorial Hrs/Week',
        # Theory Load variants  (professor's Excel uses "Theory(L)" and similar)
        'theory load': 'Theory Load',
        'theory(l)': 'Theory Load',
        'theory (l)': 'Theory Load',
        'theoryl': 'Theory Load',
        'theory load (l)': 'Theory Load',
        'theory l': 'Theory Load',
        'lecture load': 'Theory Load',
        # Practical Load variants
        'practical load': 'Practical Load',
        'practicals load': 'Practical Load',
        'practical (l)': 'Practical Load',
        'lab load': 'Practical Load',
        'pl': 'Practical Load',
        # Total Load variants
        'total load': 'Total Load',
        'total': 'Total Load',
        'total hrs': 'Total Load',
        'total hours': 'Total Load',
        'total (l)': 'Total Load',
        # Faculty variants
        'faculty name': 'Faculty',
        'faculty member': 'Faculty',
        'teacher': 'Faculty',
        'instructor': 'Faculty',
        'professor': 'Faculty',
    }

    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        CHANGE NEW-SECTION-5: Normalize df column names to canonical names.
        Uses a 3-stage approach so virtually any column name variant is handled:
          Stage 1 – exact COLUMN_ALIASES lookup (case-insensitive, stripped)
          Stage 2 – all punctuation/spaces removed, then alias lookup
          Stage 3 – keyword fuzzy fallback (checks which keywords are present)
        """
        import re

        def _strip_punct(s):
            """Remove all non-alphanumeric characters and lowercase."""
            return re.sub(r'[^a-z0-9]', '', s.lower())

        # Build a stripped-key version of the alias map for Stage 2
        stripped_aliases = {
            _strip_punct(k): v
            for k, v in DatasetUploadManager.COLUMN_ALIASES.items()
        }

        # Stage 3: keyword-fuzzy rules
        # Each rule: (must_contain_all, must_NOT_contain, canonical_name)
        # Evaluated in order; first match wins.
        FUZZY_RULES = [
            # Theory Hrs/Week  — "theory" + any hour/week indicator, but no "load"
            (['theory'], ['load', 'l)'],                  'Theory Hrs/Week'),
            # Practical Hrs/Week — "pract"/"lab" + hour/week, no "load"
            (['pract'],  ['load', 'l)'],                  'Practical Hrs/Week'),
            (['lab'],    ['load', 'l)'],                  'Practical Hrs/Week'),
            # Tutorial Hrs/Week — "tutor" + hour/week, no "load"
            (['tutor'],  ['load', 'l)'],                  'Tutorial Hrs/Week'),
            # Theory Load — "theory" + "load" OR "theory" + "(l)"
            (['theory', 'load'], [],                      'Theory Load'),
            (['theory', '(l)'],  [],                      'Theory Load'),
            (['theoryl'],        [],                      'Theory Load'),
            # Practical Load — "pract"/"lab" + "load" OR "(l)"
            (['pract', 'load'],  [],                      'Practical Load'),
            (['lab', 'load'],    [],                      'Practical Load'),
            (['pract', '(l)'],   [],                      'Practical Load'),
            # Total Load — "total"
            (['total'],          [],                      'Total Load'),
            # Module Name
            (['module'],         [],                      'Module Name'),
            (['name', 'module'], [],                      'Module Name'),
            # Faculty
            (['faculty'],        [],                      'Faculty'),
            (['teacher'],        [],                      'Faculty'),
            (['instructor'],     [],                      'Faculty'),
        ]

        rename_map = {}
        already_mapped_to = set()  # avoid mapping two cols to the same canonical name

        for col in df.columns:
            col_stripped = str(col).strip()
            col_lower = col_stripped.lower()
            col_nopunct = _strip_punct(col_lower)

            canonical = None

            # Stage 1: exact alias
            if col_lower in DatasetUploadManager.COLUMN_ALIASES:
                canonical = DatasetUploadManager.COLUMN_ALIASES[col_lower]

            # Stage 2: stripped alias
            if canonical is None and col_nopunct in stripped_aliases:
                canonical = stripped_aliases[col_nopunct]

            # Stage 3: keyword fuzzy — only if not already found
            if canonical is None:
                for must_have, must_not, target in FUZZY_RULES:
                    if (all(kw in col_lower for kw in must_have) and
                            not any(kw in col_lower for kw in must_not)):
                        canonical = target
                        break

            if canonical and canonical not in already_mapped_to:
                rename_map[col] = canonical
                already_mapped_to.add(canonical)

        if rename_map:
            df = df.rename(columns=rename_map)
        return df

    def parse_info_dataset(self, df: pd.DataFrame) -> tuple:
        """
        Parse and validate Info Dataset.
        CHANGE NEW-SECTION-5: Column names are normalized (case-insensitive, alias-aware)
        before validation. Only column NAME existence is checked — no strict value-format
        or data-type validation is enforced.
        Expected canonical columns: S.No., Program, Sem, Section, Batch, Module Name,
        Theory Hrs/Week, Practical Hrs/Week, Tutorial Hrs/Week,
        Theory Load, Practical Load, Total Load, Faculty
        """
        errors = []
        warnings = []
        records = []

        # CHANGE NEW-SECTION-5: Normalize column names before any checks
        df = self._normalize_columns(df)

        # CHANGE NEW-SECTION-5: Only check that required column NAMES exist (case-insensitive).
        # Accept any value format — do not enforce data types here.
        required_columns = [
            'S.No.', 'Program', 'Sem', 'Section', 'Batch', 'Module Name',
            'Theory Hrs/Week', 'Practical Hrs/Week', 'Tutorial Hrs/Week',
            'Theory Load', 'Practical Load', 'Total Load', 'Faculty'
        ]

        # Case-insensitive column presence check
        df_cols_lower = {c.strip().lower(): c for c in df.columns}
        missing_cols = []
        col_map = {}
        for req in required_columns:
            if req in df.columns:
                col_map[req] = req  # exact match
            elif req.lower() in df_cols_lower:
                col_map[req] = df_cols_lower[req.lower()]  # case-insensitive match
            else:
                missing_cols.append(req)

        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return [], errors, warnings

        # Helper: safely get a column value using the col_map
        def get_val(row, canonical_col):
            actual_col = col_map.get(canonical_col, canonical_col)
            return row[actual_col] if actual_col in row.index else None

        # CHANGE NEW-SECTION-5: Parse each row with lenient type coercion; no value validation
        def safe_int(val, default=0):
            try:
                if pd.isna(val):
                    return default
                # Accept roman numerals I,II,III etc. → map to int
                roman_map = {'i': 1, 'ii': 2, 'iii': 3, 'iv': 4, 'v': 5,
                             'vi': 6, 'vii': 7, 'viii': 8, 'ix': 9, 'x': 10}
                s = str(val).strip().lower()
                if s in roman_map:
                    return roman_map[s]
                return int(float(s))
            except Exception:
                return default

        def safe_float(val, default=0.0):
            try:
                if pd.isna(val):
                    return default
                return float(str(val).strip())
            except Exception:
                return default

        def safe_str(val, default=''):
            try:
                if pd.isna(val):
                    return default
                return str(val).strip()
            except Exception:
                return default

        for idx, row in df.iterrows():
            try:
                module_name = safe_str(get_val(row, 'Module Name'))
                if not module_name:
                    errors.append(f"Row {idx+1}: Module Name is required")
                    continue

                # Normalise Batch field: strip trailing .0 from float-like strings
                _raw_batch = get_val(row, 'Batch')
                _batch_str = safe_str(_raw_batch, '1')
                if _batch_str.endswith('.0'):
                    _batch_str = _batch_str[:-2]

                # Normalise Section field: convert numeric sections (1→A, 2→B …)
                _raw_section = safe_str(get_val(row, 'Section'), 'A').strip().upper()
                if _raw_section.isdigit():
                    _raw_section = chr(64 + int(_raw_section))

                record = {
                    'S.No.': safe_int(get_val(row, 'S.No.'), idx + 1),
                    'Program': safe_str(get_val(row, 'Program'), ''),
                    'Sem': safe_int(get_val(row, 'Sem'), 1),
                    'Section': _raw_section,
                    'Batch': _batch_str,
                    'Module Name': module_name,
                    'Theory Hrs/Week': safe_float(get_val(row, 'Theory Hrs/Week'), 0.0),
                    'Practical Hrs/Week': safe_float(get_val(row, 'Practical Hrs/Week'), 0.0),
                    'Tutorial Hrs/Week': safe_float(get_val(row, 'Tutorial Hrs/Week'), 0.0),
                    'Theory Load': safe_float(get_val(row, 'Theory Load'), 0.0),
                    'Practical Load': safe_float(get_val(row, 'Practical Load'), 0.0),
                    'Total Load': safe_float(get_val(row, 'Total Load'), 0.0),
                    'Faculty': safe_str(get_val(row, 'Faculty'), 'TBD') or 'TBD',
                }

                records.append(record)

            except Exception as e:
                errors.append(f"Row {idx+1}: Error parsing - {str(e)}")

        return records, errors, warnings
    
    def save_info_dataset_to_firebase(self, records: list, program: str, semester: int):
        """Save parsed Info Dataset to Firebase"""
        if self.firebase:
            success, msg = self.firebase.save_info_dataset(program, semester, records)
            return success, msg
        return False, "Firebase not connected"
    
    def parse_room_dataset(self, df: pd.DataFrame) -> tuple:
        """
        Parse and validate Room Dataset
        Expected columns: Subject, Class Type, Room No.
        """
        errors = []
        warnings = []
        records = []
        
        required_columns = ['Subject', 'Class Type', 'Room No.']
        
        # Check for required columns (Section is optional)
        missing_cols = [col for col in required_columns if col not in df.columns]
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return [], errors, warnings
        
        valid_class_types = ['theory', 'lab', 'tutorial', 'practical']
        
        for idx, row in df.iterrows():
            try:
                class_type = str(row['Class Type']).strip().lower() if pd.notna(row['Class Type']) else 'theory'
                
                record = {
                    'Subject': str(row['Subject']).strip() if pd.notna(row['Subject']) else '',
                    'Class Type': class_type,
                    'Room No.': str(row['Room No.']).strip() if pd.notna(row['Room No.']) else '',
                    'Section': str(row['Section']).strip().upper() if 'Section' in df.columns and pd.notna(row['Section']) else ''
                }
                
                # Validate class type
                if class_type not in valid_class_types:
                    warnings.append(f"Row {idx+1}: Invalid class type '{class_type}'. Using 'theory'")
                    record['Class Type'] = 'theory'
                
                # Validate subject
                if not record['Subject']:
                    errors.append(f"Row {idx+1}: Subject is required")
                    continue
                
                # Validate room
                if not record['Room No.']:
                    warnings.append(f"Row {idx+1}: Room No. is empty for subject '{record['Subject']}'")
                
                records.append(record)
                
            except Exception as e:
                errors.append(f"Row {idx+1}: Error parsing - {str(e)}")
        
        return records, errors, warnings
    
    def validate_room_against_info(self, room_records: list, info_records: list) -> tuple:
        """Validate Room Dataset subjects against Info Dataset"""
        errors = []
        warnings = []
        
        # Get all module names from Info Dataset
        info_modules = set()
        for record in info_records:
            info_modules.add(record.get('Module Name', ''))
        
        # Check each room record
        for record in room_records:
            subject = record.get('Subject', '')
            if subject and subject not in info_modules:
                warnings.append(f"Subject '{subject}' in Room Dataset not found in Info Dataset")
        
        return errors, warnings
    
    def save_room_dataset_to_firebase(self, records: list, program: str, semester: int):
        """Save parsed Room Dataset to Firebase"""
        if self.firebase:
            success, msg = self.firebase.save_room_dataset(program, semester, records)
            return success, msg
        return False, "Firebase not connected"
    
    def convert_info_to_subjects(self, info_records: list) -> list:
        """Convert Info Dataset records to subjects list for scheduling"""
        import streamlit as st
        import re
        subjects = []
        skipped_prog_sem = 0
        
        # # CHANGE TIMETABLE-FIX-4: Prevent Data Leakage
        # Determine the current program/semester to STRICTLY filter out other programs (like AIDS/MBA Tech)
        current_prog_key = st.session_state.get('selected_program', '').strip().upper()
        current_sem = str(st.session_state.get('selected_semester', '')).strip()
        
        # Get full program name if available to allow matching against it too
        prog_display_name = ""
        if current_prog_key in PROGRAM_CONFIG:
            prog_display_name = str(PROGRAM_CONFIG[current_prog_key].get('name', '')).strip().upper()

        def _clean(s):
            return re.sub(r'[^A-Z0-9]', '', str(s).upper())

        for record in info_records:
            record_prog = str(record.get('Program', '')).strip().upper()
            sem_val = str(record.get('Sem', '')).strip()
            
            # Roman to Numeric normalization for comparison if needed
            roman_map = {'i': '1', 'ii': '2', 'iii': '3', 'iv': '4', 'v': '5', 'vi': '6', 'vii': '7', 'viii': '8', 'ix': '9', 'x': '10'}
            if sem_val.lower() in roman_map:
                sem_val = roman_map[sem_val.lower()]
            
            # STRICT FILTER: Exclude other programs and semesters completely to prevent leakage
            # Robust matching: match against key (BTECH) or display name (Bachelor of Technology)
            if current_prog_key:
                clean_record_prog = _clean(record_prog)
                clean_current_prog = _clean(current_prog_key)
                clean_display_name = _clean(prog_display_name)
                
                # If record says e.g. "B.Tech" it matches BTECH. If it says "BBA" it won't match.
                if clean_record_prog != clean_current_prog and clean_record_prog != clean_display_name:
                    skipped_prog_sem += 1
                    continue
                    
            if current_sem and sem_val and sem_val != current_sem:
                skipped_prog_sem += 1
                continue
            
            # STRICT FILTER: Check sections & batches if configured
            sec_val = str(record.get('Section', 'A')).strip().upper()
            batch_val = str(record.get('Batch', '1')).strip()
            if batch_val.endswith('.0'):
                batch_val = batch_val[:-2]

            try:
                num_batches = int(batch_val)
            except ValueError:
                num_batches = 1

            school = PROGRAM_CONFIG.get(current_prog_key, {}).get('school', 'STME') 
            
            # Theory subject (combined for entire section)
            if record.get('Theory Hrs/Week', 0) > 0:
                subjects.append({
                    'name': record['Module Name'],
                    'code': f"{record['Module Name'][:3].upper()}{record.get('S.No.', '')}",
                    'type': 'Theory',
                    'weekly_hours': int(record['Theory Hrs/Week']),
                    'duration': DEFAULT_LECTURE_DURATION,
                    'school': school,
                    'program': current_prog_key,
                    'year': sem_val,
                    'semester': sem_val,
                    'section': sec_val,
                    'batch': '',  # No batch for theory (section-wide)
                    'faculty': record.get('Faculty', 'TBD'),
                    'load': record.get('Theory Load', 0)
                })
            
            # Lab/Practical subject (one per batch)
            if record.get('Practical Hrs/Week', 0) > 0:
                mod_name = record['Module Name']
                if not any(mod_name.lower().endswith(s) for s in [' lab', ' practical']):
                    name = f"{mod_name} Lab"
                else:
                    name = mod_name
                
                for b_idx in range(1, num_batches + 1):
                    # Format as 01, 02 etc for consistency with rules
                    b_str = f"0{b_idx}" if b_idx < 10 else str(b_idx)
                    subjects.append({
                        'name': name,
                        'code': f"{mod_name[:3].upper()}{record.get('S.No.', '')}L",
                        'type': 'Lab',
                        'weekly_hours': int(record['Practical Hrs/Week']),
                        'duration': DEFAULT_LAB_DURATION,
                        'school': school,
                        'program': current_prog_key,
                        'year': sem_val,
                        'semester': sem_val,
                        'section': sec_val,
                        'batch': b_str,
                        'faculty': record.get('Faculty', 'TBD'),
                        'load': record.get('Practical Load', 0)
                    })
            
            # Tutorial subject (one per batch)
            if record.get('Tutorial Hrs/Week', 0) > 0:
                mod_name = record['Module Name']
                if not mod_name.lower().endswith(' tutorial'):
                    name = f"{mod_name} Tutorial"
                else:
                    name = mod_name
 
                for b_idx in range(1, num_batches + 1):
                    # Format as 01, 02 etc for consistency with rules
                    b_str = f"0{b_idx}" if b_idx < 10 else str(b_idx)
                    subjects.append({
                        'name': name,
                        'code': f"{mod_name[:3].upper()}{record.get('S.No.', '')}T",
                        'type': 'Tutorial',
                        'weekly_hours': int(record['Tutorial Hrs/Week']),
                        'duration': DEFAULT_LECTURE_DURATION,
                        'school': school,
                        'program': current_prog_key,
                        'year': sem_val,
                        'semester': sem_val,
                        'section': sec_val,
                        'batch': b_str,
                        'faculty': record.get('Faculty', 'TBD'),
                        'load': 0
                    })
        
        if skipped_prog_sem > 0:
            st.info(f"ℹ️ Filtered dataset: {len(subjects)} subjects kept. ({skipped_prog_sem} subjects for other Programs/Semesters were skipped to prevent interference).")
        
        # ── MULTI-SECTION SUPPORT ─────────────────────────────────────────────
        # If multiple sections (A, B, C …) are configured but the info_dataset
        # only has rows for one section, we replicate every subject entry for
        # every missing section so the scheduler builds a timetable for each.
        try:
            _sel_school = st.session_state.get('selected_school', '')
            # Fallback: derive school from program config when session key is missing
            if not _sel_school:
                _sel_school = PROGRAM_CONFIG.get(current_prog_key, {}).get('school', '')
            _prog_key   = f"{_sel_school}_{current_prog_key}"
            _sem_int    = int(current_sem) if current_sem else 1

            _sections_store = st.session_state.get('sections', {})
            # Primary lookup; also try without the school prefix as a fallback
            _configured_sections = (
                _sections_store.get(_prog_key, {}).get(_sem_int, []) or
                next(
                    (v.get(_sem_int, []) for k, v in _sections_store.items()
                     if current_prog_key in k),
                    []
                )
            )

            if len(_configured_sections) > 1:
                _sections_in_subjects = set(
                    str(s.get('section', '')).strip().upper()
                    for s in subjects if s.get('section')
                )
                _missing_sections = [
                    sec for sec in _configured_sections
                    if str(sec).upper() not in _sections_in_subjects
                ]

                if _missing_sections:
                    import copy as _copy_ms
                    _extra = []
                    _source_section = next(iter(_sections_in_subjects),
                                           str(_configured_sections[0]).upper())

                    for _ms in _missing_sections:
                        for _subj in subjects:
                            if str(_subj.get('section', '')).strip().upper() != _source_section:
                                continue
                            _copy_subj = _copy_ms.deepcopy(_subj)
                            _copy_subj['section'] = str(_ms).upper()
                            _extra.append(_copy_subj)

                    subjects = subjects + _extra
                    st.info(
                        f"ℹ️ Multi-section: subjects replicated for "
                        f"Section(s) {_missing_sections} "
                        f"(total {len(subjects)} subject entries)."
                    )
        except Exception:
            pass  # Non-fatal; original subjects list is returned unchanged
        # ── END MULTI-SECTION SUPPORT ─────────────────────────────────────────

        return subjects
    
    def extract_faculty_from_info(self, info_records: list) -> list:
        """Extract unique faculty list from Info Dataset with strict filtering"""
        import streamlit as st
        import re
        faculty_dict = {}
        
        # # CHANGE TIMETABLE-FIX-4: Consistent Filtering for Faculty
        current_prog_key = st.session_state.get('selected_program', '').strip().upper()
        current_sem = str(st.session_state.get('selected_semester', '')).strip()
        prog_display_name = ""
        if current_prog_key in PROGRAM_CONFIG:
            prog_display_name = str(PROGRAM_CONFIG[current_prog_key].get('name', '')).strip().upper()

        def _clean(s):
            return re.sub(r'[^A-Z0-9]', '', str(s).upper())

        for record in info_records:
            # APPLY SAME STRICT FILTER
            record_prog = str(record.get('Program', '')).strip().upper()
            sem_val = str(record.get('Sem', '')).strip()
            
            if current_prog_key:
                clean_record_prog = _clean(record_prog)
                clean_current_prog = _clean(current_prog_key)
                clean_display_name = _clean(prog_display_name)
                if clean_record_prog != clean_current_prog and clean_record_prog != clean_display_name:
                    continue
            if current_sem and sem_val and sem_val != current_sem:
                continue

            sec_val = str(record.get('Section', 'A')).strip().upper()
            batch_val = str(record.get('Batch', '1')).strip()
            if batch_val.endswith('.0'): batch_val = batch_val[:-2]

            faculty_name = record.get('Faculty', '')
            if faculty_name and faculty_name != 'TBD':
                if faculty_name not in faculty_dict:
                    program = record.get('Program', '').upper()
                    school = PROGRAM_CONFIG.get(current_prog_key, {}).get('school', 'General')
                    
                    faculty_dict[faculty_name] = {
                        'name': faculty_name,
                        'id': f"F{len(faculty_dict)+1:03d}",
                        'department': school,
                        'subjects': [],
                        'max_hours': 20,
                        'semesters': set()  # CHANGE 4: Track semesters
                    }
                
                # Add subject and semester
                module_name = record.get('Module Name', '')
                if module_name and module_name not in faculty_dict[faculty_name]['subjects']:
                    faculty_dict[faculty_name]['subjects'].append(module_name)
                
                sem = record.get('Sem', 1)
                faculty_dict[faculty_name]['semesters'].add(sem)
        
        # Convert semesters set to list
        result = []
        for faculty in faculty_dict.values():
            faculty['semesters'] = list(faculty['semesters'])
            result.append(faculty)
        
        return result


# ==================== CHANGE 3: FACULTY MORNING CONSTRAINT MANAGER ====================

class FacultyMorningConstraintManager:
    """
    CHANGE 3: Manage faculty morning lecture constraints
    Ensures no faculty has more than 2 lectures at 9 AM per week
    """
    
    def __init__(self, firebase_manager=None):
        self.firebase = firebase_manager
        self.morning_counts = defaultdict(int)
    
    def initialize_counts(self, existing_schedules: dict = None):
        """Initialize morning counts from existing schedules"""
        self.morning_counts = defaultdict(int)
        
        if existing_schedules:
            for faculty, slots in existing_schedules.items():
                for slot_key in slots.keys():
                    if MORNING_SLOT_START in slot_key:
                        self.morning_counts[faculty] += 1
    
    def can_assign_morning(self, faculty_name: str) -> bool:
        """Check if faculty can be assigned another morning slot"""
        return self.morning_counts.get(faculty_name, 0) < FACULTY_MORNING_LIMIT
    
    def assign_morning(self, faculty_name: str) -> bool:
        """Assign a morning slot to faculty if allowed"""
        if self.can_assign_morning(faculty_name):
            self.morning_counts[faculty_name] += 1
            return True
        return False
    
    def get_morning_count(self, faculty_name: str) -> int:
        """Get current morning count for faculty"""
        return self.morning_counts.get(faculty_name, 0)
    
    def get_all_counts(self) -> dict:
        """Get all faculty morning counts"""
        return dict(self.morning_counts)
    
    def save_to_firebase(self, timetable_id: str):
        """Save morning counts to Firebase"""
        if self.firebase:
            return self.firebase.save_faculty_morning_counts(timetable_id, self.get_all_counts())
        return False, "Firebase not connected"


# ==================== CHANGE 4: FACULTY LUNCH UNION MANAGER ====================

class FacultyLunchUnionManager:
    """
    CHANGE 4: Manage faculty lunch unions for multi-semester cases
    Computes unavailable intervals for faculty teaching in multiple semesters with different lunches
    """
    
    def __init__(self, firebase_manager=None):
        self.firebase = firebase_manager
        self.faculty_unions = {}
    
    def compute_all_unions(self, faculties: List[dict], program: str):
        """Compute lunch unions for all faculty in a program"""
        self.faculty_unions = {}
        
        if not self.firebase:
            return self.faculty_unions
        
        for faculty in faculties:
            faculty_name = faculty.get('name', '')
            semesters = faculty.get('semesters', [])
            
            if len(semesters) > 1:
                # Faculty teaches in multiple semesters
                sem_lunch_configs = {}
                
                for sem in semesters:
                    config = self.firebase.get_semester_lunch_config(program, sem)
                    if config:
                        sem_lunch_configs[sem] = config
                    else:
                        # Use default
                        program_info = PROGRAM_CONFIG.get(program.upper(), {})
                        default_start = program_info.get('default_lunch_start', '13:00')
                        sem_lunch_configs[sem] = {
                            'start': default_start,
                            'duration': DEFAULT_LUNCH_DURATION
                        }
                
                # Compute union
                union = TimeSlotManager.compute_faculty_lunch_union(faculty_name, sem_lunch_configs)
                self.faculty_unions[faculty_name] = union
            else:
                # Single semester - no union needed
                self.faculty_unions[faculty_name] = []
        
        return self.faculty_unions
    
    def get_unavailable_times(self, faculty_name: str) -> List[Tuple[str, str]]:
        """Get unavailable time intervals for a faculty"""
        return self.faculty_unions.get(faculty_name, [])
    
    def is_time_available(self, faculty_name: str, time_start: str, time_end: str) -> bool:
        """Check if a time slot is available for faculty"""
        unavailable = self.faculty_unions.get(faculty_name, [])
        
        start_min = TimeSlotManager.time_to_minutes(time_start)
        end_min = TimeSlotManager.time_to_minutes(time_end)
        
        for u_start, u_end in unavailable:
            u_start_min = TimeSlotManager.time_to_minutes(u_start)
            u_end_min = TimeSlotManager.time_to_minutes(u_end)
            
            # Check for overlap
            if start_min < u_end_min and end_min > u_start_min:
                return False
        
        return True
    
    def save_to_firebase(self):
        """Save all faculty lunch unions to Firebase"""
        if self.firebase:
            return self.firebase.save_faculty_lunch_unions(self.faculty_unions)
        return False, "Firebase not connected"
    
# app.py - Part 3: SmartTimetableScheduler, Timetable Generation, Export Utilities
# Continuation from Part 2

# ==================== SMART TIMETABLE SCHEDULER (UPDATED) ====================

class SmartTimetableScheduler:
    """
    Main scheduler using Hybrid AI/ML algorithms with Firebase integration
    CHANGE 1, 2, 3, 4: Updated to support dynamic time slots, custom lunch/breaks,
    faculty morning limits, and lunch unions
    """
    
    def __init__(self, firebase_manager=None):
        self.genetic_algorithm = GeneticAlgorithm()
        self.hungarian_algorithm = HungarianAlgorithm()
        self.graph_coloring = GraphColoringAlgorithm()
        self.room_allocator = RoomAllocator(firebase_manager)
        self.morning_constraint_manager = FacultyMorningConstraintManager(firebase_manager)
        self.lunch_union_manager = FacultyLunchUnionManager(firebase_manager)
        self.firebase = firebase_manager
        self.days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        self.time_slot_manager = TimeSlotManager()
    
    def get_semester_config(self, program: str, semester: int) -> dict:
        """
        CHANGE 1, 2: Get complete configuration for a semester including lunch and breaks
        """
        config = {
            'lunch': None,
            'breaks': None,
            'time_slots': []
        }
        
        if self.firebase:
            # Get lunch config
            lunch_config = self.firebase.get_semester_lunch_config(program, semester)
            if lunch_config:
                config['lunch'] = lunch_config
            else:
                # Use default
                program_info = PROGRAM_CONFIG.get(program.upper(), {})
                config['lunch'] = {
                    'custom': False,
                    'start': program_info.get('default_lunch_start', '13:00'),
                    'duration': DEFAULT_LUNCH_DURATION,
                    'end': TimeSlotManager.minutes_to_time(
                        TimeSlotManager.time_to_minutes(program_info.get('default_lunch_start', '13:00')) + DEFAULT_LUNCH_DURATION
                    )
                }
            
            # Get break config
            break_config = self.firebase.get_semester_break_config(program, semester)
            if break_config and break_config.get('enabled', False):
                config['breaks'] = break_config
        
        # Generate time slots based on config
        config['time_slots'] = TimeSlotManager.generate_semester_slots(
            program, semester, config['lunch'], config['breaks']
        )
        
        return config
    
    def generate_dynamic_schedule_structure(self, program: str, semester: int, 
                                            batches: List[str]) -> dict:
        """
        CHANGE 1: Generate empty schedule structure with dynamic time slots
        Returns schedule structure with proper time columns based on lunch/break config
        """
        config = self.get_semester_config(program, semester)
        time_slots = config['time_slots']
        
        schedule_structure = {}
        
        for batch in batches:
            batch_key = f"Sem_{semester}_Section_{batch}"
            schedule_structure[batch_key] = {
                'config': config,
                'schedule': {}
            }
            
            for day in self.days:
                schedule_structure[batch_key]['schedule'][day] = {}
                
                for slot in time_slots:
                    slot_key = TimeSlotManager.get_slot_key(slot)
                    
                    if slot['type'] == 'lunch':
                        schedule_structure[batch_key]['schedule'][day][slot_key] = {
                            'subject': '🍴 LUNCH BREAK',
                            'faculty': '',
                            'room': 'Cafeteria',
                            'type': 'LUNCH',
                            'duration': slot['duration'],
                            'start': slot['start'],
                            'end': slot['end']
                        }
                    elif slot['type'] == 'break':
                        schedule_structure[batch_key]['schedule'][day][slot_key] = {
                            'subject': '☕ BREAK',
                            'faculty': '',
                            'room': '',
                            'type': 'BREAK',
                            'duration': slot['duration'],
                            'start': slot['start'],
                            'end': slot['end']
                        }
                    else:
                        schedule_structure[batch_key]['schedule'][day][slot_key] = None
        
        return schedule_structure
    
    def generate_hybrid_timetable(self, schools_data, faculties, subjects, rooms, 
                                   algorithm_choice='hybrid', room_allocations=None,
                                   program: str = None, semester: int = None):
        """
        CHANGE 1, 2, 3, 4: Generate timetable using selected algorithm with dynamic time slots
        """
        
        # Create main progress container
        progress_container = st.container()
        
        with progress_container:
            st.markdown("### 🔄 Timetable Generation in Progress...")
            st.markdown("---")
            
            # Progress tracking
            overall_progress = st.progress(0)
            status_text = st.empty()
            details_container = st.container()
            metrics_container = st.empty()
            
            # CHANGE 1: Get semester configuration
            semester_config = None
            if program and semester:
                status_text.info("📋 Loading semester configuration...")
                semester_config = self.get_semester_config(program, semester)
                
                with details_container:
                    st.markdown("#### ⚙️ Semester Configuration")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        lunch_info = semester_config.get('lunch', {})
                        lunch_str = f"{lunch_info.get('start', '13:00')} - {lunch_info.get('end', '13:50')}"
                        st.info(f"🍴 Lunch: {lunch_str} ({lunch_info.get('duration', 50)} min)")
                    with col2:
                        breaks_info = semester_config.get('breaks', {})
                        if breaks_info and breaks_info.get('enabled'):
                            st.info(f"☕ Breaks: {breaks_info.get('duration', 10)} min after lectures {breaks_info.get('placements', [])}")
                        else:
                            st.info("☕ Breaks: None configured")
                    with col3:
                        st.info(f"📊 Time Slots: {len(semester_config.get('time_slots', []))}")
                
                overall_progress.progress(5)
                time_module.sleep(0.3)
            
            # Apply room allocations - STRICT: each subject gets ONLY the room from the dataset
            # Keys in room_allocations: "Subject_type_Section" (e.g. "Quantum Physics_lab_A")
            # Section-specific keys are tried first before generic fallbacks.
            if room_allocations:
                status_text.info("Applying strict room assignments from Room Dataset...")
                for subject in subjects:
                    raw_name  = subject.get('name', '')
                    class_type = subject.get('type', 'Theory').lower()
                    sec_str = str(subject.get('section', '')).strip().upper()
                    
                    # Strip Lab/Tutorial/Practical suffix to get base name
                    stripped_name = raw_name
                    for sfx in (' Lab', ' Tutorial', ' Practical'):
                        if stripped_name.lower().endswith(sfx.lower()):
                            stripped_name = stripped_name[:-len(sfx)].strip()
                            break
                    
                    # Try section-specific keys first, then generic fallbacks
                    keys_to_try = []
                    if sec_str:
                        keys_to_try += [
                            f"{stripped_name}_{class_type}_{sec_str}",
                            f"{raw_name}_{class_type}_{sec_str}",
                        ]
                        if class_type == 'lab':
                            keys_to_try.append(f"{stripped_name}_lab_{sec_str}")
                        else:
                            keys_to_try.append(f"{stripped_name}_theory_{sec_str}")
                            keys_to_try.append(f"{stripped_name}_tutorial_{sec_str}")
                    # Generic fallbacks (no section suffix)
                    keys_to_try += [
                        f"{stripped_name}_{class_type}",
                        f"{raw_name}_{class_type}",
                        f"{stripped_name}_lab" if class_type == 'lab' else f"{stripped_name}_theory",
                    ]
                    
                    for k in keys_to_try:
                        if k in room_allocations:
                            val = room_allocations[k]
                            # val is a list; key is already section-specific so take val[0]
                            if isinstance(val, list) and len(val) > 0:
                                subject['assigned_room'] = val[0]
                            elif isinstance(val, str) and val:
                                subject['assigned_room'] = val
                            break
                overall_progress.progress(10)
                time_module.sleep(0.3)
            
            # Initialize morning constraint manager
            status_text.info("🌅 Initializing faculty morning constraints...")
            existing_faculty_schedules = {}
            existing_room_schedules = {}
            if self.firebase:
                # Exclude the current semester's own timetable so regenerating
                # does not over-block every slot that was in the previous run.
                _excl_key = None
                if program and semester:
                    _prog_schools = list(schools_data.keys())
                    _school_pfx = _prog_schools[0] if _prog_schools else ''
                    _excl_key = f"{_school_pfx}_Sem{semester}"
                existing_faculty_schedules = self.firebase.get_all_faculty_schedules(
                    exclude_timetable_key=_excl_key
                )
                existing_room_schedules = self.firebase.get_all_room_schedules(
                    exclude_timetable_key=_excl_key
                )
            
            self.morning_constraint_manager.initialize_counts(existing_faculty_schedules)
            overall_progress.progress(15)
            
            # CHANGE 4: Compute faculty lunch unions
            if program:
                status_text.info("👨‍🏫 Computing faculty lunch unions...")
                self.lunch_union_manager.compute_all_unions(faculties, program)
                overall_progress.progress(20)
            
            if algorithm_choice == 'hybrid':
                schedule = self._generate_hybrid_with_progress(
                    schools_data, faculties, subjects, rooms,
                    overall_progress, status_text, details_container, metrics_container,
                    existing_faculty_schedules, existing_room_schedules,
                    semester_config, program, semester
                )
            elif algorithm_choice == 'genetic_only':
                schedule = self._generate_ga_only_with_progress(
                    schools_data, faculties, subjects, rooms,
                    overall_progress, status_text, details_container,
                    existing_faculty_schedules, existing_room_schedules,
                    semester_config, program, semester
                )
            elif algorithm_choice == 'hungarian_graph':
                schedule = self._generate_hungarian_graph_with_progress(
                    schools_data, faculties, subjects, rooms,
                    overall_progress, status_text, details_container,
                    existing_faculty_schedules, existing_room_schedules,
                    semester_config, program, semester
                )
            else:
                schedule = self._generate_fallback_schedule(
                    schools_data, faculties, subjects, rooms,
                    semester_config, program, semester,
                    existing_faculty_schedules=existing_faculty_schedules,
                    existing_room_schedules=existing_room_schedules,
                )
            
            # CHANGE 3: Save morning constraint counts
            if program and semester:
                timetable_key = f"{program}_Sem{semester}"
                self.morning_constraint_manager.save_to_firebase(timetable_key)
            
            # CHANGE 4: Save faculty lunch unions
            self.lunch_union_manager.save_to_firebase()
            
            overall_progress.progress(100)
            status_text.success("✅ Timetable generation complete!")
            
            time_module.sleep(1)
            st.balloons()
            
            return schedule, semester_config
    
    def _generate_hybrid_with_progress(self, schools_data, faculties, subjects, rooms,
                                       overall_progress, status_text, details_container,
                                       metrics_container, existing_faculty_schedules,
                                       existing_room_schedules, semester_config,
                                       program, semester):
        """Generate using hybrid algorithm with progress tracking"""
        
        # ==================== PHASE 1: HUNGARIAN ALGORITHM ====================
        with details_container:
            st.markdown("#### 🎯 Phase 1: Hungarian Algorithm")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Faculty Count", len(faculties))
            with col2:
                st.metric("Subjects Count", len(subjects))
            with col3:
                st.metric("Status", "Running...")
            
            hungarian_progress = st.progress(0)
            hungarian_status = st.empty()
        
        status_text.info("🎯 Phase 1: Running Hungarian Algorithm for optimal faculty-course assignment...")
        
        # CHANGE 3: Get morning counts for Hungarian
        morning_counts = self.morning_constraint_manager.get_all_counts()
        
        hungarian_steps = [
            ("Creating cost matrix...", 20),
            ("Calculating optimal assignments...", 50),
            ("Validating faculty constraints...", 70),
            ("Finalizing assignments...", 100)
        ]
        
        for step_name, step_progress in hungarian_steps:
            hungarian_status.text(f"   └─ {step_name}")
            hungarian_progress.progress(step_progress)
            time_module.sleep(0.4)
        
        faculty_assignments = self.hungarian_algorithm.solve(faculties, subjects, morning_counts)
        
        assignments_made = 0
        for subject in subjects:
            if subject['name'] in faculty_assignments:
                subject['faculty'] = faculty_assignments[subject['name']]
                assignments_made += 1
        
        with details_container:
            st.success(f"✅ Phase 1 Complete: {assignments_made} faculty assignments optimized")
        
        overall_progress.progress(35)
        time_module.sleep(0.5)
        
        # ==================== PHASE 2: GRAPH COLORING ====================
        with details_container:
            st.markdown("#### 🎨 Phase 2: Graph Coloring Algorithm")
            col1, col2, col3 = st.columns(3)
            
            graph_progress = st.progress(0)
            graph_status = st.empty()
        
        status_text.info("🎨 Phase 2: Applying Graph Coloring for conflict-free slot allocation...")
        
        # Build classes list
        classes = []
        graph_status.text("   └─ Building conflict graph...")
        graph_progress.progress(10)
        time_module.sleep(0.3)
        
        # CHANGE 1: Get dynamic time slots
        time_slots = semester_config.get('time_slots', []) if semester_config else []
        lecture_slots = TimeSlotManager.get_lecture_slots_only(time_slots)
        
        for school_key, school_data in schools_data.items():
            # Only iterate semesters that have batches configured
            configured_sems = list(school_data.get('batches', {}).keys())
            if not configured_sems:
                configured_sems = [semester] if semester else [1]
            for year in configured_sems:
                batches = school_data.get('batches', {}).get(year, [1])
                for batch in batches:
                    def _match_section(s_val, b_val):
                        s_val = str(s_val).strip().upper() if s_val else ''
                        b_val = str(b_val).strip().upper()
                        if s_val == b_val: return True
                        if b_val.isdigit(): return s_val == chr(64 + int(b_val))
                        return False

                    batch_subjects = [s for s in subjects
                                    if (s.get('school', '').upper() in school_key.upper() or
                                        s.get('program', '').upper() in school_key.upper()) and
                                    (str(s.get('year')) == str(year) or str(s.get('semester')) == str(year)) and 
                                    _match_section(s.get('section'), str(batch))]

                    for subject in batch_subjects:
                        for session in range(max(1, subject.get('weekly_hours', 3))):
                            classes.append({
                                'school': school_key,
                                'batch': f"Sem_{year}_Section_{batch}",
                                'subject': subject['name'],
                                'faculty': subject.get('faculty', 'TBD'),
                                'type': subject.get('type', 'Theory'),
                                'room': subject.get('assigned_room', 'TBD'),
                                'duration': subject.get('duration', DEFAULT_LECTURE_DURATION),
                                'real_batch': subject.get('batch', '1') # Required for parallel labs
                            })
        
        with details_container:
            col1.metric("Classes to Schedule", len(classes))
        
        graph_status.text("   └─ Identifying conflicts...")
        graph_progress.progress(30)
        time_module.sleep(0.3)
        
        # Build available slots from dynamic time slots
        available_slots = []
        for day in self.days[:5]:  # Mon-Fri
            for slot in lecture_slots:
                slot_key = TimeSlotManager.get_slot_key(slot)
                available_slots.append((day, slot_key, slot))
        
        with details_container:
            col2.metric("Available Slots", len(available_slots))
        
        graph_status.text("   └─ Applying Welsh-Powell coloring...")
        graph_progress.progress(60)
        time_module.sleep(0.3)
        
        # CHANGE 4: Apply graph coloring with faculty lunch unions
        faculty_lunch_unions = self.lunch_union_manager.faculty_unions
        slot_assignments = self.graph_coloring.color_graph(classes, [(s[0], s[1]) for s in available_slots])
        
        graph_status.text("   └─ Validating slot assignments...")
        graph_progress.progress(90)
        time_module.sleep(0.3)
        
        colors_used = len(set(slot_assignments.values()))
        with details_container:
            col3.metric("Time Slots Used", colors_used)
        
        graph_progress.progress(100)
        
        with details_container:
            st.success(f"✅ Phase 2 Complete: {len(classes)} classes assigned to {colors_used} unique time slots")
        
        overall_progress.progress(55)
        time_module.sleep(0.5)
        
        # ==================== PHASE 3: GENETIC ALGORITHM ====================
        with details_container:
            st.markdown("#### 🧬 Phase 3: Genetic Algorithm Optimization")
            col1, col2, col3, col4 = st.columns(4)
            
            ga_progress = st.progress(0)
            ga_status = st.empty()
            ga_metrics = st.empty()
        
        status_text.info("🧬 Phase 3: Running Genetic Algorithm for final optimization...")
        
        ga_status.text("   └─ Creating initial population...")
        ga_progress.progress(5)
        
        # Convert to schedule format with dynamic slots
        initial_schedule = self._convert_to_dynamic_schedule(
            classes, slot_assignments, schools_data, rooms,
            semester_config, program, semester
        )
        
        # Create constraints with all new features
        constraints = create_constraints(schools_data, subjects, faculties, rooms)
        constraints['initial_schedule'] = initial_schedule
        constraints['existing_faculty_schedules'] = existing_faculty_schedules
        constraints['existing_room_schedules'] = existing_room_schedules
        constraints['semester_config'] = semester_config
        constraints['semester'] = semester
        constraints['faculty_morning_counts'] = self.morning_constraint_manager.get_all_counts()
        constraints['faculty_lunch_unions'] = self.lunch_union_manager.faculty_unions
        # Elective groups: subjects in the same group are alternatives — only ONE
        # needs to be scheduled per section.  GA uses this to skip extras.
        constraints['elective_groups'] = getattr(self, 'elective_groups', [])
        
        with details_container:
            col1.metric("Population Size", self.genetic_algorithm.population_size)
            col2.metric("Generations", "30")
            col3.metric("Mutation Rate", f"{self.genetic_algorithm.mutation_rate*100:.0f}%")
        
        ga_status.text("   └─ Evolving population...")
        
        # Run GA with progress
        optimized_schedule = self._evolve_with_progress(
            constraints, 
            generations=30,
            ga_progress=ga_progress,
            ga_status=ga_status,
            ga_metrics_placeholder=ga_metrics,
            details_col4=col4
        )
        
        with details_container:
            st.success("✅ Phase 3 Complete: Schedule optimized!")

        overall_progress.progress(90)

        # ── PHASE 4: Agentic AI clash repair + legacy fallback ──────────────
        status_text.info("🔍 Phase 4: Detecting clashes and running Agentic AI repair...")
        clash_detector = ClashDetector(self.firebase)
        _p4_existing_fac = constraints.get('existing_faculty_schedules', {})
        _p4_existing_room = constraints.get('existing_room_schedules', {})
        remaining_clashes = clash_detector.detect_all_clashes(
            optimized_schedule,
            existing_faculty_schedules=_p4_existing_fac,
            existing_room_schedules=_p4_existing_room,
        )

        agent_log_lines = []
        agent_summary = {}

        if remaining_clashes:
            with details_container:
                _cross_count = sum(1 for c in remaining_clashes
                                   if 'Cross-Semester' in c.get('type', ''))
                _intra_count = len(remaining_clashes) - _cross_count
                st.markdown(
                    f"#### 🤖 Phase 4: Agentic Clash Repair  "
                    f"({_intra_count} intra-schedule, {_cross_count} cross-semester)"
                )
                agent_log_placeholder = st.empty()
                repair_progress = st.progress(0)

            from agent.integration import format_turn_log_entry, run_agentic_clash_repair

            def _phase4_turn_callback(turn, response):
                agent_log_lines.extend(format_turn_log_entry(turn, response))
                agent_log_placeholder.markdown(
                    "##### AGENT LOG (live)\n\n" + "\n\n".join(agent_log_lines[-20:])
                )
                repair_progress.progress(min(90, turn * 10))

            optimized_schedule, agent_summary, remaining_clashes, agent_log_lines = (
                run_agentic_clash_repair(
                    firebase_manager=self.firebase,
                    genetic_algorithm=self.genetic_algorithm,
                    schedule=optimized_schedule,
                    clashes=remaining_clashes,
                    constraints=constraints,
                    program=program,
                    semester=semester,
                    clash_detector=clash_detector,
                    on_turn_callback=_phase4_turn_callback,
                )
            )

            repair_progress.progress(100)
            st.session_state['last_agent_session'] = agent_summary
            st.session_state['last_agent_log'] = agent_log_lines

            final_clash_count = len(remaining_clashes)
            with details_container:
                if agent_log_lines:
                    with st.expander("🤖 Agent Repair Log", expanded=False):
                        st.markdown("\n\n".join(agent_log_lines))
                if final_clash_count == 0:
                    st.success("✅ Phase 4 Complete: Zero clashes!")
                else:
                    st.warning(
                        f"⚠️ Phase 4: {final_clash_count} clash(es) could not be auto-resolved. "
                        "Use the 🤖 AI Agent or Edit tab to fix manually."
                    )
        else:
            final_clash_count = 0
            with details_container:
                st.success("✅ Phase 4: No clashes detected — schedule is clean!")

        overall_progress.progress(95)

        # ── PHASE 5: Lecture count validation ─────────────────────────────────
        status_text.info("📊 Phase 5: Verifying lecture counts against info_dataset...")
        _hour_violations = self.genetic_algorithm._check_hour_violations(
            optimized_schedule, constraints
        )
        # Build a per-subject completion report for display
        _completion_report = []
        _subjects_for_check = constraints.get('subjects', [])
        if _subjects_for_check:
            from collections import defaultdict as _ddc
            _sched_counts = _ddc(int)
            for _sch in optimized_schedule:
                for _bk in optimized_schedule[_sch]:
                    for _d in optimized_schedule[_sch][_bk]:
                        for _sk, _sc in optimized_schedule[_sch][_bk][_d].items():
                            _items = _sc if isinstance(_sc, list) else [_sc]
                            for _ci in _items:
                                if isinstance(_ci, dict) and _ci.get('type') not in ('LUNCH','BREAK',None):
                                    _nm = str(_ci.get('subject','')).strip().upper()
                                    _ct = str(_ci.get('type','')).upper()
                                    _is_lab = any(t in _ct for t in ['LAB','TUTORIAL','PRACTICAL'])
                                    _b = str(_ci.get('batch','1')).strip().replace('.0','').upper()
                                    _ckey = (_nm, _b if _is_lab else '_ALL')
                                    _sched_counts[_ckey] += 1

            _seen_keys = set()
            for _s in _subjects_for_check:
                _sn = str(_s.get('name','')).strip().upper()
                _sec = str(_s.get('section','')).strip().upper()
                _is_lab = any(t in str(_s.get('type','')).upper() for t in ['LAB','TUTORIAL','PRACTICAL'])
                _b = str(_s.get('batch','1')).strip().replace('.0','').upper()
                _ckey = (_sn, _b if _is_lab else '_ALL')
                if _ckey in _seen_keys:
                    continue
                _seen_keys.add(_ckey)
                _target = int(_s.get('weekly_hours', 0) or 0)
                if _target <= 0:
                    continue
                _actual = _sched_counts.get(_ckey, 0)
                _completion_report.append({
                    'subject': _s.get('name',''),
                    'section': _sec,
                    'type': _s.get('type','Theory'),
                    'required': _target,
                    'scheduled': _actual,
                    'ok': _actual >= _target,
                })

        with details_container:
            st.markdown("#### 📋 Phase 5: Lecture Count Report")
            if _completion_report:
                _ok_count = sum(1 for r in _completion_report if r['ok'])
                _fail_count = len(_completion_report) - _ok_count
                if _fail_count == 0:
                    st.success(f"✅ All {_ok_count} subject(s) have their required lecture counts!")
                else:
                    st.warning(f"⚠️ {_fail_count} subject(s) could not be fully scheduled "
                               f"(faculty/slot shortage). {_ok_count} complete.")
                    import pandas as _pd_rep
                    _rep_rows = [r for r in _completion_report if not r['ok']]
                    _rep_df = _pd_rep.DataFrame(_rep_rows)[['subject','section','type','required','scheduled']]
                    _rep_df.columns = ['Subject','Section','Type','Required/Week','Scheduled']
                    st.dataframe(_rep_df, use_container_width=True, hide_index=True)
            else:
                st.info("ℹ️ No subject data available for count validation.")

        overall_progress.progress(98)

        # Finalize
        status_text.info("🏁 Finalizing timetable...")
        self._add_lunch_and_breaks(optimized_schedule, semester_config)

        # Show final stats
        with metrics_container:
            st.markdown("#### 📊 Generation Summary")
            final_col1, final_col2, final_col3, final_col4 = st.columns(4)

            _total_subjects = len(set(r['subject'] for r in _completion_report)) if _completion_report else len(classes)
            _complete_subjs = sum(1 for r in _completion_report if r['ok']) if _completion_report else _total_subjects
            final_col1.metric("✅ Subjects Complete",
                              f"{_complete_subjs}/{_total_subjects}",
                              delta="All done" if _complete_subjs == _total_subjects else f"{_total_subjects-_complete_subjs} short")
            final_col2.metric("🎯 Faculty Assigned", assignments_made)
            clash_icon = "✅" if final_clash_count == 0 else "⚠️"
            final_col3.metric(f"{clash_icon} Clashes", final_clash_count)
            final_col4.metric("⏱️ Algorithm", "Hybrid")

        return optimized_schedule
    
    def _evolve_with_progress(self, constraints, generations, ga_progress, ga_status, 
                               ga_metrics_placeholder, details_col4):
        """Run genetic algorithm with visual progress"""
        
        ga_status.text("   └─ Initializing population...")
        ga_progress.progress(10)
        
        population = []
        for i in range(self.genetic_algorithm.population_size):
            individual = self.genetic_algorithm.create_individual(constraints)
            individual['fitness'] = self.genetic_algorithm.fitness(individual, constraints)
            population.append(individual)
        
        best_individual = None
        best_fitness = -float('inf')
        
        for generation in range(generations):
            progress_pct = 10 + int((generation / generations) * 85)
            ga_progress.progress(progress_pct)
            ga_status.text(f"   └─ Generation {generation + 1}/{generations}")
            
            for ind in population:
                ind['fitness'] = self.genetic_algorithm.fitness(ind, constraints)
            
            population.sort(key=lambda x: x['fitness'], reverse=True)
            
            current_best = population[0]
            if current_best['fitness'] > best_fitness:
                best_fitness = current_best['fitness']
                best_individual = copy.deepcopy(current_best)
            
            with ga_metrics_placeholder:
                mcol1, mcol2 = st.columns(2)
                mcol1.metric("Best Fitness", f"{best_fitness:.0f}/1000")
                mcol2.metric("Clashes", current_best.get('clashes', 0))
            
            details_col4.metric("Generation", f"{generation + 1}/{generations}")
            
            if current_best.get('clashes', 0) == 0 and current_best['fitness'] >= 900:
                ga_status.text(f"   └─ Perfect solution found at generation {generation + 1}!")
                break
            
            # Create new population
            new_population = []
            new_population.extend(copy.deepcopy(population[:self.genetic_algorithm.elitism_size]))
            
            while len(new_population) < self.genetic_algorithm.population_size:
                parent1 = self.genetic_algorithm._tournament_selection(population)
                parent2 = self.genetic_algorithm._tournament_selection(population)
                
                if random.random() < self.genetic_algorithm.crossover_rate:
                    child = self.genetic_algorithm.crossover(parent1, parent2)
                else:
                    child = copy.deepcopy(parent1)
                
                if random.random() < self.genetic_algorithm.mutation_rate:
                    child = self.genetic_algorithm.mutate(child, constraints)
                
                new_population.append(child)
            
            population = new_population
            time_module.sleep(0.1)
        
        ga_progress.progress(100)

        return best_individual['schedule'] if best_individual else {}
    
    def _convert_to_dynamic_schedule(self, classes, slot_assignments, schools_data, 
                                      rooms, semester_config, program, semester):
        """
        CHANGE 1: Convert graph coloring output to schedule with dynamic time slots
        """
        schedule = {}
        time_slots = semester_config.get('time_slots', []) if semester_config else []

        for school_key in schools_data:
            schedule[school_key] = {}
            # Only iterate semesters that are actually configured (not all years 1-N)
            _conv_sems = list(schools_data[school_key].get('batches', {}).keys())
            if not _conv_sems:
                _conv_sems = [semester] if semester else [1]
            for year in _conv_sems:
                batches = schools_data[school_key].get('batches', {}).get(year, [1])
                for batch in batches:
                    batch_key = f"Sem_{year}_Section_{batch}"
                    schedule[school_key][batch_key] = {}
                    
                    for day in self.days[:5]:
                        schedule[school_key][batch_key][day] = {}
                        
                        for slot in time_slots:
                            slot_key = TimeSlotManager.get_slot_key(slot)
                            
                            if slot['type'] == 'lunch':
                                schedule[school_key][batch_key][day][slot_key] = {
                                    'subject': '🍴 LUNCH BREAK',
                                    'faculty': '',
                                    'room': 'Cafeteria',
                                    'type': 'LUNCH',
                                    'duration': slot['duration'],
                                    'start': slot['start'],
                                    'end': slot['end']
                                }
                            elif slot['type'] == 'break':
                                schedule[school_key][batch_key][day][slot_key] = {
                                    'subject': '☕ BREAK',
                                    'faculty': '',
                                    'room': '',
                                    'type': 'BREAK',
                                    'duration': slot['duration'],
                                    'start': slot['start'],
                                    'end': slot['end']
                                }
                            else:
                                schedule[school_key][batch_key][day][slot_key] = None
        
        # Assign classes from graph coloring
        room_index = 0
        for i, class_info in enumerate(classes):
            if i in slot_assignments:
                day, slot_key = slot_assignments[i]
                school = class_info['school']
                batch = class_info['batch']
                
                if school in schedule and batch in schedule[school]:
                    if day in schedule[school][batch] and slot_key in schedule[school][batch][day]:
                        # Skip lunch and break slots
                        current = schedule[school][batch][day].get(slot_key)
                        if isinstance(current, dict) and current.get('type') in ['LUNCH', 'BREAK']:
                            continue
                        
                        # ROOM-FIX: track whether room came from dataset (locked) or fallback
                        if class_info.get('room') and class_info['room'] != 'TBD':
                            room_name = class_info['room']
                            room_is_locked = True   # from Room Dataset - never overwrite
                        elif rooms:
                            room = rooms[room_index % len(rooms)]
                            room_name = room.get('name', 'TBD')
                            room_index += 1
                            room_is_locked = False
                        else:
                            room_name = 'TBD'
                            room_is_locked = False

                        # Find slot info
                        slot_info = next((s for s in time_slots 
                                         if TimeSlotManager.get_slot_key(s) == slot_key), None)
                        
                        new_class = {
                            'subject': class_info['subject'],
                            'faculty': class_info['faculty'],
                            'room': room_name,
                            'room_locked': room_is_locked,   # ROOM-FIX: preserved through GA
                            'type': class_info['type'],
                            'duration': class_info.get('duration', DEFAULT_LECTURE_DURATION),
                            'start': slot_info['start'] if slot_info else '',
                            'end': slot_info['end'] if slot_info else '',
                            'batch': class_info.get('real_batch', '1')
                        }
                        
                        current_slot_val = schedule[school][batch][day].get(slot_key)
                        
                        if current_slot_val is None:
                            schedule[school][batch][day][slot_key] = new_class
                        else:
                            # Safely handle parallel labs if they got mapped to the same slot
                            new_type = str(new_class.get('type', '')).upper()
                            if 'LAB' in new_type or 'TUTORIAL' in new_type or 'PRACTICAL' in new_type:
                                if isinstance(current_slot_val, dict):
                                    curr_type = str(current_slot_val.get('type', '')).upper()
                                    if 'LAB' in curr_type or 'TUTORIAL' in curr_type or 'PRACTICAL' in curr_type:
                                        schedule[school][batch][day][slot_key] = [current_slot_val, new_class]
                                    else:
                                        schedule[school][batch][day][slot_key] = new_class
                                elif isinstance(current_slot_val, list):
                                    current_slot_val.append(new_class)
                            else:
                                schedule[school][batch][day][slot_key] = new_class
        
        return schedule
    
    def _generate_ga_only_with_progress(self, schools_data, faculties, subjects, rooms,
                                        overall_progress, status_text, details_container,
                                        existing_faculty_schedules, existing_room_schedules,
                                        semester_config, program, semester):
        """Generate using only Genetic Algorithm with progress"""
        
        with details_container:
            st.markdown("#### 🧬 Genetic Algorithm Only Mode")
            col1, col2, col3 = st.columns(3)
            col1.metric("Population", self.genetic_algorithm.population_size)
            col2.metric("Generations", "50")
            col3.metric("Status", "Initializing...")
            
            ga_progress = st.progress(0)
            ga_status = st.empty()
        
        constraints = create_constraints(schools_data, subjects, faculties, rooms)
        constraints['existing_faculty_schedules'] = existing_faculty_schedules
        constraints['existing_room_schedules'] = existing_room_schedules
        constraints['semester_config'] = semester_config
        constraints['semester'] = semester
        constraints['faculty_morning_counts'] = self.morning_constraint_manager.get_all_counts()
        constraints['faculty_lunch_unions'] = self.lunch_union_manager.faculty_unions
        
        status_text.info("🧬 Running Genetic Algorithm optimization...")
        
        optimized_schedule = self.genetic_algorithm.evolve(constraints, generations=50, verbose=False)
        
        if optimized_schedule:
            self._add_lunch_and_breaks(optimized_schedule, semester_config)
            overall_progress.progress(100)
            status_text.success("✅ Genetic Algorithm completed!")
            
            with details_container:
                col3.metric("Status", "Complete ✅")
            
            return optimized_schedule
        else:
            return self._generate_fallback_schedule(schools_data, faculties, subjects, rooms,
                                                    semester_config, program, semester)
    
    def _generate_hungarian_graph_with_progress(self, schools_data, faculties, subjects, rooms,
                                                overall_progress, status_text, details_container,
                                                existing_faculty_schedules=None, existing_room_schedules=None,
                                                semester_config=None, program=None, semester=None):
        """Generate using Hungarian + Graph Coloring with progress"""
        
        with details_container:
            st.markdown("#### 🎯 Hungarian + Graph Coloring Mode")
            col1, col2 = st.columns(2)
            
            hungarian_progress = st.progress(0)
            graph_progress = st.progress(0)
        
        # Hungarian Algorithm
        status_text.info("🎯 Running Hungarian Algorithm...")
        col1.metric("Hungarian", "Running...")
        
        morning_counts = self.morning_constraint_manager.get_all_counts()
        faculty_assignments = self.hungarian_algorithm.solve(faculties, subjects, morning_counts)
        
        for subject in subjects:
            if subject['name'] in faculty_assignments:
                subject['faculty'] = faculty_assignments[subject['name']]
        
        hungarian_progress.progress(100)
        col1.metric("Hungarian", "Complete ✅")
        overall_progress.progress(50)
        
        # Graph Coloring
        status_text.info("🎨 Applying Graph Coloring...")
        col2.metric("Graph Coloring", "Running...")
        
        schedule = self._generate_with_graph_coloring(
            schools_data, subjects, faculties, rooms,
            semester_config, program, semester,
            existing_faculty_schedules=existing_faculty_schedules or {},
            existing_room_schedules=existing_room_schedules or {},
        )
        
        graph_progress.progress(100)
        col2.metric("Graph Coloring", "Complete ✅")
        overall_progress.progress(100)
        
        status_text.success("✅ Hungarian + Graph Coloring completed!")
        
        self._add_lunch_and_breaks(schedule, semester_config)
        return schedule
    
    def _generate_with_graph_coloring(self, schools_data, subjects, faculties, rooms,
                                       semester_config, program, semester,
                                       existing_faculty_schedules=None,
                                       existing_room_schedules=None):
        """Generate schedule using graph coloring only with dynamic slots.

        Pre-blocks slots that are already occupied by OTHER semesters (loaded
        from Firebase) so graph coloring never creates cross-semester faculty/room clashes.
        """
        from collections import defaultdict as _dd
        existing_faculty_schedules = existing_faculty_schedules or {}
        existing_room_schedules = existing_room_schedules or {}

        # Build per-faculty and per-room blocked slot sets from Firebase data
        _fac_blocked = _dd(set)   # faculty_name -> set of "Day_slot" strings
        for _fac, _slots in existing_faculty_schedules.items():
            _fac_blocked[_fac].update(_slots.keys())
        _room_blocked = _dd(set)
        for _rm, _slots in existing_room_schedules.items():
            _room_blocked[_rm].update(_slots.keys())

        schedule = {}
        time_slots = semester_config.get('time_slots', []) if semester_config else []
        lecture_slots_only = [s for s in time_slots if s.get('type', 'lecture') == 'lecture']
        days = self.days[:5]

        # Build subject lookup by section/year
        def _subj_key(s):
            return (str(s.get('school', '')).upper(),
                    str(s.get('year', s.get('semester', ''))),
                    str(s.get('section', '')).upper())

        subj_by_key = _dd(list)
        for s in subjects:
            subj_by_key[_subj_key(s)].append(s)

        for school_key, school_data in schools_data.items():
            schedule[school_key] = {}
            school_type = school_key.split('_')[0] if '_' in school_key else school_key
            configured_sems = list(school_data.get('batches', {}).keys())
            if not configured_sems:
                configured_sems = [semester] if semester else [1]
            for year in configured_sems:
                batches = school_data.get('batches', {}).get(year, [1])
                for batch_idx in batches:
                    sec_letter = chr(64 + batch_idx)  # 1→A, 2→B …
                    batch_key = f"Sem_{year}_Section_{batch_idx}"
                    batch_schedule = {}

                    for day in days:
                        batch_schedule[day] = {}
                        for slot in time_slots:
                            slot_key = TimeSlotManager.get_slot_key(slot)
                            if slot['type'] == 'lunch':
                                batch_schedule[day][slot_key] = {
                                    'subject': '🍴 LUNCH BREAK', 'faculty': '',
                                    'room': 'Cafeteria', 'type': 'LUNCH',
                                    'duration': slot['duration'],
                                    'start': slot['start'], 'end': slot['end']
                                }
                            elif slot['type'] == 'break':
                                batch_schedule[day][slot_key] = {
                                    'subject': '☕ BREAK', 'faculty': '',
                                    'room': '', 'type': 'BREAK',
                                    'duration': slot['duration'],
                                    'start': slot['start'], 'end': slot['end']
                                }
                            else:
                                batch_schedule[day][slot_key] = None

                    # Try to assign subjects for this section using graph coloring
                    # with cross-semester awareness
                    sk_candidates = [
                        (school_type, str(year), sec_letter),
                        (school_type, str(year), ''),
                        ('', str(year), sec_letter),
                        ('', str(year), ''),
                    ]
                    sec_subjects = []
                    for sk in sk_candidates:
                        sec_subjects = subj_by_key.get(sk, [])
                        if sec_subjects:
                            break

                    if sec_subjects:
                        # Track per-batch used slots to avoid intra-schedule clashes
                        used_slots = set()  # "Day_slot_key" strings
                        fac_used = _dd(set)
                        room_used = _dd(set)

                        for subj in sec_subjects:
                            wh = int(subj.get('weekly_hours', 3) or 3)
                            faculty = subj.get('faculty', 'TBD') or 'TBD'
                            # STRICT: prefer assigned_room, then fall back to 'room' key
                            room = subj.get('assigned_room') or subj.get('room', 'TBD') or 'TBD'
                            scheduled = 0
                            for day in days:
                                if scheduled >= wh:
                                    break
                                for slot in lecture_slots_only:
                                    if scheduled >= wh:
                                        break
                                    sk = TimeSlotManager.get_slot_key(slot)
                                    day_sk = f"{day}_{sk}"
                                    if day_sk in used_slots:
                                        continue
                                    if faculty != 'TBD':
                                        if day_sk in _fac_blocked[faculty]:
                                            continue  # faculty busy in another semester
                                        if day_sk in fac_used[faculty]:
                                            continue  # faculty busy within this semester
                                    if room != 'TBD':
                                        if day_sk in _room_blocked[room]:
                                            continue  # room busy in another semester
                                        if day_sk in room_used[room]:
                                            continue  # room busy within this semester
                                        # STRICT: room is assigned — if occupied we skip, never substitute
                                    # Place the class
                                    batch_schedule[day][sk] = {
                                        'subject': subj['name'],
                                        'faculty': faculty,
                                        'room': room,
                                        'room_locked': room != 'TBD',   # ROOM-FIX: lock dataset rooms
                                        'type': subj.get('type', 'Theory'),
                                        'batch': str(subj.get('batch', '')),
                                        'section': sec_letter,
                                        'duration': slot.get('duration', 60),
                                        'start': slot.get('start', ''),
                                        'end': slot.get('end', ''),
                                    }
                                    used_slots.add(day_sk)
                                    if faculty != 'TBD':
                                        fac_used[faculty].add(day_sk)
                                        _fac_blocked[faculty].add(day_sk)  # block for subsequent sections
                                    if room != 'TBD':
                                        room_used[room].add(day_sk)
                                        _room_blocked[room].add(day_sk)
                                    scheduled += 1

                    schedule[school_key][batch_key] = batch_schedule

        return schedule
    
    def _add_lunch_and_breaks(self, schedule, semester_config):
        """
        CHANGE 1, 2: Add lunch and breaks to the schedule based on configuration
        """
        if not semester_config:
            return
        
        time_slots = semester_config.get('time_slots', [])
        
        for school_key in schedule:
            for batch in schedule[school_key]:
                for day in self.days[:5]:
                    if day not in schedule[school_key][batch]:
                        continue
                    
                    for slot in time_slots:
                        slot_key = TimeSlotManager.get_slot_key(slot)
                        
                        if slot['type'] == 'lunch':
                            schedule[school_key][batch][day][slot_key] = {
                                'subject': '🍴 LUNCH BREAK',
                                'faculty': '',
                                'room': 'Cafeteria',
                                'type': 'LUNCH',
                                'duration': slot['duration'],
                                'start': slot['start'],
                                'end': slot['end']
                            }
                        elif slot['type'] == 'break':
                            schedule[school_key][batch][day][slot_key] = {
                                'subject': '☕ BREAK',
                                'faculty': '',
                                'room': '',
                                'type': 'BREAK',
                                'duration': slot['duration'],
                                'start': slot['start'],
                                'end': slot['end']
                            }
    
    def _generate_fallback_schedule(self, schools_data, faculties, subjects, rooms,
                                     semester_config=None, program=None, semester=None,
                                     existing_faculty_schedules=None,
                                     existing_room_schedules=None):
        """Fallback schedule generation with dynamic time slots"""
        schedule = {}
        faculty_tracker = defaultdict(lambda: defaultdict(list))
        room_tracker = defaultdict(lambda: defaultdict(list))

        # Pre-block slots occupied by OTHER semesters in Firebase
        for _fac, _slots in (existing_faculty_schedules or {}).items():
            for _sk in _slots.keys():
                faculty_tracker[_fac][_sk].append({'source': 'firebase'})
        for _rm, _slots in (existing_room_schedules or {}).items():
            for _sk in _slots.keys():
                room_tracker[_rm][_sk].append({'source': 'firebase'})

        # CHANGE 1: Get time slots from config
        time_slots = semester_config.get('time_slots', []) if semester_config else []
        lecture_slots = TimeSlotManager.get_lecture_slots_only(time_slots)

        # FIXED: Build a subject-name → assigned_room cache so that
        # each lab/tutorial gets its specific room rather than Physics Lab-1.
        # Key = subject['name'] (with " Lab" / " Tutorial" suffixes intact).
        _subj_room_cache = {}
        for _s in subjects:
            _rn = _s.get('assigned_room', '')
            if _rn:
                _subj_room_cache[_s.get('name', '')] = _rn
        
        for school_key, school_data in schools_data.items():
            school_name = 'STME' if 'STME' in school_key else ('SOC' if 'SOC' in school_key else 'SOL')
            schedule[school_key] = {}
            # Only iterate semesters that have batches configured
            configured_sems = list(school_data.get('batches', {}).keys())
            if not configured_sems:
                configured_sems = [semester] if semester else [1]
            for year in configured_sems:
                batches = school_data.get('batches', {}).get(year, [1])

                for batch in batches:
                    batch_key = f"Sem_{year}_Section_{batch}"
                    batch_schedule = {}
                    
                    for day in self.days[:5]:
                        batch_schedule[day] = {}
                        
                        for slot in time_slots:
                            slot_key = TimeSlotManager.get_slot_key(slot)
                            
                            if slot['type'] == 'lunch':
                                batch_schedule[day][slot_key] = {
                                    'subject': '🍴 LUNCH BREAK',
                                    'faculty': '',
                                    'room': 'Cafeteria',
                                    'type': 'LUNCH',
                                    'duration': slot['duration'],
                                    'start': slot['start'],
                                    'end': slot['end']
                                }
                            elif slot['type'] == 'break':
                                batch_schedule[day][slot_key] = {
                                    'subject': '☕ BREAK',
                                    'faculty': '',
                                    'room': '',
                                    'type': 'BREAK',
                                    'duration': slot['duration'],
                                    'start': slot['start'],
                                    'end': slot['end']
                                }
                            else:
                                batch_schedule[day][slot_key] = None
                    
                    # Assign subjects to lecture slots
                    # FIXED: filter by BOTH year AND section so that when the info_dataset
                    # contains entries for multiple sections/batches (e.g. A/1, A/2, B/1, B/2)
                    # only the subjects that belong to this section are scheduled here.
                    # We derive the section letter from the 1-based batch index so that
                    # batch=1 → section A, batch=2 → section B, etc.
                    _sections_for_year = sorted(
                        school_data.get('section_batches', {}).get(year, {}).keys()
                    )  # e.g. ['A', 'B']
                    if _sections_for_year and isinstance(batch, int) and 0 < batch <= len(_sections_for_year):
                        _sec_letter = _sections_for_year[batch - 1]
                    elif isinstance(batch, int):
                        _sec_letter = chr(64 + batch)
                    else:
                        _sec_letter = None

                    def _subj_matches_section(s):
                        """True if the subject record belongs to this section."""
                        _s_sec = str(s.get('section', '')).strip().upper()
                        _s_bch = str(s.get('batch', '')).strip()
                        # If subject has no section info, include it (backward compat)
                        if not _s_sec:
                            return True
                        # Match by section letter if we know it
                        if _sec_letter and _s_sec != _sec_letter.upper():
                            return False
                        return True

                    # Collect subjects for this section.
                    # PARALLEL LABS: For lab subjects, we group ALL batch-entries
                    # (e.g. Batch 1 and Batch 2 of the same lab) so they can be
                    # placed into the SAME time slot as a stacked list — exactly
                    # like the reference timetable where Mon 11:50-1:50 shows
                    # Batch 1: English Comm Lab  AND  Batch 2: OOP Lab in one cell.
                    _seen_theory = set()
                    theory_subjects = []   # theory/tutorial – one per name+type
                    lab_groups = {}        # key=(name,type) -> list of subject records

                    for s in subjects:
                        if not (s.get('school', '').upper() == school_name.upper() and
                                (s.get('year') == year or s.get('semester') == year)):
                            continue
                        if not _subj_matches_section(s):
                            continue
                        _stype = str(s.get('type', '')).strip().upper()
                        _is_lab = _stype in ('LAB', 'PRACTICAL', 'TUTORIAL')
                        _dedup_key = (s.get('name', ''), s.get('type', ''))
                        if _is_lab:
                            lab_groups.setdefault(_dedup_key, []).append(s)
                        else:
                            if _dedup_key not in _seen_theory:
                                _seen_theory.add(_dedup_key)
                                theory_subjects.append(s)

                    available_slot_keys = [TimeSlotManager.get_slot_key(s) for s in lecture_slots]

                    # ---- Schedule theory/tutorial subjects (one slot each) ----
                    for subject in theory_subjects:
                        weekly_hours = subject.get('weekly_hours', 3)
                        sessions_needed = int(weekly_hours)
                        sessions_scheduled = 0

                        for _ in range(sessions_needed):
                            for attempt in range(100):
                                day = random.choice(self.days[:5])
                                slot_key = random.choice(available_slot_keys)

                                if batch_schedule[day].get(slot_key) is None:
                                    faculty = subject.get('faculty', 'TBD')

                                    # CHANGE 3: Check morning limit
                                    if '09:00' in slot_key:
                                        if not self.morning_constraint_manager.can_assign_morning(faculty):
                                            continue

                                    key = f"{day}_{slot_key}"
                                    if key not in faculty_tracker[faculty][key]:
                                        room_found = None

                                        if subject.get('assigned_room'):
                                            room_name = subject['assigned_room']
                                            if key not in room_tracker[room_name][key]:
                                                room_found = {'name': room_name}
                                            else:
                                                # STRICT: assigned room is occupied — skip this slot entirely
                                                continue
                                        else:
                                            for room in rooms:
                                                if key not in room_tracker[room['name']][key]:
                                                    room_found = room
                                                    break

                                        if room_found:
                                            slot_info = next(
                                                (s for s in time_slots
                                                 if TimeSlotManager.get_slot_key(s) == slot_key), None)

                                            batch_schedule[day][slot_key] = {
                                                'subject': subject['name'],
                                                'subject_code': subject.get('code', ''),
                                                'faculty': faculty,
                                                'room': room_found['name'],
                                                'room_locked': True,   # ROOM-FIX: GA path, always locked
                                                'type': subject.get('type', 'Theory'),
                                                'duration': slot_info['duration'] if slot_info else DEFAULT_LECTURE_DURATION,
                                                'start': slot_info['start'] if slot_info else '',
                                                'end': slot_info['end'] if slot_info else ''
                                            }

                                            faculty_tracker[faculty][key].append(batch_key)
                                            room_tracker[room_found['name']][key].append(batch_key)

                                            if '09:00' in slot_key:
                                                self.morning_constraint_manager.assign_morning(faculty)

                                            sessions_scheduled += 1
                                            break
                        if sessions_scheduled >= sessions_needed:
                                break

                    # FIXED: Parallel Labs Pairing (Different Subjects) 
                    # Instead of putting all batches of the same lab together, 
                    # we pair different lab subjects (e.g., Subj A and Subj B).
                    # Slot 1: Subj A (Batch 1) + Subj B (Batch 2)
                    # Slot 2: Subj B (Batch 1) + Subj A (Batch 2)
                    
                    lab_keys = sorted(lab_groups.keys())
                    lab_pool = []
                    for k in lab_keys:
                        # Find batches 1 and 2
                        b1 = next((s for s in lab_groups[k] if str(s.get('batch')) == '1'), None)
                        b2 = next((s for s in lab_groups[k] if str(s.get('batch')) == '2'), None)
                        if b1 and b2:
                            lab_pool.append({'name': k[0], 'type': k[1], 'sessions': max(1, int(b1.get('weekly_hours', 2))), 'b1': b1, 'b2': b2})
                        elif b1:
                            lab_pool.append({'name': k[0], 'type': k[1], 'sessions': max(1, int(b1.get('weekly_hours', 2))), 'solo': [b1]})
                        elif b2:
                            lab_pool.append({'name': k[0], 'type': k[1], 'sessions': max(1, int(b2.get('weekly_hours', 2))), 'solo': [b2]})
                        else:
                            lab_pool.append({'name': k[0], 'type': k[1], 'sessions': max(1, int(lab_groups[k][0].get('weekly_hours', 2))), 'solo': lab_groups[k]})

                    # Processing the pool in pairs to create parallel labs
                    idx = 0
                    while idx < len(lab_pool):
                        group1 = lab_pool[idx]
                        group2 = lab_pool[idx+1] if idx + 1 < len(lab_pool) else None
                        
                        sessions_to_schedule = group1['sessions'] if not group2 else max(group1['sessions'], group2['sessions'])
                        
                        for session_idx in range(sessions_to_schedule):
                            for attempt in range(250):
                                day = random.choice(self.days[:5])
                                slot_key = random.choice(available_slot_keys)
                                key = f"{day}_{slot_key}"
                                
                                if batch_schedule[day].get(slot_key) is not None:
                                    continue
                                
                                # Parallel pairing logic
                                if group2:
                                    if session_idx % 2 == 0:
                                        subjects_to_assign = [group1['b1'], group2['b2']]
                                    else:
                                        subjects_to_assign = [group1['b2'], group2['b1']]
                                else:
                                    subjects_to_assign = group1.get('solo', [group1.get('b1'), group1.get('b2')])
                                    subjects_to_assign = [s for s in subjects_to_assign if s]

                                # Plan placement
                                temp_entries = []
                                resources_planned = [] # list of (fac, room)
                                can_place = True
                                
                                for subj in subjects_to_assign:
                                    fac = subj.get('faculty', 'TBD')
                                    if key in faculty_tracker[fac][key]:
                                        can_place = False; break
                                    
                                    if '09:00' in slot_key and not self.morning_constraint_manager.can_assign_morning(fac):
                                        can_place = False; break

                                    # Room
                                    room_name = subj.get('assigned_room') or _subj_room_cache.get(subj.get('name'))
                                    if room_name:
                                        # Strict mode: If room is assigned, we MUST use it. No fallbacks.
                                        if key in room_tracker[room_name][key]:
                                            room_name = None # Occupied, fail to place in this slot.
                                    else:
                                        # No assigned room, fallback to any available lab room
                                        cand = [r['name'] for r in rooms if r.get('type', '').lower() in ('lab', 'practical')]
                                        random.shuffle(cand)
                                        room_name = next((r_name for r_name in cand if key not in room_tracker[r_name][key]), None)
                                    
                                    if not room_name:
                                        can_place = False; break
                                    
                                    # Ensure unique room and faculty for this slot planning
                                    if any(r == room_name for _, r in resources_planned):
                                        can_place = False; break
                                    if fac != 'TBD' and any(f == fac for f, _ in resources_planned):
                                        can_place = False; break
                                    
                                    resources_planned.append((fac, room_name))
                                    
                                    slot_info = next((s for s in time_slots if TimeSlotManager.get_slot_key(s) == slot_key), None)
                                    temp_entries.append({
                                        'subject': subj['name'],
                                        'subject_code': subj.get('code', ''),
                                        'faculty': fac,
                                        'room': room_name,
                                        'room_locked': True,   # ROOM-FIX: lab path, always locked
                                        'batch': str(subj.get('batch', '1')),
                                        'type': 'Lab',
                                        'duration': slot_info['duration'] if slot_info else 60,
                                        'start': slot_info['start'] if slot_info else '',
                                        'end': slot_info['end'] if slot_info else ''
                                    })
                                    resources_planned.append((fac, room_name))

                                if can_place and len(temp_entries) == len(subjects_to_assign):
                                    # Commit assignment
                                    batch_schedule[day][slot_key] = temp_entries if len(temp_entries) > 1 else temp_entries[0]
                                    for f, r in resources_planned:
                                        faculty_tracker[f][key].append(batch_key)
                                        room_tracker[r][key].append(batch_key)
                                        if '09:00' in slot_key:
                                            self.morning_constraint_manager.assign_morning(f)
                                    break # success for this session_idx
                        idx += 2 if group2 else 1

                    schedule[school_key][batch_key] = batch_schedule
         
        return schedule
        
        return schedule


# ==================== EXPORT UTILITIES - UPDATED FOR DYNAMIC SLOTS ====================

class ExportManager:
    """Handle exporting timetables to various formats - Updated for dynamic time slots"""
    
    @staticmethod
    def get_time_slots_from_schedule(schedule_data: dict) -> List[str]:
        """
        CHANGE 1: Extract time slot keys from schedule data
        Returns sorted list of time slot keys
        """
        slots = set()
        
        for day in schedule_data:
            if isinstance(schedule_data[day], dict):
                for slot_key in schedule_data[day].keys():
                    slots.add(slot_key)
        
        # Sort by start time
        sorted_slots = sorted(list(slots), key=lambda x: TimeSlotManager.time_to_minutes(x.split('-')[0]))
        return sorted_slots
    
    @staticmethod
    def format_slot_for_display(slot_key: str) -> str:
        """Format time slot key for display (e.g. 9:00 am to 10:00 am)"""
        parts = slot_key.split('-')
        if len(parts) == 2:
            return f"{parts[0].lower()} to {parts[1].lower()}"
        return slot_key.lower()
    
    @staticmethod
    def export_to_pdf(schedule_data, filename="timetable.pdf"):
        """Export timetable to PDF with dynamic time slots"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        
        buffer = io.BytesIO()
        page_width, page_height = landscape(A4)
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4),
            leftMargin=0.5*cm,
            rightMargin=0.5*cm,
            topMargin=0.5*cm,
            bottomMargin=0.5*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=6,
            leading=8,
            alignment=TA_CENTER,
        )
        
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            alignment=TA_CENTER,
            textColor=colors.whitesmoke,
        )
        
        # Add dynamic header rows
        curr_year = datetime.now().year
        ay_str = f"Time Table AY {curr_year}-{str(curr_year+1)[2:]} wef {datetime.now().strftime('%d.%m.%Y')}"
        
        header_title_style = ParagraphStyle(
            'HeaderTitleStyle',
            parent=styles['Heading1'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.whitesmoke,
            backColor=colors.black,
            borderPadding=2
        )
        elements.append(Paragraph(ay_str, header_title_style))
        elements.append(Spacer(1, 10))
        
        # CHANGE 1: Get dynamic time slots from schedule
        time_slots = ExportManager.get_time_slots_from_schedule(schedule_data)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        # Create header row
        header_row = [Paragraph('<b>Day</b>', header_style)]
        for slot in time_slots:
            display_slot = ExportManager.format_slot_for_display(slot)
            header_row.append(Paragraph(f'<b>{display_slot}</b>', header_style))
        
        data = [header_row]
        
        # Add data rows
        for day in days:
            row = [Paragraph(f'<b>{day}</b>', cell_style)]
            
            for slot in time_slots:
                cell_content = ""
                
                if day in schedule_data and slot in schedule_data[day]:
                    class_info = schedule_data[day][slot]
                    if class_info:
                        if isinstance(class_info, list):
                            parts = []
                            for entry in class_info:
                                ci = entry.get('ci', entry) if isinstance(entry, dict) else {}
                                sub = str(ci.get('subject', 'N/A'))[:15]
                                parts.append(sub)
                            cell_content = " + ".join(parts)
                        elif class_info.get('type') == 'LUNCH':
                            cell_content = "🍴 LUNCH"
                        elif class_info.get('type') == 'BREAK':
                            cell_content = "☕ BREAK"
                        else:
                            subject = str(class_info.get('subject', 'N/A'))[:18]
                            faculty = str(class_info.get('faculty', 'TBD'))[:12]
                            room = str(class_info.get('room', 'TBD'))[:8]
                            cell_content = f"{subject}<br/>{faculty}<br/>{room}"
                    else:
                        cell_content = "FREE"
                else:
                    cell_content = "FREE"
                
                row.append(Paragraph(cell_content, cell_style))
            
            data.append(row)
        
        # Calculate column widths
        available_width = page_width - 1*cm
        day_col_width = 1.5*cm
        slot_col_width = (available_width - day_col_width) / len(time_slots)
        
        col_widths = [day_col_width] + [slot_col_width] * len(time_slots)
        
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Apply styles
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90d9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#e8e8e8')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]
        
        table.setStyle(TableStyle(style_commands))
        
        # Highlight lunch and break rows
        for i, day in enumerate(days):
            row_idx = i + 1
            for j, slot in enumerate(time_slots):
                if day in schedule_data and slot in schedule_data[day]:
                    class_info = schedule_data[day][slot]
                    # Guard: list slots are combined lab/tutorial entries – never LUNCH/BREAK
                    if class_info and isinstance(class_info, dict):
                        if class_info.get('type') == 'LUNCH':
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (j+1, row_idx), (j+1, row_idx), colors.HexColor('#fff3cd')),
                            ]))
                        elif class_info.get('type') == 'BREAK':
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (j+1, row_idx), (j+1, row_idx), colors.HexColor('#dfe6e9')),
                            ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_to_pdf_detailed(schedule_data, school_name="", batch_name="", filename="timetable.pdf"):
        """Export timetable to PDF with title and additional details - Dynamic slots"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        buffer = io.BytesIO()
        page_width, page_height = landscape(A4)
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4),
            leftMargin=0.75*cm,
            rightMargin=0.75*cm,
            topMargin=0.75*cm,
            bottomMargin=0.75*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=5,
            textColor=colors.HexColor('#2c3e50')
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            spaceAfter=15,
            textColor=colors.HexColor('#7f8c8d')
        )
        
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=6,
            leading=8,
            alignment=TA_CENTER,
        )
        
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            alignment=TA_CENTER,
            textColor=colors.whitesmoke,
        )
        
        # Add dynamic header rows
        curr_year = datetime.now().year
        ay_str = f"Time Table AY {curr_year}-{str(curr_year+1)[2:]} wef {datetime.now().strftime('%d.%m.%Y')}"
        class_info_str = f"{school_name} {batch_name.replace('|', '').strip()}" if school_name and batch_name else "Weekly Timetable"

        header_title_style = ParagraphStyle(
            'HeaderTitleStyle',
            parent=styles['Heading1'],
            fontSize=12,
            alignment=TA_CENTER,
            textColor=colors.whitesmoke,
            backColor=colors.black,
            borderPadding=2
        )
        
        header_info_style = ParagraphStyle(
            'HeaderInfoStyle',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=TA_CENTER,
            textColor=colors.black,
            backColor=colors.HexColor('#ffc107'),
            borderPadding=4
        )
        
        elements.append(Paragraph(ay_str, header_title_style))
        elements.append(Paragraph(class_info_str, header_info_style))
        
        elements.append(Spacer(1, 10))
        
        # CHANGE 1: Get dynamic time slots
        time_slots = ExportManager.get_time_slots_from_schedule(schedule_data)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        if any('Saturday' in str(k) for k in schedule_data.keys()):
             days.append('Saturday')
        
        # Header row
        header_row = [Paragraph('<b>Day</b>', header_style)]
        for slot in time_slots:
            display_slot = ExportManager.format_slot_for_display(slot)
            header_row.append(Paragraph(f'<b>{display_slot}</b>', header_style))
        
        data = [header_row]
        
        # Data rows
        for day in days:
            row = [Paragraph(f'<b>{day}</b>', cell_style)]
            
            for slot in time_slots:
                cell_content = ""
                
                if day in schedule_data and slot in schedule_data[day]:
                    class_info = schedule_data[day][slot]
                    if class_info is None:
                        cell_content = "<font color='gray'>FREE</font>"
                    elif isinstance(class_info, list):
                        # CHANGE UNIVERSAL-BATCH-FIX-6: Combined lab/tutorial slot: stack each batch entry
                        parts = []
                        for _entry in class_info:
                            _ci = _entry.get('ci', _entry) if isinstance(_entry, dict) else {}
                            _bl = str(_entry.get('batch', _entry.get('_batch_label', '')) if isinstance(_entry, dict) else '').strip()
                            if _bl.endswith('.0'): _bl = _bl[:-2]
                            _subj = str(_ci.get('subject', 'N/A'))[:20]
                            _fac  = str(_ci.get('faculty', 'TBD'))[:15]
                            _room = str(_ci.get('room', 'TBD'))[:10]
                            parts.append(f"<b>{_subj} Batch {_bl}</b><br/>{_fac} {_room}")
                        cell_content = "<br/>" + "<br/>---<br/>".join(parts)
                    elif class_info.get('type') == 'LUNCH':
                        duration = class_info.get('duration', 50)
                        cell_content = f"<b>LUNCH</b><br/>({duration} min)"
                    elif class_info.get('type') == 'BREAK':
                        duration = class_info.get('duration', 10)
                        cell_content = f"<b>BREAK</b><br/>({duration} min)"
                    else:
                        subject = str(class_info.get('subject', 'N/A'))[:20]
                        faculty = str(class_info.get('faculty', 'TBD'))[:15]
                        room = str(class_info.get('room', 'TBD'))[:10]
                        class_type = str(class_info.get('type', ''))[:8]
                        
                        cell_content = f"<b>{subject}</b><br/>"
                        cell_content += f"{faculty}<br/>"
                        cell_content += f"{room}"
                        if class_type and class_type not in ['Theory', 'LUNCH', 'BREAK']:
                            cell_content += f"<br/><i>({class_type})</i>"
                else:
                    cell_content = "<font color='gray'>FREE</font>"
                
                row.append(Paragraph(cell_content, cell_style))
            
            data.append(row)
        
        # Calculate column widths
        available_width = page_width - 1.5*cm
        day_col_width = 1.8*cm
        slot_col_width = (available_width - day_col_width) / len(time_slots)
        
        col_widths = [day_col_width] + [slot_col_width] * len(time_slots)
        
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Table styling
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#2c3e50')),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]
        
        # Alternating colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                style_commands.append(('BACKGROUND', (1, i), (-1, i), colors.HexColor('#f8f9fa')))
        
        table.setStyle(TableStyle(style_commands))
        
        # Highlight lunch and breaks
        for i, day in enumerate(days):
            row_idx = i + 1
            for j, slot in enumerate(time_slots):
                if day in schedule_data and slot in schedule_data[day]:
                    class_info = schedule_data[day][slot]
                    # Guard: list slots are combined lab/tutorial entries – never LUNCH/BREAK
                    if class_info and isinstance(class_info, dict):
                        if class_info.get('type') == 'LUNCH':
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (j+1, row_idx), (j+1, row_idx), colors.HexColor('#ffeaa7')),
                            ]))
                        elif class_info.get('type') == 'BREAK':
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (j+1, row_idx), (j+1, row_idx), colors.HexColor('#dfe6e9')),
                            ]))
        
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 15))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#95a5a6')
        )
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", footer_style))
        
        doc.build(elements)
        
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_to_excel_formatted(schedule_data, school_name="", batch_name=""):
        """Export timetable to Excel with dynamic time slots and proper formatting"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        day_font = Font(bold=True, size=9)
        day_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        
        cell_font = Font(size=8)
        cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        lunch_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
        break_fill = PatternFill(start_color="dfe6e9", end_color="dfe6e9", fill_type="solid")
        free_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add dynamic header rows
        curr_year = datetime.now().year
        ay_str = f"Time Table AY {curr_year}-{str(curr_year+1)[2:]} wef {datetime.now().strftime('%d.%m.%Y')}"
        class_info_str = f"{school_name} {batch_name.replace('|', '').strip()}" if school_name and batch_name else "Weekly Timetable"

        # FIXED: Derive time_slots and days dynamically from schedule_data
        time_slots = ExportManager.get_time_slots_from_schedule(schedule_data)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

        # Academic Year Header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(time_slots) + 1)
        ay_cell = ws.cell(row=1, column=1, value=ay_str)
        ay_cell.font = Font(bold=True, size=11, color="FFFFFF")
        ay_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ay_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # Class Info Header
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(time_slots) + 1)
        info_cell = ws.cell(row=2, column=1, value=class_info_str)
        info_cell.font = Font(bold=True, size=14, color="000000")
        info_cell.fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
        info_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 35

        # Time slot labels in 'am to pm' format
        def _fmt_ts(sk):
            pts = sk.split('-')
            return f"{pts[0].lower()} to {pts[1].lower()}" if len(pts) == 2 else sk

        # Headers Row 3
        headers = ['Day'] + [_fmt_ts(slot) for slot in time_slots]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[3].height = 35
        
        # Data rows
        for row_idx, day in enumerate(days, 4):
            # Day column
            day_cell = ws.cell(row=row_idx, column=1, value=day)
            day_cell.font = day_font
            day_cell.fill = day_fill
            day_cell.alignment = cell_alignment
            day_cell.border = thin_border
            
            # Time slot columns
            for col_idx, slot in enumerate(time_slots, 2):
                cell_value = ""
                cell_fill = None
                
                if day in schedule_data and slot in schedule_data[day]:
                    class_info = schedule_data[day][slot]
                    if class_info is None:
                        cell_value = "FREE"
                        cell_fill = free_fill
                    elif isinstance(class_info, list):
                        # CHANGE UNIVERSAL-BATCH-FIX-6: Combined lab/tutorial slot: stack each batch entry as text
                        parts = []
                        for _entry in class_info:
                            _ci = _entry.get('ci', _entry) if isinstance(_entry, dict) else {}
                            _bl = str(_entry.get('batch', _entry.get('_batch_label', '')) if isinstance(_entry, dict) else '').strip()
                            if _bl.endswith('.0'): _bl = _bl[:-2]
                            _subj = str(_ci.get('subject', 'N/A'))
                            _fac  = str(_ci.get('faculty', 'TBD'))
                            _room = str(_ci.get('room', 'TBD'))
                            parts.append(f"🔬 {_subj} Batch {_bl}\n{_fac} {_room}")
                        cell_value = "\n".join(parts)
                        cell_fill = PatternFill(start_color="FFA726", end_color="FFA726", fill_type="solid")
                    elif class_info.get('type') == 'LUNCH':
                        duration = class_info.get('duration', 50)
                        cell_value = f"🍴 LUNCH BREAK\n({duration} min)"
                        cell_fill = lunch_fill
                    elif class_info.get('type') == 'BREAK':
                        duration = class_info.get('duration', 10)
                        cell_value = f"☕ BREAK\n({duration} min)"
                        cell_fill = break_fill
                    else:
                        subject = str(class_info.get('subject', 'N/A'))
                        faculty = str(class_info.get('faculty', 'TBD'))
                        room = str(class_info.get('room', 'TBD'))
                        class_type = str(class_info.get('type', ''))
                        cell_value = f"📚 {subject}\n👨‍🏫 {faculty}\n📍 {room}"
                        if class_type and class_type not in ['Theory', 'LUNCH', 'BREAK']:
                            cell_value += f"\n({class_type})"
                else:
                    cell_value = "FREE"
                    cell_fill = free_fill
                
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                if cell_fill:
                    cell.fill = cell_fill
            
            ws.row_dimensions[row_idx].height = 70
        
        # Set column widths
        ws.column_dimensions['A'].width = 12  # Day column
        for col in range(2, len(time_slots) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    @staticmethod
    def export_schedule_to_dataframe(schedule_data) -> pd.DataFrame:
        """
        CHANGE 1: Convert schedule to DataFrame with dynamic time slots
        """
        time_slots = ExportManager.get_time_slots_from_schedule(schedule_data)
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        data = []
        for day in days:
            row = {'Day': day}
            for slot in time_slots:
                display_slot = ExportManager.format_slot_for_display(slot)
                
                if day in schedule_data and slot in schedule_data[day]:
                    class_info = schedule_data[day][slot]
                    if class_info is None:
                        row[display_slot] = "FREE"
                    elif isinstance(class_info, list):
                        # CHANGE UNIVERSAL-BATCH-FIX-6: Strip .0 from batch labels in DataFrame export
                        parts = []
                        for _e in class_info:
                            _ci = _e.get('ci', _e) if isinstance(_e, dict) else {}
                            _bl = str(_e.get('batch', _e.get('_batch_label', '')) if isinstance(_e, dict) else '').strip()
                            if _bl.endswith('.0'): _bl = _bl[:-2]
                            parts.append(f"{_ci.get('subject','N/A')} Batch {_bl} {_ci.get('faculty','TBD')} {_ci.get('room','TBD')}")
                        row[display_slot] = " / ".join(parts)
                    elif class_info.get('type') == 'LUNCH':
                        duration = class_info.get('duration', 50)
                        row[display_slot] = f"🍴 LUNCH ({duration} min)"
                    elif class_info.get('type') == 'BREAK':
                        duration = class_info.get('duration', 10)
                        row[display_slot] = f"☕ BREAK ({duration} min)"
                    else:
                        subject = class_info.get('subject', 'N/A')
                        faculty = class_info.get('faculty', 'TBD')
                        room = class_info.get('room', 'TBD')
                        row[display_slot] = f"📚 {subject}\n👨‍🏫 {faculty}\n📍 {room}"
                else:
                    row[display_slot] = "FREE"
            
            data.append(row)
        
        return pd.DataFrame(data)
    
# app.py - Part 4: Report Generator, Admin Dashboard, Main Application
# Continuation from Part 3

# ==================== REPORT GENERATOR ====================

class ReportGenerator:
    """Generate professional reports for administration"""
    
    def __init__(self, firebase_manager):
        self.firebase = firebase_manager
    
    def generate_faculty_workload_report(self) -> pd.DataFrame:
        """Generate faculty workload analysis"""
        timetables = self.firebase.get_all_timetables()
        
        workload_data = defaultdict(lambda: {
            'total_hours': 0,
            'theory_hours': 0,
            'lab_hours': 0,
            'tutorial_hours': 0,
            'subjects': set(),
            'programs': set(),
            'morning_slots': 0  # CHANGE 3: Track morning slots
        })
        
        for timetable in timetables:
            schedule = timetable.get('schedule', {})
            for school in schedule:
                for batch in schedule[school]:
                    for day in schedule[school][batch]:
                        for slot, class_info in schedule[school][batch][day].items():
                            # Safely handle combined lab/tutorial list slots
                            entries = class_info if isinstance(class_info, list) else ([class_info] if class_info else [])
                            for class_info in entries:
                                if not isinstance(class_info, dict):
                                    continue
                                if class_info.get('faculty') and class_info.get('type') not in ['LUNCH', 'BREAK']:
                                    faculty = class_info['faculty']
                                    workload_data[faculty]['total_hours'] += 1
                                    workload_data[faculty]['subjects'].add(class_info.get('subject', ''))
                                    workload_data[faculty]['programs'].add(school)
                                    
                                    # CHANGE 3: Track morning slots
                                    if '09:00' in slot:
                                        workload_data[faculty]['morning_slots'] += 1
                                    
                                    class_type = class_info.get('type', 'Theory').lower()
                                    if 'lab' in class_type:
                                        workload_data[faculty]['lab_hours'] += 1
                                    elif 'tutorial' in class_type:
                                        workload_data[faculty]['tutorial_hours'] += 1
                                    else:
                                        workload_data[faculty]['theory_hours'] += 1
        
        # Convert to DataFrame
        report_data = []
        for faculty, data in workload_data.items():
            report_data.append({
                'Faculty Name': faculty,
                'Total Hours/Week': data['total_hours'],
                'Theory Hours': data['theory_hours'],
                'Lab Hours': data['lab_hours'],
                'Tutorial Hours': data['tutorial_hours'],
                'Morning Slots (9AM)': data['morning_slots'],  # CHANGE 3
                'Subjects Count': len(data['subjects']),
                'Programs': ', '.join(data['programs']),
                'Workload Status': 'Overloaded' if data['total_hours'] > 20 else ('Optimal' if data['total_hours'] >= 15 else 'Underloaded'),
                'Morning Limit Status': '⚠️ At Limit' if data['morning_slots'] >= FACULTY_MORNING_LIMIT else '✅ OK'  # CHANGE 3
            })
        
        return pd.DataFrame(report_data)
    
    def generate_room_utilization_report(self) -> pd.DataFrame:
        """Generate room utilization analysis"""
        timetables = self.firebase.get_all_timetables()
        
        total_available_slots = 5 * 6  # 5 days * ~6 usable slots
        room_usage = defaultdict(lambda: {'used_slots': 0, 'programs': set()})
        
        for timetable in timetables:
            schedule = timetable.get('schedule', {})
            for school in schedule:
                for batch in schedule[school]:
                    for day in schedule[school][batch]:
                        for slot, class_info in schedule[school][batch][day].items():
                            # Safely handle combined lab/tutorial list slots
                            entries = class_info if isinstance(class_info, list) else ([class_info] if class_info else [])
                            for class_info in entries:
                                if not isinstance(class_info, dict):
                                    continue
                                if class_info.get('room') and class_info.get('type') not in ['LUNCH', 'BREAK']:
                                    room = class_info['room']
                                    if room not in ['TBD', 'Cafeteria', '']:
                                        room_usage[room]['used_slots'] += 1
                                        room_usage[room]['programs'].add(school)
        
        report_data = []
        for room, data in room_usage.items():
            utilization = (data['used_slots'] / total_available_slots) * 100
            report_data.append({
                'Room': room,
                'Used Slots/Week': data['used_slots'],
                'Available Slots': total_available_slots,
                'Utilization %': f"{utilization:.1f}%",
                'Programs Using': ', '.join(data['programs']),
                'Status': 'High' if utilization > 70 else ('Medium' if utilization > 40 else 'Low')
            })
        
        return pd.DataFrame(report_data)
    
    def generate_program_summary_report(self) -> pd.DataFrame:
        """Generate program-wise summary report"""
        timetables = self.firebase.get_all_timetables()
        
        program_data = defaultdict(lambda: {
            'total_classes': 0,
            'theory_classes': 0,
            'lab_classes': 0,
            'tutorial_classes': 0,
            'faculty_count': set(),
            'rooms_used': set(),
            'batches': set()
        })
        
        for timetable in timetables:
            schedule = timetable.get('schedule', {})
            for school in schedule:
                for batch in schedule[school]:
                    program_data[school]['batches'].add(batch)
                    for day in schedule[school][batch]:
                        for slot, class_info in schedule[school][batch][day].items():
                            # Safely handle combined lab/tutorial list slots
                            entries = class_info if isinstance(class_info, list) else ([class_info] if class_info else [])
                            for ci in entries:
                                if not isinstance(ci, dict):
                                    continue
                                if ci.get('type') in ['LUNCH', 'BREAK', None]:
                                    continue
                                program_data[school]['total_classes'] += 1
                                
                                if ci.get('faculty'):
                                    program_data[school]['faculty_count'].add(ci['faculty'])
                                
                                if ci.get('room') and ci['room'] not in ['TBD', 'Cafeteria']:
                                    program_data[school]['rooms_used'].add(ci['room'])
                                
                                class_type = ci.get('type', 'Theory').lower()
                                if 'lab' in class_type:
                                    program_data[school]['lab_classes'] += 1
                                elif 'tutorial' in class_type:
                                    program_data[school]['tutorial_classes'] += 1
                                else:
                                    program_data[school]['theory_classes'] += 1
        
        report_data = []
        for program, data in program_data.items():
            report_data.append({
                'Program': program,
                'Total Batches': len(data['batches']),
                'Total Classes/Week': data['total_classes'],
                'Theory Classes': data['theory_classes'],
                'Lab Classes': data['lab_classes'],
                'Tutorial Classes': data['tutorial_classes'],
                'Faculty Engaged': len(data['faculty_count']),
                'Rooms Used': len(data['rooms_used'])
            })
        
        return pd.DataFrame(report_data)
    
    def generate_clash_history_report(self) -> pd.DataFrame:
        """Generate clash history report"""
        clashes = self.firebase.get_unresolved_clashes()
        
        report_data = []
        for clash in clashes:
            report_data.append({
                'Clash Type': clash.get('type', 'Unknown'),
                'Severity': clash.get('severity', 'Unknown'),
                'Details': clash.get('details', 'No details'),
                'Time': clash.get('time', 'Unknown'),
                'Faculty': clash.get('faculty', 'N/A'),
                'Room': clash.get('room', 'N/A'),
                'Status': 'Resolved' if clash.get('resolved', False) else 'Unresolved',
                'Detected At': clash.get('detected_at', 'Unknown')
            })
        
        if not report_data:
            report_data.append({
                'Clash Type': 'No clashes',
                'Severity': '-',
                'Details': 'No clashes detected in the system',
                'Time': '-',
                'Faculty': '-',
                'Room': '-',
                'Status': '-',
                'Detected At': '-'
            })
        
        return pd.DataFrame(report_data)
    
    def generate_agent_repair_history_report(self) -> pd.DataFrame:
        """Generate report of agent repair actions from Firebase repair_history."""
        history = self.firebase.get_repair_history(limit=200)
        if not history:
            return pd.DataFrame(
                [{
                    'Repair ID': 'No data',
                    'Session ID': '-',
                    'Action Type': '-',
                    'Clash Type': '-',
                    'Faculty/Room': '-',
                    'From Slot': '-',
                    'To Slot': '-',
                    'Reason': 'No agent repair history found',
                    'Success': '-',
                    'Timestamp': '-',
                }]
            )

        rows = []
        for entry in history:
            rows.append({
                'Repair ID': entry.get('repair_id', entry.get('id', '')),
                'Session ID': entry.get('session_id', ''),
                'Action Type': entry.get('action_type', ''),
                'Clash Type': entry.get('clash_type', ''),
                'Faculty/Room': entry.get('faculty_or_room', ''),
                'From Slot': json.dumps(entry.get('from_slot', {})),
                'To Slot': json.dumps(entry.get('to_slot', {})),
                'Reason': entry.get('reason', ''),
                'Success': entry.get('success', ''),
                'Timestamp': str(entry.get('timestamp', entry.get('timestamp_server', ''))),
            })
        return pd.DataFrame(rows)

    # CHANGE 1, 2: New report for semester configurations
    def generate_semester_config_report(self) -> pd.DataFrame:
        """Generate report of all semester configurations"""
        lunch_configs = self.firebase.get_all_lunch_configs()
        break_configs = self.firebase.get_all_break_configs()
        
        # Combine configs
        config_dict = defaultdict(lambda: {'lunch': None, 'breaks': None})
        
        for config in lunch_configs:
            key = f"{config.get('program', 'N/A')}_Sem{config.get('semester', 'N/A')}"
            config_dict[key]['lunch'] = config
        
        for config in break_configs:
            key = f"{config.get('program', 'N/A')}_Sem{config.get('semester', 'N/A')}"
            config_dict[key]['breaks'] = config
        
        report_data = []
        for key, data in config_dict.items():
            lunch = data['lunch'] or {}
            breaks = data['breaks'] or {}
            
            report_data.append({
                'Program/Semester': key,
                'Custom Lunch': 'Yes' if lunch.get('custom', False) else 'No',
                'Lunch Time': f"{lunch.get('start', 'N/A')} - {lunch.get('end', 'N/A')}",
                'Lunch Duration': f"{lunch.get('duration', 50)} min",
                'Lunch Locked': '🔒' if lunch.get('locked', False) else '🔓',
                'Breaks Enabled': 'Yes' if breaks.get('enabled', False) else 'No',
                'Break Duration': f"{breaks.get('duration', 0)} min" if breaks.get('enabled') else '-',
                'Break Placements': str(breaks.get('placements', [])) if breaks.get('enabled') else '-'
            })
        
        if not report_data:
            report_data.append({
                'Program/Semester': 'No configs',
                'Custom Lunch': '-',
                'Lunch Time': '-',
                'Lunch Duration': '-',
                'Lunch Locked': '-',
                'Breaks Enabled': '-',
                'Break Duration': '-',
                'Break Placements': '-'
            })
        
        return pd.DataFrame(report_data)
    
    def export_faculty_report_to_excel(self) -> bytes:
        """Export faculty workload report to Excel"""
        df = self.generate_faculty_workload_report()
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Faculty Workload', index=False)
            
            worksheet = writer.sheets['Faculty Workload']
            for idx, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
        
        output.seek(0)
        return output.getvalue()
    
    def export_room_report_to_excel(self) -> bytes:
        """Export room utilization report to Excel"""
        df = self.generate_room_utilization_report()
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Room Utilization', index=False)
            
            worksheet = writer.sheets['Room Utilization']
            for idx, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
        
        output.seek(0)
        return output.getvalue()
    
    def export_comprehensive_report_to_excel(self) -> bytes:
        """Export all reports to a single Excel file with multiple sheets"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Faculty Workload Sheet
            faculty_df = self.generate_faculty_workload_report()
            faculty_df.to_excel(writer, sheet_name='Faculty Workload', index=False)
            
            # Room Utilization Sheet
            room_df = self.generate_room_utilization_report()
            room_df.to_excel(writer, sheet_name='Room Utilization', index=False)
            
            # Program Summary Sheet
            program_df = self.generate_program_summary_report()
            program_df.to_excel(writer, sheet_name='Program Summary', index=False)
            
            # Clash History Sheet
            clash_df = self.generate_clash_history_report()
            clash_df.to_excel(writer, sheet_name='Clash History', index=False)
            
            # CHANGE 1, 2: Semester Configs Sheet
            config_df = self.generate_semester_config_report()
            config_df.to_excel(writer, sheet_name='Semester Configs', index=False)
            
            # Auto-adjust column widths for all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value) or "") for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        output.seek(0)
        return output.getvalue()


# ==================== ADMIN DASHBOARD ====================

def show_admin_dashboard(firebase_mgr):
    """Show comprehensive admin dashboard with analytics"""
    
    st.markdown("## 📊 Admin Dashboard")
    st.markdown("Real-time overview of your timetable management system")
    
    if not firebase_mgr:
        st.error("❌ Firebase not connected. Dashboard requires Firebase connection.")
        return
    
    # Initialize Report Generator
    report_gen = ReportGenerator(firebase_mgr)
    
    # ==================== KEY METRICS ====================
    st.markdown("### 🎯 Key Metrics")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    # Get data for metrics
    all_timetables = firebase_mgr.get_all_timetables()
    all_clashes = firebase_mgr.get_unresolved_clashes()
    all_rooms = firebase_mgr.get_rooms_list()
    all_info_datasets = firebase_mgr.get_info_dataset()
    
    # CHANGE 1, 2: Get config counts
    all_lunch_configs = firebase_mgr.get_all_lunch_configs()
    all_break_configs = firebase_mgr.get_all_break_configs()
    
    # Calculate faculty count from info datasets
    faculty_set = set()
    subjects_count = 0
    if all_info_datasets:
        for dataset in all_info_datasets:
            if 'data' in dataset:
                for record in dataset['data']:
                    subjects_count += 1
                    if record.get('Faculty') and record.get('Faculty') != 'TBD':
                        faculty_set.add(record['Faculty'])
    
    with col1:
        st.metric(
            "📅 Total Timetables", 
            len(all_timetables),
            delta=f"{len(all_timetables)} active" if all_timetables else None
        )
    
    with col2:
        st.metric(
            "📚 Active Programs", 
            len(PROGRAM_CONFIG),
            delta=None
        )
    
    with col3:
        clash_count = len(all_clashes)
        st.metric(
            "⚠️ Unresolved Clashes", 
            clash_count,
            delta="All clear!" if clash_count == 0 else f"{clash_count} need attention",
            delta_color="inverse" if clash_count > 0 else "normal"
        )
    
    with col4:
        st.metric(
            "👨‍🏫 Total Faculty", 
            len(faculty_set),
            delta=None
        )
    
    with col5:
        st.metric(
            "🏢 Total Rooms", 
            len(all_rooms),
            delta=None
        )
    
    # Second row of metrics
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.metric(
            "📖 Total Subjects",
            subjects_count,
            delta=None
        )
    
    with col2:
        info_count = len(all_info_datasets) if all_info_datasets else 0
        st.metric(
            "📤 Info Datasets",
            info_count,
            delta=None
        )
    
    with col3:
        room_datasets = firebase_mgr.get_room_dataset()
        room_dataset_count = len(room_datasets) if room_datasets else 0
        st.metric(
            "🏠 Room Datasets",
            room_dataset_count,
            delta=None
        )
    
    # CHANGE 1, 2: Show config counts
    with col4:
        custom_lunch_count = sum(1 for c in all_lunch_configs if c.get('custom', False))
        st.metric(
            "🍴 Custom Lunch Configs",
            custom_lunch_count,
            delta=None
        )
    
    with col5:
        enabled_break_count = sum(1 for c in all_break_configs if c.get('enabled', False))
        st.metric(
            "☕ Break Configs",
            enabled_break_count,
            delta=None
        )
    
    st.markdown("---")
    
    # CHANGE 3: Faculty Morning Constraint Summary
    st.markdown("### 🌅 Faculty Morning Constraint Status")
    st.markdown(f'<span class="morning-limit-badge">Max {FACULTY_MORNING_LIMIT} lectures at 9AM per faculty per week</span>', unsafe_allow_html=True)
    
    # Get morning counts
    faculty_morning_counts = firebase_mgr.get_faculty_morning_counts()
    if faculty_morning_counts:
        at_limit = [f for f, c in faculty_morning_counts.items() if c >= FACULTY_MORNING_LIMIT]
        if at_limit:
            st.warning(f"⚠️ {len(at_limit)} faculty at morning limit: {', '.join(at_limit[:5])}{'...' if len(at_limit) > 5 else ''}")
        else:
            st.success("✅ All faculty within morning slot limits")
    else:
        st.info("No morning constraint data available yet")
    
    st.markdown("---")
    
    # ==================== CHARTS SECTION ====================
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📈 Program Distribution")
        
        program_classes = defaultdict(int)
        for timetable in all_timetables:
            schedule = timetable.get('schedule', {})
            for school in schedule:
                for batch in schedule[school]:
                    for day in schedule[school][batch]:
                        for slot, slot_val in schedule[school][batch][day].items():
                            # slot_val may be a list (combined lab slots) or a plain dict
                            _slot_items = slot_val if isinstance(slot_val, list) else [slot_val]
                            for class_info in _slot_items:
                                if (class_info and isinstance(class_info, dict)
                                        and class_info.get('type') not in ['LUNCH', 'BREAK', None]):
                                    program_classes[school] += 1
        
        if program_classes:
            fig = px.pie(
                values=list(program_classes.values()),
                names=list(program_classes.keys()),
                title="Classes Distribution by Program",
                color_discrete_sequence=px.colors.qualitative.Set3
            )
            fig.update_traces(textposition='inside', textinfo='percent+label')
            fig.update_layout(height=350)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No timetable data available for chart.")
    
    with col2:
        st.markdown("### 🏢 Room Utilization Overview")
        
        room_usage = defaultdict(int)
        
        for timetable in all_timetables:
            schedule = timetable.get('schedule', {})
            for school in schedule:
                for batch in schedule[school]:
                    for day in schedule[school][batch]:
                        for slot, slot_val in schedule[school][batch][day].items():
                            _slot_items = slot_val if isinstance(slot_val, list) else [slot_val]
                            for class_info in _slot_items:
                                if (class_info and isinstance(class_info, dict)
                                        and class_info.get('room')
                                        and class_info.get('type') not in ['LUNCH', 'BREAK']):
                                    room = class_info['room']
                                    if room not in ['TBD', 'Cafeteria', '']:
                                        room_usage[room] += 1
        
        if room_usage:
            sorted_rooms = sorted(room_usage.items(), key=lambda x: x[1], reverse=True)[:10]
            rooms = [r[0] for r in sorted_rooms]
            usage = [r[1] for r in sorted_rooms]
            
            fig = px.bar(
                x=rooms,
                y=usage,
                title="Top 10 Most Used Rooms",
                labels={'x': 'Room', 'y': 'Classes/Week'},
                color=usage,
                color_continuous_scale='Blues'
            )
            fig.update_layout(height=350, xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No room usage data available for chart.")
    
    st.markdown("---")
    
    # ==================== QUICK ACTIONS ====================
    st.markdown("### ⚡ Quick Actions")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("📤 Upload New Dataset", use_container_width=True):
            st.session_state.dashboard_action = 'upload'
            st.info("👆 Go to 'Dataset Upload' tab to upload new data")
    
    with col2:
        if st.button("🚀 Generate Timetable", use_container_width=True):
            st.session_state.dashboard_action = 'generate'
            st.info("👆 Go to 'Generate Timetable' tab to create new timetable")
    
    with col3:
        if st.button("📊 View Reports", use_container_width=True):
            st.session_state.dashboard_action = 'reports'
            st.info("👆 Go to 'Reports & Analytics' tab for detailed reports")
    
    with col4:
        if st.button("🔄 Refresh Dashboard", use_container_width=True):
            st.rerun()
    
    # ==================== SYSTEM INFORMATION ====================
    with st.expander("ℹ️ System Information", expanded=False):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("**Available Programs:**")
            for program, config in PROGRAM_CONFIG.items():
                st.write(f"• {program}: {config['semesters']} semesters")
        
        with col2:
            st.markdown("**Default Lunch Settings:**")
            st.write(f"• Duration: {DEFAULT_LUNCH_DURATION} min")
            st.write(f"• STME: {DEFAULT_LUNCH_START_TIMES.get('STME', '13:00')}")
            st.write(f"• SOC: {DEFAULT_LUNCH_START_TIMES.get('SOC', '11:00')}")
            st.write(f"• SOL: {DEFAULT_LUNCH_START_TIMES.get('SOL', '12:00')}")
        
        with col3:
            st.markdown("**System Status:**")
            st.write(f"• Firebase: {'🟢 Connected' if firebase_mgr else '🔴 Disconnected'}")
            st.write(f"• Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            st.write(f"• Morning Limit: {FACULTY_MORNING_LIMIT}/faculty/week")
            st.write(f"• Version: 2.1 (Dynamic Slots)")


# ==================== HELPER FUNCTIONS ====================

def get_faculty_primary_school(faculty_name, faculties_list):
    """Determine the primary school for a faculty based on their department"""
    for faculty in faculties_list:
        if faculty['name'] == faculty_name:
            dept = faculty.get('department', '')
            
            if dept == 'STME' or 'Computer' in dept or 'Engineering' in dept or 'Technology' in dept:
                return 'STME'
            elif dept == 'SOC' or 'Commerce' in dept or 'Business' in dept:
                return 'SOC'
            elif dept == 'SOL' or 'Law' in dept:
                return 'SOL'
    
    return 'STME'


def display_faculty_timetable(faculty_name, faculty_schedule, faculty_metadata, tab_id=""):
    """Display a faculty member's complete timetable with details"""
    
    key_prefix = f"{tab_id}_{faculty_name.replace(' ', '_').replace('.', '')}"
    
    with st.expander(f"📘 {faculty_name}'s Timetable", expanded=True):
        
        # Faculty Info Header
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            schools = faculty_metadata.get('schools', set())
            school_list = ', '.join([s.split('_')[0] if '_' in s else s for s in schools])
            st.info(f"**Schools:** {school_list}")
        
        with col2:
            subjects = faculty_metadata.get('subjects', set())
            st.info(f"**Subjects:** {len(subjects)}")
        
        with col3:
            total_hours = faculty_metadata.get('total_hours', 0)
            status = "🔴 Overloaded" if total_hours > 20 else ("🟢 Optimal" if total_hours >= 15 else "🟡 Underloaded")
            st.info(f"**Weekly Hours:** {total_hours} ({status})")
        
        # CHANGE 3: Show morning slot count
        with col4:
            morning_count = faculty_metadata.get('morning_slots', 0)
            morning_status = "⚠️ At Limit" if morning_count >= FACULTY_MORNING_LIMIT else "✅ OK"
            st.info(f"**9AM Slots:** {morning_count}/{FACULTY_MORNING_LIMIT} ({morning_status})")
        
        # Get time slots from schedule
        all_slots = set()
        for day, day_schedule in faculty_schedule.items():
            all_slots.update(day_schedule.keys())
        
        time_slots = sorted(list(all_slots), key=lambda x: TimeSlotManager.time_to_minutes(x.split('-')[0]) if '-' in x else 0)
        
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        
        # View type selector
        view_type = st.radio(
            "View Type", 
            ["📅 Week View", "📆 Day View", "📊 Summary"], 
            horizontal=True, 
            key=f"view_{key_prefix}"
        )
        
        if view_type == "📅 Week View":
            # Weekly Timetable View
            timetable_data = []
            teaching_hours = 0
            
            for day in days:
                row = {'Day': day}
                for slot in time_slots:
                    display_slot = ExportManager.format_slot_for_display(slot) if '-' in slot else slot
                    
                    if day in faculty_schedule and slot in faculty_schedule[day]:
                        slot_val = faculty_schedule[day][slot]
                        if slot_val:
                            # Handle both list (clashes/combined) and single dict
                            items = slot_val if isinstance(slot_val, list) else [slot_val]
                            
                            is_lunch = any(i.get('type') == 'LUNCH' for i in items if isinstance(i, dict))
                            is_break = any(i.get('type') == 'BREAK' for i in items if isinstance(i, dict))
                            
                            if is_lunch:
                                row[display_slot] = "🍴 LUNCH"
                            elif is_break:
                                row[display_slot] = "☕ BREAK"
                            else:
                                subjs = []
                                rooms = []
                                batches = []
                                for class_info in items:
                                    if isinstance(class_info, dict):
                                        subjs.append(class_info.get('subject', 'N/A')[:15])
                                        rooms.append(class_info.get('room', 'TBD')[:10])
                                        b = str(class_info.get('batch', ''))
                                        batches.append(b.split('_')[-1] if '_' in b else b)
                                
                                # Join unique values
                                row[display_slot] = f"📚 {' & '.join(sorted(list(set(subjs))))}\n📍 {'/'.join(sorted(list(set(rooms))))}\n👥 {' & '.join(sorted(list(set(batches))))}"
                                teaching_hours += 1
                        else:
                            row[display_slot] = "FREE"
                    else:
                        row[display_slot] = "FREE"
                timetable_data.append(row)
            
            # Display metrics
            metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
            with metric_col1:
                st.metric("Total Classes", teaching_hours)
            with metric_col2:
                avg_daily = teaching_hours / 5 if teaching_hours > 0 else 0
                st.metric("Avg Daily", f"{avg_daily:.1f}")
            with metric_col3:
                free_slots = (5 * len([s for s in time_slots if 'LUNCH' not in s and 'BREAK' not in s])) - teaching_hours
                st.metric("Free Slots", max(0, free_slots))
            with metric_col4:
                unique_subjects = len(faculty_metadata.get('subjects', set()))
                st.metric("Subjects", unique_subjects)
            
            # Display timetable
            df = pd.DataFrame(timetable_data)
            st.dataframe(df, use_container_width=True, height=350)
            
            # Export options
            st.markdown("##### 📥 Export Options")
            exp_col1, exp_col2 = st.columns(2)
            
            with exp_col1:
                csv = df.to_csv(index=False)
                st.download_button(
                    label="📥 Download CSV",
                    data=csv,
                    file_name=f"faculty_{faculty_name.replace(' ', '_')}_timetable.csv",
                    mime="text/csv",
                    key=f"csv_{key_prefix}"
                )
            
            with exp_col2:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Timetable', index=False)
                output.seek(0)
                st.download_button(
                    label="📥 Download Excel",
                    data=output,
                    file_name=f"faculty_{faculty_name.replace(' ', '_')}_timetable.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"excel_{key_prefix}"
                )
        
        elif view_type == "📆 Day View":
            selected_day = st.selectbox(
                "Select Day", 
                days, 
                key=f"day_{key_prefix}"
            )
            
            st.markdown(f"#### {selected_day}'s Schedule")
            
            day_data = []
            for slot in time_slots:
                display_slot = ExportManager.format_slot_for_display(slot) if '-' in slot else slot
                
                if selected_day in faculty_schedule and slot in faculty_schedule[selected_day]:
                    slot_val = faculty_schedule[selected_day][slot]
                    if slot_val:
                        items = slot_val if isinstance(slot_val, list) else [slot_val]
                        for class_info in items:
                            if not isinstance(class_info, dict): continue
                            
                            if class_info.get('type') == 'LUNCH':
                                day_data.append({
                                    "Time": display_slot,
                                    "Subject": "🍴 LUNCH BREAK",
                                    "Room": "Cafeteria",
                                    "Class": "-",
                                    "Type": "LUNCH"
                                })
                            elif class_info.get('type') == 'BREAK':
                                day_data.append({
                                    "Time": display_slot,
                                    "Subject": "☕ BREAK",
                                    "Room": "-",
                                    "Class": "-",
                                    "Type": "BREAK"
                                })
                            else:
                                day_data.append({
                                    "Time": display_slot,
                                    "Subject": class_info.get('subject', 'N/A'),
                                    "Room": class_info.get('room', 'TBD'),
                                    "Class": class_info.get('batch', 'N/A'),
                                    "Type": class_info.get('type', 'Theory')
                                })
                    else:
                        day_data.append({
                            "Time": display_slot,
                            "Subject": "FREE",
                            "Room": "-",
                            "Class": "-",
                            "Type": "-"
                        })
                else:
                    day_data.append({
                        "Time": display_slot,
                        "Subject": "FREE",
                        "Room": "-",
                        "Class": "-",
                        "Type": "-"
                    })
            
            df_day = pd.DataFrame(day_data)
            st.dataframe(df_day, use_container_width=True, hide_index=True)
            
            teaching_count = sum(1 for d in day_data if d['Subject'] not in ['FREE', '🍴 LUNCH BREAK', '☕ BREAK'])
            st.caption(f"📊 Teaching {teaching_count} classes on {selected_day}")
        
        elif view_type == "📊 Summary":
            st.markdown("#### 📊 Teaching Summary")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("##### 📚 Subjects Teaching")
                subjects = faculty_metadata.get('subjects', set())
                for subject in sorted(subjects):
                    if subject:
                        st.write(f"  • {subject}")
            
            with col2:
                st.markdown("##### 🏫 Schools & Programs")
                schools = faculty_metadata.get('schools', set())
                for school in sorted(schools):
                    st.write(f"  • {school}")
            
            st.markdown("##### ⏱️ Workload Analysis")
            total_hours = faculty_metadata.get('total_hours', 0)
            
            if total_hours > 20:
                st.error(f"⚠️ Faculty is OVERLOADED with {total_hours} hours/week (Max recommended: 20)")
            elif total_hours >= 15:
                st.success(f"✅ Faculty has OPTIMAL workload: {total_hours} hours/week")
            else:
                st.warning(f"⚠️ Faculty is UNDERLOADED with only {total_hours} hours/week (Min recommended: 15)")
            
            # CHANGE 3: Morning constraint status
            morning_count = faculty_metadata.get('morning_slots', 0)
            if morning_count >= FACULTY_MORNING_LIMIT:
                st.warning(f"⚠️ Faculty at MORNING LIMIT: {morning_count} 9AM lectures (Max: {FACULTY_MORNING_LIMIT})")
            else:
                st.success(f"✅ Morning slots OK: {morning_count}/{FACULTY_MORNING_LIMIT} used")


# ==================== MAIN APPLICATION ====================

def main():
    logger.info("Smart Timetable Scheduler starting")
    # Show Firebase connection status
    if firebase_manager:
        st.markdown('<div class="firebase-status firebase-connected">🔥 Firebase Connected</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="firebase-status firebase-disconnected">⚠️ Firebase Disconnected</div>', unsafe_allow_html=True)
    
    # Initialize session state
    if 'portal' not in st.session_state:
        st.session_state.portal = None
    if 'schools_data' not in st.session_state:
        st.session_state.schools_data = {}
    if 'faculties' not in st.session_state:
        st.session_state.faculties = []
    if 'subjects' not in st.session_state:
        st.session_state.subjects = []
    if 'rooms' not in st.session_state:
        st.session_state.rooms = []
    if 'info_dataset' not in st.session_state:
        st.session_state.info_dataset = []
    if 'room_dataset' not in st.session_state:
        st.session_state.room_dataset = []
    if 'room_allocations' not in st.session_state:
        st.session_state.room_allocations = {}
    if 'generated_schedules' not in st.session_state:
        st.session_state.generated_schedules = {}
    if 'current_schedule' not in st.session_state:
        st.session_state.current_schedule = None
    if 'current_semester_config' not in st.session_state:
        st.session_state.current_semester_config = None
    if 'edit_mode' not in st.session_state:
        st.session_state.edit_mode = False
    if 'editor' not in st.session_state:
        st.session_state.editor = TimetableEditor(firebase_manager)
    if 'edited_schedule' not in st.session_state:
        st.session_state.edited_schedule = None
    if 'detected_clashes' not in st.session_state:
        st.session_state.detected_clashes = []
    if 'selected_program' not in st.session_state:
        st.session_state.selected_program = None
    if 'selected_semester' not in st.session_state:
        st.session_state.selected_semester = 1
    if 'section_batch_schedules' not in st.session_state:
        st.session_state.section_batch_schedules = {}
    if 'best_schedule' not in st.session_state:
        st.session_state.best_schedule = None
    
    # Portal Selection Page
    if st.session_state.portal is None:
        st.markdown('<h1 class="main-header">🎓 Smart Classroom & Timetable Scheduler</h1>', unsafe_allow_html=True)
        
        st.markdown("### Select Your Portal")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("👨‍💼 Admin Portal", use_container_width=True, type="primary", key="admin_btn"):
                st.session_state.portal = 'admin'
                st.rerun()
        
        with col2:
            if st.button("👨‍🏫 Faculty Portal", use_container_width=True, type="primary", key="faculty_btn"):
                st.session_state.portal = 'faculty'
                st.rerun()
        
        with col3:
            if st.button("👨‍🎓 Student Portal", use_container_width=True, type="primary", key="student_btn"):
                st.session_state.portal = 'student'
                st.rerun()
        
        # Feature highlights
        st.markdown("---")
        st.markdown("### ✨ Key Features")
        
        feat_col1, feat_col2, feat_col3, feat_col4 = st.columns(4)
        
        with feat_col1:
            st.markdown("#### 🍴 Custom Lunch")
            st.write("Configure lunch duration (30-90 min) per semester")
        
        with feat_col2:
            st.markdown("#### ☕ Break Support")
            st.write("Add breaks after specific lectures")
        
        with feat_col3:
            st.markdown("#### 🌅 Morning Limits")
            st.write(f"Max {FACULTY_MORNING_LIMIT} lectures at 9AM per faculty")
        
        with feat_col4:
            st.markdown("#### 🧬 AI Scheduling")
            st.write("Hybrid algorithms for optimal schedules")
    
    # ==================== ADMIN PORTAL ====================
    elif st.session_state.portal == 'admin':
        st.markdown('<h1 class="main-header">👨‍💼 Admin Portal</h1>', unsafe_allow_html=True)
        
        if st.button("← Back to Portal Selection"):
            st.session_state.portal = None
            st.rerun()
        
        # Sidebar Configuration
        st.sidebar.markdown("## 🏫 School & Program Selection")
        
        # Step 1: Select School
        selected_school = st.sidebar.selectbox(
            "1️⃣ Select School",
            list(SCHOOL_CONFIG.keys()),
            format_func=lambda x: f"{x} - {SCHOOL_CONFIG[x]['name']}"
        )
        
        # CHANGE NEW-SECTION-6 fix: persist to session_state so render_semester_config_sidebar can access it
        st.session_state.selected_school = selected_school
        
        if selected_school:
            school_info = SCHOOL_CONFIG[selected_school]
            
            # Step 2: Select Program
            available_programs = school_info.get('programs', [])
            selected_program = st.sidebar.selectbox(
                "2️⃣ Select Program",
                available_programs,
                format_func=lambda x: f"{x} - {PROGRAM_CONFIG[x]['name']}"
            )
            
            st.session_state.selected_program = selected_program
            
            if selected_program:
                program_info = PROGRAM_CONFIG[selected_program]
                
                # Step 3: Select Semester
                max_semesters = program_info['semesters']
                selected_semester = st.sidebar.selectbox(
                    "3️⃣ Select Semester",
                    range(1, max_semesters + 1),
                    format_func=lambda x: f"Semester {x}"
                )
                
                st.session_state.selected_semester = selected_semester
                
                # CHANGE 1, 2: Render semester configuration
                render_semester_config_sidebar(firebase_manager, selected_program, selected_semester)
                
                # CHANGE SECTION-BATCH-1: Unified Sections & Batches sidebar expander
                # Replaces previous CHANGE NEW-SECTION-1 and CHANGE NEW-SECTION-2 blocks.
                # Sections auto-increment (A→B→C…); each section has independent batch counter (1→2→3…).
                st.sidebar.markdown("---")
                st.sidebar.markdown("### 3️⃣b Sections & Batches")

                program_key = f"{selected_school}_{selected_program}"

                # ---- Initialise session_state structures ----
                if 'sections' not in st.session_state:
                    st.session_state.sections = {}
                if program_key not in st.session_state.sections:
                    st.session_state.sections[program_key] = {}
                if selected_semester not in st.session_state.sections[program_key]:
                    # Try to restore from Firebase before defaulting to ['A']
                    _restored_secs = ['A']
                    if firebase_manager:
                        try:
                            _fb_batches = firebase_manager.get_batches(program=selected_program)
                            _sem_str = str(selected_semester)
                            for _fb_b in _fb_batches:
                                if (str(_fb_b.get('semester', '')) == _sem_str and
                                        _fb_b.get('school', '') == selected_school):
                                    _fb_all = _fb_b.get('all_sections')
                                    if _fb_all and isinstance(_fb_all, list) and len(_fb_all) > len(_restored_secs):
                                        _restored_secs = list(_fb_all)
                        except Exception:
                            pass
                    st.session_state.sections[program_key][selected_semester] = _restored_secs

                with st.sidebar.expander("📂 Sections & Batches", expanded=False):
                    st.caption("Theory → shared by whole section  |  Labs/Tutorials → per batch")

                    # ---- Convenience aliases ----
                    _sections_list = st.session_state.sections[program_key][selected_semester]
                    # Guard: initialise schools_data[program_key] if it doesn't exist yet
                    if program_key not in st.session_state.schools_data:
                        st.session_state.schools_data[program_key] = {
                            'name': program_info.get('name', program_key),
                            'school': selected_school,
                            'program': selected_program,
                            'years': max_semesters,
                            'semesters': max_semesters,
                            'batches': {},
                            'section_batches': {},
                        }
                    if 'section_batches' not in st.session_state.schools_data[program_key]:
                        st.session_state.schools_data[program_key]['section_batches'] = {}
                    _sb = st.session_state.schools_data[program_key] \
                              .setdefault('section_batches', {}) \
                              .setdefault(selected_semester, {})

                    # Ensure every section has at least batch [1];
                    # also try to restore saved batch counts from Firebase.
                    if firebase_manager:
                        try:
                            _fb_batches_sb = firebase_manager.get_batches(program=selected_program)
                            _sem_str_sb = str(selected_semester)
                            for _fb_b_sb in _fb_batches_sb:
                                if (str(_fb_b_sb.get('semester', '')) == _sem_str_sb and
                                        _fb_b_sb.get('school', '') == selected_school):
                                    _sec_name = _fb_b_sb.get('section', 'A')
                                    _sec_batches_saved = _fb_b_sb.get('section_batches', [1])
                                    if isinstance(_sec_batches_saved, list) and _sec_batches_saved:
                                        _sb.setdefault(_sec_name, _sec_batches_saved)
                        except Exception:
                            pass
                    for _s in _sections_list:
                        _sb.setdefault(_s, [1])

                    # ---- Add Section button (finds first missing letter) ----
                    _next_letter = 'A'
                    for i in range(26):
                        _cand = chr(ord('A') + i)
                        if _cand not in _sections_list:
                            _next_letter = _cand
                            break
                    
                    if st.button(f"➕ Add Section {_next_letter}",
                                 key=f"sb_add_sec_{program_key}_{selected_semester}"):
                        if _next_letter not in _sections_list:
                            _sections_list.append(_next_letter)
                            _sections_list.sort() # Keep sections alphabetical
                            _sb[_next_letter] = [1]
                            st.rerun()

                    st.markdown("---")

                    # ---- Per-section rows ----
                    for _sec in list(_sections_list):
                        _batches_for_sec = _sb.get(_sec, [1])
                        _sec_col1, _sec_col2 = st.columns([3, 1])
                        with _sec_col1:
                            st.markdown(f"**Section {_sec}** — Batches: {', '.join(map(str, _batches_for_sec))}")
                        with _sec_col2:
                            if len(_sections_list) > 1:
                                if st.button("🗑️", key=f"sb_del_sec_{program_key}_{selected_semester}_{_sec}",
                                             help=f"Delete Section {_sec}"):
                                    _sections_list.remove(_sec)
                                    _sb.pop(_sec, None)
                                    st.rerun()

                        # Add / Remove Batch (per section — does NOT touch batches[sem] section-index list)
                        _next_batch_num = (max(_batches_for_sec) + 1) if _batches_for_sec else 1
                        _b_col1, _b_col2 = st.columns(2)
                        with _b_col1:
                            if st.button(f"➕ Batch {_next_batch_num}",
                                         key=f"sb_add_batch_{program_key}_{selected_semester}_{_sec}"):
                                _batches_for_sec.append(_next_batch_num)
                                _sb[_sec] = _batches_for_sec
                                # Do NOT overwrite batches[sem] here — that holds the section index list
                                # and is recomputed from len(_sections_list) below every render.
                                st.rerun()
                        with _b_col2:
                            if len(_batches_for_sec) > 1:
                                if st.button(f"➖ Batch {_batches_for_sec[-1]}",
                                             key=f"sb_rm_batch_{program_key}_{selected_semester}_{_sec}"):
                                    _sb[_sec] = _batches_for_sec[:-1]
                                    st.rerun()

                        st.markdown("")  # spacing

                    # ---- Save to Firebase (existing save_batch, extended fields) ----
                    if st.button("💾 Save Sections & Batches",
                                 key=f"sb_save_{program_key}_{selected_semester}"):
                        for _sec in _sections_list:
                            _bi = {
                                'school': selected_school,
                                'program': selected_program,
                                'semester': selected_semester,
                                'section': _sec,
                                'section_batches': _sb.get(_sec, [1]),
                                'all_sections': _sections_list,
                            }
                            if firebase_manager:
                                firebase_manager.save_batch(_bi)
                        st.success(f"✅ Saved {len(_sections_list)} section(s)")

                # Sync schools_data.batches[semester] → section index list [1, 2, …, N].
                # The GA iterates this list and creates one schedule entry per section:
                #   batch=1 → Section A (matches subjects with section='A')
                #   batch=2 → Section B (matches subjects with section='B') … etc.
                # Per-section lab batches are read from the subjects' 'batch' field
                # (set in convert_info_to_subjects from the Batch column), NOT from here.
                _all_sections_list = st.session_state.sections[program_key][selected_semester]
                _section_index_list = list(range(1, len(_all_sections_list) + 1))
                st.session_state.schools_data[program_key] \
                    .setdefault('batches', {})[selected_semester] = _section_index_list

                # Also keep section_batches in sync (stores per-section lab batch lists)
                _sb_sync = st.session_state.schools_data[program_key] \
                               .setdefault('section_batches', {}) \
                               .setdefault(selected_semester, {})
                for _s_sync in _all_sections_list:
                    _sb_sync.setdefault(_s_sync, [1])

                # Firebase Operations
                st.sidebar.markdown("---")
                st.sidebar.markdown("### 🔥 Firebase Operations")
                
                timetable_key = f"{program_key}_Sem{selected_semester}"
                
                if st.sidebar.button("📥 Load from Firebase", key="load_firebase"):
                    if firebase_manager:
                        timetable_data = firebase_manager.load_timetable(timetable_key)
                        if timetable_data:
                            _loaded_schedule = timetable_data.get('schedule')
                            st.session_state.current_schedule = _loaded_schedule
                            st.session_state.current_semester_config = timetable_data.get('semester_config')
                            st.session_state.generated_schedules[timetable_key] = _loaded_schedule

                            # Rebuild section_batch_schedules from the loaded schedule
                            # so Tab 3 continues to show correct data after a page reload.
                            if _loaded_schedule:
                                import copy as _copy_load
                                _lsbs_prog = f"{selected_school}_{selected_program}"
                                _lsbs_sem = selected_semester
                                _lsbs_raw_sk = next(iter(_loaded_schedule), _lsbs_prog)
                                _lsbs_flat = _loaded_schedule.get(_lsbs_raw_sk, {})
                                _lsbs_bkeys = [k for k in _lsbs_flat if k not in
                                               {'Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'}]

                                # Determine sections: try to load individual SecX_Combined docs
                                # from Firebase first (most reliable source after a restart).
                                _lsbs_combined_loaded = {}
                                if firebase_manager:
                                    _alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                                    for _lsi in range(max(len(_lsbs_bkeys), 1)):
                                        _lsec_l = _alpha[_lsi]
                                        _sec_doc_key = f"{_lsbs_prog}_Sem{_lsbs_sem}_Sec{_lsec_l}_Combined"
                                        try:
                                            _sec_doc = firebase_manager.load_timetable(_sec_doc_key)
                                            if _sec_doc and _sec_doc.get('schedule'):
                                                _sec_inner = _sec_doc['schedule']
                                                _sec_school_k = next(iter(_sec_inner), None)
                                                if _sec_school_k:
                                                    _sec_combined_inner = _sec_inner[_sec_school_k]
                                                    _actual_combined = next(
                                                        (v for v in _sec_combined_inner.values()
                                                         if isinstance(v, dict) and
                                                         any(d in v for d in ['Monday','Tuesday','Wednesday'])),
                                                        None
                                                    )
                                                    if _actual_combined:
                                                        _lsbs_combined_loaded[_lsec_l] = _actual_combined
                                        except Exception:
                                            pass

                                st.session_state.section_batch_schedules \
                                    .setdefault(_lsbs_prog, {})[_lsbs_sem] = {}
                                _lsbs_sbs = st.session_state.section_batch_schedules[_lsbs_prog][_lsbs_sem]

                                # Derive section count: max of batch keys count vs loaded combined docs
                                _lsbs_n_sects = max(len(_lsbs_bkeys), len(_lsbs_combined_loaded), 1)
                                _alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                                for _li in range(_lsbs_n_sects):
                                    _lsec = _alpha[_li]
                                    _lbk = _lsbs_bkeys[_li] if _li < len(_lsbs_bkeys) else next(iter(_lsbs_flat), None)
                                    _lsec_raw = _lsbs_flat.get(_lbk, {}) if _lbk else {}
                                    # Prefer the separately-saved Combined doc over raw schedule
                                    _combined_to_use = (
                                        _lsbs_combined_loaded.get(_lsec) or
                                        _copy_load.deepcopy(_lsec_raw)
                                    )
                                    _lsbs_sbs[_lsec] = {
                                        '_combined': _combined_to_use,
                                        '_section': _copy_load.deepcopy(_lsec_raw),
                                        1: _copy_load.deepcopy(_lsec_raw),
                                    }
                                    # Also restore sections session_state so sidebar stays in sync
                                    if _lsec not in st.session_state.sections.get(_lsbs_prog, {}).get(_lsbs_sem, []):
                                        st.session_state.sections \
                                            .setdefault(_lsbs_prog, {}) \
                                            .setdefault(_lsbs_sem, ['A'])
                                        if _lsec not in st.session_state.sections[_lsbs_prog][_lsbs_sem]:
                                            st.session_state.sections[_lsbs_prog][_lsbs_sem].append(_lsec)

                            st.sidebar.success("✅ Loaded from Firebase")
                        else:
                            st.sidebar.warning("No timetable found in Firebase")
                
                if timetable_key in st.session_state.generated_schedules:
                    if st.sidebar.button("💾 Save to Firebase", key="save_firebase"):
                        if firebase_manager:
                            batch_info = {
                                'start_date': st.session_state.get(
                                    f"start_date_{selected_program}_{selected_semester}_val",
                                    datetime.now().strftime('%Y-%m-%d')),
                                'duration_days': st.session_state.get(
                                    f"duration_{selected_program}_{selected_semester}_val", 90),
                            }
                            
                            # Get semester config
                            sem_config = None
                            if firebase_manager:
                                lunch_cfg = firebase_manager.get_semester_lunch_config(selected_program, selected_semester)
                                break_cfg = firebase_manager.get_semester_break_config(selected_program, selected_semester)
                                sem_config = {'lunch': lunch_cfg, 'breaks': break_cfg}
                            
                            success, msg = firebase_manager.save_timetable(
                                timetable_key,
                                st.session_state.generated_schedules[timetable_key],
                                batch_info,
                                sem_config
                            )
                            if success:
                                st.sidebar.success(msg)
                            else:
                                st.sidebar.error(msg)
                    
                    if st.sidebar.button("🗑️ Delete from Firebase", key="delete_firebase"):
                        if firebase_manager:
                            success, msg = firebase_manager.delete_timetable(timetable_key)
                            if success:
                                st.sidebar.success(msg)
                                if timetable_key in st.session_state.generated_schedules:
                                    del st.session_state.generated_schedules[timetable_key]
                                st.rerun()
                            else:
                                st.sidebar.error(msg)
                
                # Timetable status
                st.sidebar.markdown("---")
                st.sidebar.markdown("### 📋 Timetable Status")
                
                if timetable_key in st.session_state.generated_schedules:
                    st.sidebar.success("✅ GENERATED")
                else:
                    st.sidebar.warning("⏳ Not Generated")
        
        # Main content area
        st.markdown("---")
        
        # Tabs
        tab0, tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
            "🏠 Dashboard",
            "📤 Dataset Upload", 
            "📅 Generate Timetable", 
            "📊 Generated Timetables",
            "✏️ Edit & Update Timetable",
            "🔥 Firebase Management",
            "🤖 AI Agent",
            "📈 Reports & Analytics"
        ])
        
        # ==================== TAB 0: DASHBOARD ====================
        with tab0:
            show_admin_dashboard(firebase_manager)
        
        # ==================== TAB 1: DATASET UPLOAD ====================
        with tab1:
            st.markdown("### 📤 Dataset Upload")
            st.markdown("Upload your datasets to configure the timetable generation system.")
            
            # CHANGE: Nuclear Reset Button
            if st.button("🧨 NUCLEAR RESET: Clear All Local Session Data", type="secondary", help="Use this if you feel data is 'stuck' or cross-contaminating from another semester."):
                reset_dataset_session_state()
                st.rerun()
            
            upload_manager = DatasetUploadManager(firebase_manager)
            
            current_program = st.session_state.get('selected_program', 'BTECH')
            current_semester = st.session_state.get('selected_semester', 1)
            
            st.info(f"📌 Uploading for: **{current_program}** - **Semester {current_semester}**")
            
            col1, col2 = st.columns(2)
            
            # Info Dataset Upload
            with col1:
                st.markdown("#### 📚 Info Dataset")
                st.markdown('<p class="tooltip-text">Contains subject, faculty, and load information</p>', unsafe_allow_html=True)
                
                with st.expander("📋 Column Descriptions", expanded=False):
                    for col_name, description in INFO_DATASET_COLUMNS.items():
                        st.markdown(f'<div class="column-info"><b>{col_name}</b>: {description}</div>', unsafe_allow_html=True)
                
                info_file = st.file_uploader(
                    "Choose Info Dataset (CSV/Excel)",
                    type=['csv', 'xlsx', 'xls'],
                    key="info_dataset_upload",
                    help="Upload CSV or Excel file with subject and faculty information",
                    on_change=reset_dataset_session_state
                )
                
                if info_file:
                    try:
                        if info_file.name.endswith('.csv'):
                            df = pd.read_csv(info_file)
                        else:
                            df = pd.read_excel(info_file)
                        
                        st.markdown("##### Preview:")
                        st.dataframe(df.head(10), use_container_width=True)
                        
                        st.markdown(f"**Total Records:** {len(df)}")
                        
                        records, errors, warnings = upload_manager.parse_info_dataset(df)
                        
                        # NOTE: semester-mismatch check intentionally removed here.
                        # The filter below (filtered_records) already handles this correctly —
                        # only records matching the selected program+semester are imported.
                        
                        if errors:
                            st.error("❌ Validation Errors:")
                            for error in errors:
                                st.write(f"  • {error}")
                        
                        if warnings:
                            st.warning("⚠️ Warnings:")
                            for warning in warnings[:5]:
                                st.write(f"  • {warning}")
                        
                        if records and not errors:
                            st.success(f"✅ Successfully parsed {len(records)} records")
                            
                            # ── FIX BUG 1: Filter records to ONLY the currently selected
                            #    program + semester so we never store/import data from other
                            #    semesters that happen to be in the same Excel file.
                            import re as _re
                            def _clean_prog(s):
                                return _re.sub(r'[^A-Z0-9]', '', str(s).upper())
                            
                            _cp_key = str(current_program).strip().upper()
                            _cp_name = str(PROGRAM_CONFIG.get(_cp_key, {}).get('name', '')).strip().upper()
                            _cs_str  = str(current_semester).strip()
                            _roman_fwd = {'1':'i','2':'ii','3':'iii','4':'iv','5':'v',
                                          '6':'vi','7':'vii','8':'viii','9':'ix','10':'x'}
                            _roman_rev = {v: k for k, v in _roman_fwd.items()}

                            def _record_matches(rec):
                                """Return True if record belongs to current_program & current_semester."""
                                rp = str(rec.get('Program', '')).strip().upper()
                                rs = str(rec.get('Sem', '')).strip().lower()
                                # normalise float → int string (Excel reads integers as 4.0)
                                if '.' in rs:
                                    try:
                                        rs = str(int(float(rs)))
                                    except (ValueError, TypeError):
                                        pass
                                # normalise roman → arabic
                                if rs in _roman_rev:
                                    rs = _roman_rev[rs]
                                # program match (accept empty Program field too)
                                if not rp:
                                    prog_ok = True
                                else:
                                    prog_ok = (_clean_prog(rp) == _clean_prog(_cp_key) or
                                               _clean_prog(rp) == _clean_prog(_cp_name))
                                # semester match (accept empty Sem field too)
                                if not rs or not _cs_str:
                                    sem_ok = True
                                else:
                                    sem_ok = (rs == _cs_str)
                                return prog_ok and sem_ok

                            filtered_records = [r for r in records if _record_matches(r)]

                            if not filtered_records:
                                # Fallback: if nothing matched (file has no program/sem columns
                                # filled in), use all records but warn the user.
                                filtered_records = records
                                st.warning(
                                    f"⚠️ Could not filter records by Program/Semester. "
                                    f"All {len(records)} records will be used. "
                                    f"Please ensure your file has correct Program and Sem columns."
                                )
                            else:
                                st.info(
                                    f"🔍 Filtered: {len(filtered_records)} records match "
                                    f"**{current_program} – Semester {current_semester}** "
                                    f"(skipped {len(records) - len(filtered_records)} records "
                                    f"from other programs/semesters)."
                                )
                            
                            col_a, col_b = st.columns(2)
                            
                            with col_a:
                                if st.button("📥 Import to Session", key="import_info_session"):
                                    # Clear old data first to prevent sticking
                                    reset_dataset_session_state()
                                    
                                    st.session_state.info_dataset = filtered_records
                                    st.session_state.subjects = upload_manager.convert_info_to_subjects(filtered_records)
                                    st.session_state.faculties = upload_manager.extract_faculty_from_info(filtered_records)
                                    st.success(f"✅ Imported {len(st.session_state.info_dataset)} records to session")
                                    st.info(f"📚 {len(st.session_state.subjects)} subject entries created")
                                    st.info(f"👨‍🏫 {len(st.session_state.faculties)} faculty members identified")
                            
                            with col_b:
                                if st.button("☁️ Save to Firebase", key="import_info_firebase"):
                                    if firebase_manager:
                                        success, msg = upload_manager.save_info_dataset_to_firebase(
                                            filtered_records, current_program, current_semester
                                        )
                                        if success:
                                            st.session_state.info_dataset = filtered_records
                                            st.session_state.subjects = upload_manager.convert_info_to_subjects(filtered_records)
                                            st.session_state.faculties = upload_manager.extract_faculty_from_info(filtered_records)
                                            st.success(f"✅ {msg}")
                                        else:
                                            st.error(f"❌ {msg}")
                                    else:
                                        st.error("Firebase not connected")
                    
                    except Exception as e:
                        st.error(f"Error reading file: {str(e)}")
                
                st.markdown("---")
                if st.button("📥 Load Info Dataset from Firebase", key="load_info_firebase"):
                    if firebase_manager:
                        info_data = firebase_manager.get_info_dataset(current_program, current_semester)
                        if info_data and 'data' in info_data:
                            st.session_state.info_dataset = info_data['data']
                            st.session_state.subjects = upload_manager.convert_info_to_subjects(info_data['data'])
                            st.session_state.faculties = upload_manager.extract_faculty_from_info(info_data['data'])
                            st.success(f"✅ Loaded {len(info_data['data'])} records from Firebase")
                        else:
                            st.warning("No Info Dataset found in Firebase for this program/semester")
            
            # CHANGE NEW-SECTION-4: Elective Subjects Configuration
            # Shown below both dataset columns when info_dataset is available in session
            if st.session_state.get('info_dataset'):
                st.markdown("---")
                with st.expander("🎯 Elective Configuration", expanded=False):
                    st.markdown(
                        "Group subjects into elective sets. All subjects in one elective group "
                        "are scheduled **in parallel** (same time, different rooms)."
                    )

                    # Get unique module names from the info dataset
                    info_records = st.session_state.info_dataset
                    all_modules = sorted(set(r.get('Module Name', '') for r in info_records if r.get('Module Name')))

                    if not all_modules:
                        st.info("No modules found in the loaded Info Dataset.")
                    else:
                        # Initialize elective_groups in session_state if not present
                        if 'elective_groups' not in st.session_state:
                            st.session_state.elective_groups = []

                        # Button to add a new elective group
                        if st.button("➕ Add New Elective Group",
                                     key="add_elective_group"):
                            st.session_state.elective_groups.append({
                                'name': f"Elective {len(st.session_state.elective_groups) + 1}",
                                'subjects': []
                            })

                        # Render each elective group
                        groups_to_keep = []
                        for g_idx, group in enumerate(st.session_state.elective_groups):
                            with st.container():
                                g_col1, g_col2 = st.columns([4, 1])
                                with g_col1:
                                    group_name = st.text_input(
                                        "Group Name",
                                        value=group['name'],
                                        key=f"elec_grp_name_{g_idx}"
                                    )
                                    st.session_state.elective_groups[g_idx]['name'] = group_name
                                with g_col2:
                                    if st.button("🗑️ Remove Group",
                                                 key=f"remove_elec_grp_{g_idx}"):
                                        continue  # mark for removal by not appending

                                st.markdown(f"**Select subjects for {group_name}:**")
                                # Checkboxes for each module
                                selected_subjects = []
                                subj_cols = st.columns(3)
                                for m_idx, module in enumerate(all_modules):
                                    already_in = module in group.get('subjects', [])
                                    checked = subj_cols[m_idx % 3].checkbox(
                                        module,
                                        value=already_in,
                                        key=f"elec_{g_idx}_{m_idx}"
                                    )
                                    if checked:
                                        selected_subjects.append(module)

                                st.session_state.elective_groups[g_idx]['subjects'] = selected_subjects
                                if selected_subjects:
                                    st.success(f"📌 {group_name}: {', '.join(selected_subjects)}")
                                else:
                                    st.info("No subjects selected yet for this group.")
                                st.markdown("---")
                                groups_to_keep.append(g_idx)

                        # Remove deleted groups (maintain list integrity)
                        st.session_state.elective_groups = [
                            g for i, g in enumerate(st.session_state.elective_groups)
                            if i in groups_to_keep
                        ]

                        # Save to Firebase inside existing info_dataset document
                        if st.button("💾 Save Elective Groups to Firebase",
                                     key="save_elective_groups"):
                            if firebase_manager and st.session_state.elective_groups:
                                try:
                                    doc_id = f"{current_program}_Sem{current_semester}"
                                    elective_data = [
                                        {'name': g['name'], 'subjects': g['subjects']}
                                        for g in st.session_state.elective_groups
                                    ]
                                    firebase_manager.db.collection(
                                        firebase_manager.collections['info_dataset']
                                    ).document(doc_id).set(
                                        {'electives': elective_data},
                                        merge=True
                                    )
                                    st.success("✅ Elective groups saved to Firebase!")
                                except Exception as e:
                                    st.error(f"❌ Error saving electives: {str(e)}")
                            elif not firebase_manager:
                                st.error("Firebase not connected")
                            else:
                                st.warning("No elective groups to save.")

            # Room Dataset Upload
            with col2:

                st.markdown("#### 🏢 Room Dataset")
                st.markdown('<p class="tooltip-text">Maps subjects to rooms by class type</p>', unsafe_allow_html=True)
                
                with st.expander("📋 Column Descriptions", expanded=False):
                    for col_name, description in ROOM_DATASET_COLUMNS.items():
                        st.markdown(f'<div class="column-info"><b>{col_name}</b>: {description}</div>', unsafe_allow_html=True)
                
                room_file = st.file_uploader(
                    "Choose Room Dataset (CSV/Excel)",
                    type=['csv', 'xlsx', 'xls'],
                    key="room_dataset_upload",
                    help="Upload CSV or Excel file with room assignments",
                    on_change=reset_dataset_session_state
                )
                
                if room_file:
                    try:
                        if room_file.name.endswith('.csv'):
                            df = pd.read_csv(room_file)
                        else:
                            df = pd.read_excel(room_file)
                        
                        st.markdown("##### Preview:")
                        st.dataframe(df.head(10), use_container_width=True)
                        
                        st.markdown(f"**Total Records:** {len(df)}")
                        
                        records, errors, warnings = upload_manager.parse_room_dataset(df)
                        
                        if errors:
                            st.error("❌ Validation Errors:")
                            for error in errors:
                                st.write(f"  • {error}")
                        
                        if warnings:
                            st.warning("⚠️ Warnings:")
                            for warning in warnings[:5]:
                                st.write(f"  • {warning}")
                        
                        if records and not errors:
                            st.success(f"✅ Successfully parsed {len(records)} room mappings")
                            
                            unique_rooms = set()
                            for record in records:
                                if record.get('Room No.'):
                                    unique_rooms.add(record['Room No.'])
                            
                            st.info(f"🏢 {len(unique_rooms)} unique rooms identified")
                            
                            col_a, col_b = st.columns(2)
                            
                            with col_a:
                                if st.button("📥 Import to Session", key="import_room_session"):
                                    st.session_state.room_dataset = records
                                    rooms_list = []
                                    for room_no in unique_rooms:
                                        room_type = 'Classroom'
                                        for r in records:
                                            if r.get('Room No.') == room_no and r.get('Class Type', '').lower() == 'lab':
                                                room_type = 'Lab'
                                                break
                                        
                                        rooms_list.append({
                                            'room_id': room_no,
                                            'name': room_no,
                                            'capacity': 60,
                                            'building': 'Main',
                                            'type': room_type,
                                            'equipment': ['Projector', 'Whiteboard'] if room_type == 'Classroom' else ['Computers', 'Projector']
                                        })
                                    
                                    st.session_state.rooms = rooms_list
                                    st.success(f"✅ Imported {len(records)} room mappings to session")
                            
                            with col_b:
                                if st.button("☁️ Save to Firebase", key="import_room_firebase"):
                                    if firebase_manager:
                                        success, msg = upload_manager.save_room_dataset_to_firebase(
                                            records, current_program, current_semester
                                        )
                                        if success:
                                            st.session_state.room_dataset = records
                                            st.success(f"✅ {msg}")
                                        else:
                                            st.error(f"❌ {msg}")
                                    else:
                                        st.error("Firebase not connected")
                    
                    except Exception as e:
                        st.error(f"Error reading file: {str(e)}")
                
                st.markdown("---")
                if st.button("📥 Load Room Dataset from Firebase", key="load_room_firebase"):
                    if firebase_manager:
                        room_data = firebase_manager.get_room_dataset(current_program, current_semester)
                        if room_data and 'data' in room_data:
                            st.session_state.room_dataset = room_data['data']
                            st.success(f"✅ Loaded {len(room_data['data'])} room mappings from Firebase")
                        else:
                            st.warning("No Room Dataset found in Firebase for this program/semester")
            
            # Room Allocation Section
            st.markdown("---")
            st.markdown("### 🔧 Room Allocation")
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                if st.session_state.info_dataset:
                    st.success(f"✅ Info Dataset loaded: {len(st.session_state.info_dataset)} records")
                else:
                    st.warning("⚠️ Please upload Info Dataset first")
                
                if st.session_state.room_dataset:
                    st.success(f"✅ Room Dataset loaded: {len(st.session_state.room_dataset)} mappings")
                else:
                    st.info("ℹ️ Room Dataset not loaded - auto-allocation will be used")
            
            with col2:
                if st.button("🔄 Allocate Rooms", type="primary", key="allocate_rooms"):
                    if st.session_state.info_dataset:
                        room_allocator = RoomAllocator(firebase_manager)
                        allocations, msg = room_allocator.allocate_rooms(current_program, current_semester)
                        
                        if allocations:
                            st.session_state.room_allocations = allocations
                            st.success(f"✅ {msg}")
                            
                            with st.expander("📋 Room Allocation Summary", expanded=True):
                                alloc_df = pd.DataFrame([
                                    {'Subject': k.rsplit('_', 1)[0], 'Type': k.rsplit('_', 1)[1], 'Room': ", ".join(v) if isinstance(v, list) else v}
                                    for k, v in allocations.items()
                                ])
                                st.dataframe(alloc_df, use_container_width=True)
                        else:
                            st.error(f"❌ {msg}")
                    else:
                        st.error("Please upload Info Dataset first")
            
            # Current Data Summary
            st.markdown("---")
            st.markdown("### 📊 Current Data Summary")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Schools Configured", len(st.session_state.schools_data))
            col2.metric("Total Subjects", len(st.session_state.subjects))
            col3.metric("Total Faculty", len(st.session_state.faculties))
            col4.metric("Total Rooms", len(st.session_state.rooms))
            
            # CHANGE 3: Morning limit info
            st.markdown(f'<span class="morning-limit-badge">🌅 Morning Limits: Enforced ({FACULTY_MORNING_LIMIT} max/faculty/week at 9AM)</span>', unsafe_allow_html=True)
        
        # ==================== TAB 2: GENERATE TIMETABLE ====================
        with tab2:
            st.markdown("### 🚀 Generate Timetable with AI/ML Algorithms")

            # ── Auto-load elective groups from Firebase if not already in session ──
            if 'elective_groups' not in st.session_state or not st.session_state.elective_groups:
                _t2_current_program = st.session_state.get('selected_program', '')
                _t2_current_sem = st.session_state.get('selected_semester', 1)
                if firebase_manager and _t2_current_program:
                    try:
                        _el_doc_id = f"{_t2_current_program}_Sem{_t2_current_sem}"
                        _el_doc = firebase_manager.db.collection(
                            firebase_manager.collections['info_dataset']
                        ).document(_el_doc_id).get()
                        if _el_doc.exists:
                            _el_data = _el_doc.to_dict().get('electives', [])
                            if _el_data:
                                st.session_state.elective_groups = _el_data
                    except Exception:
                        pass

            col1, col2 = st.columns([2, 1])
            
            with col1:
                ready = True
                missing = []
                
                if not st.session_state.schools_data:
                    missing.append("❌ School/Program configuration")
                    ready = False
                else:
                    st.success("✅ Schools/Programs configured")
                
                if not st.session_state.subjects:
                    missing.append("❌ Subjects (from Info Dataset)")
                    ready = False
                else:
                    st.success(f"✅ {len(st.session_state.subjects)} subjects loaded")
                
                if not st.session_state.faculties:
                    missing.append("❌ Faculty (from Info Dataset)")
                    ready = False
                else:
                    st.success(f"✅ {len(st.session_state.faculties)} faculty members loaded")
                
                if not st.session_state.rooms:
                    missing.append("❌ Rooms")
                    ready = False
                else:
                    st.success(f"✅ {len(st.session_state.rooms)} rooms loaded")
                
                # CHANGE 1, 2: Show config status
                if firebase_manager:
                    lunch_config = firebase_manager.get_semester_lunch_config(
                        st.session_state.get('selected_program', 'BTECH'),
                        st.session_state.get('selected_semester', 1)
                    )
                    if lunch_config and lunch_config.get('custom'):
                        st.success(f"✅ Custom lunch: {lunch_config.get('start')} - {lunch_config.get('end')} ({lunch_config.get('duration')} min)")
                    else:
                        st.info(f"ℹ️ Using default lunch ({DEFAULT_LUNCH_DURATION} min)")
                    
                    break_config = firebase_manager.get_semester_break_config(
                        st.session_state.get('selected_program', 'BTECH'),
                        st.session_state.get('selected_semester', 1)
                    )
                    if break_config and break_config.get('enabled'):
                        st.success(f"✅ Breaks: {break_config.get('duration')} min after lectures {break_config.get('placements')}")
                    else:
                        st.info("ℹ️ No breaks configured")
                
                if missing:
                    st.warning("Missing data:")
                    for item in missing:
                        st.write(item)
            
            with col2:
                st.markdown("#### Algorithm Selection")
                algorithm_choice = st.selectbox(
                    "Choose Algorithm",
                    ["hybrid", "genetic_only", "hungarian_graph"],
                    format_func=lambda x: {
                        "hybrid": "🧬 Hybrid (Hungarian + Graph + GA)",
                        "genetic_only": "🧬 Genetic Algorithm Only",
                        "hungarian_graph": "🎯 Hungarian + Graph Coloring"
                    }[x]
                )
                
                if algorithm_choice == "hybrid":
                    st.info("Uses all three algorithms for optimal results")
                elif algorithm_choice == "genetic_only":
                    st.slider("GA Generations", 20, 100, 50, key="ga_gens")
                    st.slider("GA Population Size", 50, 200, 100, key="ga_pop")
                
                if st.button("🚀 GENERATE TIMETABLE", type="primary", disabled=not ready):
                    # SAFETY CHECK: Sync validation
                    valid_data = True
                    if st.session_state.subjects:
                        sample_subj = st.session_state.subjects[0]
                        session_prog = sample_subj.get('program', '')
                        session_sem = sample_subj.get('semester', 0)
                        
                        if str(session_prog).strip().upper() != str(selected_program).strip().upper() or \
                           str(session_sem).strip() != str(selected_semester).strip():
                            st.error(f"⚠️ DATA MISMATCH DETECTED!")
                            st.write(f"Loaded data is for **{session_prog} Sem {session_sem}**, but you have selected **{selected_program} Sem {selected_semester}**.")
                            st.warning("Please upload the correct dataset for this selection or click 'Import to Session' again if you just changed the dropdown.")
                            valid_data = False
                    
                    if valid_data:
                        with st.spinner("Running AI/ML optimization..."):
                            scheduler = SmartTimetableScheduler(firebase_manager)
                            
                            current_program = st.session_state.get('selected_program', 'BTECH')
                            current_semester = st.session_state.get('selected_semester', 1)

                            # ── FIX BUG 2 (generate-time): Ensure every configured section
                            # has its own set of subjects before the scheduler runs.
                            # This is the authoritative fix — it runs AFTER sections are
                            # confirmed in session_state (sidebar has already set them up).
                            _gen_prog_key_fix = f"{selected_school}_{current_program}"
                            _gen_sem_fix = int(current_semester) if str(current_semester).isdigit() else 1
                            _configured_secs_fix = (
                                st.session_state.get('sections', {})
                                               .get(_gen_prog_key_fix, {})
                                               .get(_gen_sem_fix, ['A'])
                            )

                            if len(_configured_secs_fix) > 1:
                                import copy as _copy_fix

                                _subjects_work = list(st.session_state.subjects)

                                # Find which sections already have subjects
                                _secs_present = set(
                                    str(s.get('section', '')).strip().upper()
                                    for s in _subjects_work if s.get('section')
                                )
                                _source_sec = next(iter(_secs_present), 'A')

                                # Replicate for each missing section
                                _extra_subjects = []
                                _newly_added_secs = []
                                for _ms in _configured_secs_fix:
                                    _ms_up = str(_ms).upper()
                                    if _ms_up not in _secs_present:
                                        _newly_added_secs.append(_ms_up)
                                        for _s in _subjects_work:
                                            if str(_s.get('section', '')).strip().upper() == _source_sec:
                                                _ns = _copy_fix.deepcopy(_s)
                                                _ns['section'] = _ms_up
                                                _extra_subjects.append(_ns)

                                if _extra_subjects:
                                    st.session_state.subjects = _subjects_work + _extra_subjects
                                    st.info(
                                        f"ℹ️ Generate-time: subjects replicated for "
                                        f"Section(s) {_newly_added_secs} from Section {_source_sec} "
                                        f"({len(_extra_subjects)} extra entries)."
                                    )
                            # ── END MULTI-SECTION GENERATE-TIME FIX ─────────────────────────
                            
                            # Pass elective groups so the scheduler knows which
                            # subjects are optional alternatives (only 1 from each
                            # group needs to be scheduled per section).
                            scheduler.elective_groups = list(
                                st.session_state.get('elective_groups', [])
                            )

                            # Always re-allocate rooms based on the current dataset in Firebase before generation
                            room_allocator = RoomAllocator(firebase_manager)
                            allocations, _ = room_allocator.allocate_rooms(current_program, current_semester)
                            if allocations:
                                st.session_state.room_allocations = allocations

                            schedule, semester_config = scheduler.generate_hybrid_timetable(
                                st.session_state.schools_data,
                                st.session_state.faculties,
                                st.session_state.subjects,
                                st.session_state.rooms,
                                algorithm_choice,
                                room_allocations=st.session_state.room_allocations,
                                program=current_program,
                                semester=current_semester
                            )
                            
                            # Store generated schedule
                            for school_key in st.session_state.schools_data:
                                for sem in st.session_state.schools_data[school_key].get('batches', {}).keys():
                                    timetable_key = f"{school_key}_Sem{sem}"
                                    st.session_state.generated_schedules[timetable_key] = schedule
                            
                            st.session_state.current_schedule = schedule
                            st.session_state.current_semester_config = semester_config
                            
                            # Save to Firebase
                            if firebase_manager and schedule:
                                timetable_key = f"{selected_school}_{current_program}_Sem{current_semester}"
                                
                                firebase_manager.save_timetable(timetable_key, schedule, {
                                    'program': current_program,
                                    'semester': current_semester,
                                    'generated_at': datetime.now().isoformat()
                                }, semester_config)
                            
                            st.success("✅ Timetable generated successfully!")

                            # Post-process schedule per section.
                            _sel_prog_key = f"{st.session_state.get('selected_school', selected_school)}_{current_program}"
                            _sel_sem = current_semester

                            # ── DERIVE section list from THREE sources; use the largest count ──
                            # Source 1: what we explicitly replicated subjects for
                            _src1 = list(_configured_secs_fix) if _configured_secs_fix else ['A']
                            # Source 2: what schools_data.batches says (set by sidebar every render)
                            _batches_val = (
                                st.session_state.schools_data
                                  .get(_sel_prog_key, {})
                                  .get('batches', {})
                                  .get(_sel_sem, [1])
                            )
                            _src2 = [chr(65 + i) for i in range(len(_batches_val))]
                            # Source 3: what the raw schedule actually contains (count batch keys)
                            _raw_sk_tmp = next(iter(schedule), _sel_prog_key)
                            _raw_flat_tmp = schedule.get(_raw_sk_tmp, {})
                            _day_set_tmp = {'Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'}
                            _batch_keys_tmp = [k for k in _raw_flat_tmp if k not in _day_set_tmp]
                            _src3 = [chr(65 + i) for i in range(len(_batch_keys_tmp))]
                            # Use the source with the most sections (guards against partial resets)
                            _gen_sections = max([_src1, _src2, _src3], key=len)
                            if not _gen_sections:
                                _gen_sections = ['A']
                            st.info(f"📋 Processing {len(_gen_sections)} section(s): {_gen_sections}")
                            _gen_sb = st.session_state.schools_data.get(_sel_prog_key, {}) \
                                          .get('section_batches', {}).get(_sel_sem, {})

                            import copy as _copy

                            # Initialise section_batch_schedules store (clear previous)
                            if 'section_batch_schedules' not in st.session_state:
                                st.session_state.section_batch_schedules = {}
                            st.session_state.section_batch_schedules \
                                .setdefault(_sel_prog_key, {})[_sel_sem] = {}
                            _sbs = st.session_state.section_batch_schedules[_sel_prog_key][_sel_sem]

                            # Locate the school key in the generated schedule
                            _raw_school_key = next(iter(schedule), _sel_prog_key)
                            _raw_flat = schedule.get(_raw_school_key, {})
                            _raw_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
                            _day_names_set = {'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'}

                            # Check schedule structure
                            _first_child = next(iter(_raw_flat), None)
                            _flat_by_day = _first_child in _day_names_set

                            # ── BUILD A DIRECT MAPPING: section index → raw schedule key ──
                            # Use exact string matching first ("Sem_{sem}_Section_{N}") so that
                            # the semester number cannot accidentally be confused with the section
                            # number when both happen to be equal (e.g. Sem_1_Section_1).
                            _num_to_sched = {}
                            if not _flat_by_day:
                                for _sec_idx, _sec_label in enumerate(_gen_sections):
                                    _sec_idx_1 = _sec_idx + 1  # 1-based
                                    _exact_keys = [
                                        f"Sem_{_sel_sem}_Section_{_sec_idx_1}",
                                        f"Sem_{_sel_sem}_Section_{_sec_label}",
                                        f"Sem{_sel_sem}_Section_{_sec_idx_1}",
                                        f"Sem{_sel_sem}_Section_{_sec_label}",
                                        f"Section_{_sec_label}",
                                        str(_sec_idx_1),
                                    ]
                                    for _ek in _exact_keys:
                                        if _ek in _raw_flat:
                                            _num_to_sched[_sec_idx_1] = _raw_flat[_ek]
                                            break

                            # Helper: check whether a day-schedule dict has at least one real class
                            def _has_real_classes(day_sched: dict) -> bool:
                                for _v in day_sched.values():
                                    if _v is None:
                                        continue
                                    _items = _v if isinstance(_v, list) else [_v]
                                    for _i in _items:
                                        if isinstance(_i, dict) and _i.get('type') not in ('LUNCH', 'BREAK', None):
                                            return True
                                return False

                            def _build_combined(raw_sched: dict) -> dict:
                                """Flatten a raw GA schedule dict into the combined format."""
                                _out = {}
                                for _day in _raw_days:
                                    _day_slots = raw_sched.get(_day, {})
                                    _out[_day] = {}
                                    for _slot_key, _ci_raw in _day_slots.items():
                                        if not _ci_raw:
                                            _out[_day][_slot_key] = None
                                            continue
                                        _classes_to_process = _ci_raw if isinstance(_ci_raw, list) else [_ci_raw]
                                        _lab_entries = []
                                        _theory_entry = None
                                        _fixed_entry = None
                                        for _ci in _classes_to_process:
                                            _ctype = str(_ci.get('type', '') or '').strip().upper()
                                            if _ctype in ('LUNCH', 'BREAK'):
                                                _fixed_entry = _copy.deepcopy(_ci)
                                            elif 'LAB' in _ctype or 'TUTORIAL' in _ctype or 'PRACTICAL' in _ctype:
                                                _b_val = str(_ci.get('batch', '1')).strip()
                                                if _b_val.endswith('.0'):
                                                    _b_val = _b_val[:-2]
                                                _lab_entries.append({'batch': _b_val, 'ci': _copy.deepcopy(_ci)})
                                            else:
                                                if _theory_entry is None:
                                                    _tc = _copy.deepcopy(_ci)
                                                    _tc.pop('batch', None)
                                                    _theory_entry = _tc
                                        if _fixed_entry:
                                            _out[_day][_slot_key] = _fixed_entry
                                        elif _lab_entries:
                                            _out[_day][_slot_key] = _lab_entries
                                        elif _theory_entry:
                                            _out[_day][_slot_key] = _theory_entry
                                return _out

                            for _sec_i, _sec in enumerate(_gen_sections):
                                _sec_num = _sec_i + 1  # 1-based index

                                if _flat_by_day:
                                    # Single-section fallback: all sections share the raw schedule
                                    _sec_raw_schedule = _raw_flat
                                else:
                                    _sec_raw_schedule = _num_to_sched.get(_sec_num)

                                    # If still not found, try remaining keys in _raw_flat
                                    if _sec_raw_schedule is None:
                                        _already_used = set(_num_to_sched.values())
                                        for _rk, _rv in _raw_flat.items():
                                            if _rv not in _already_used:
                                                _sec_raw_schedule = _rv
                                                break

                                # Build the combined schedule from the raw GA output
                                if _sec_raw_schedule:
                                    _combined_sec_schedule = _build_combined(_sec_raw_schedule)
                                else:
                                    _combined_sec_schedule = {_d: {} for _d in _raw_days}

                                # ── FALLBACK: if this section has no real classes, copy theory
                                #    from Section A (index 0).  In most colleges theory lectures
                                #    are shared across sections; labs remain per-batch.
                                _sec_has_classes = any(
                                    _has_real_classes(_combined_sec_schedule.get(_d, {}))
                                    for _d in _raw_days
                                )
                                if not _sec_has_classes and _sec_i > 0:
                                    # Section A must already be in _sbs (processed first)
                                    _sec_a_label = _gen_sections[0]
                                    _sec_a_combined = _sbs.get(_sec_a_label, {}).get('_combined', {})
                                    if _sec_a_combined:
                                        # Copy all entries from Section A, stamping correct section
                                        _combined_sec_schedule = {}
                                        for _d in _raw_days:
                                            _combined_sec_schedule[_d] = {}
                                            for _sk, _sv in _sec_a_combined.get(_d, {}).items():
                                                if _sv is None:
                                                    _combined_sec_schedule[_d][_sk] = None
                                                elif isinstance(_sv, dict):
                                                    _cv = _copy.deepcopy(_sv)
                                                    _combined_sec_schedule[_d][_sk] = _cv
                                                elif isinstance(_sv, list):
                                                    _combined_sec_schedule[_d][_sk] = _copy.deepcopy(_sv)
                                        st.info(
                                            f"ℹ️ Section {_sec}: schedule mirrored from Section "
                                            f"{_sec_a_label} (combined-lecture model)."
                                        )

                                _sbs[_sec] = {'_combined': _combined_sec_schedule, '_section': _combined_sec_schedule}

                                # Save combined section-level to Firebase
                                if firebase_manager and _combined_sec_schedule:
                                    _sec_tkey = f"{_sel_prog_key}_Sem{_sel_sem}_Sec{_sec}_Combined"
                                    try:
                                        firebase_manager.save_timetable(
                                            _sec_tkey,
                                            {_raw_school_key: {f'Section{_sec}Combined': _combined_sec_schedule}},
                                            {'program': current_program, 'semester': _sel_sem,
                                             'section': _sec,
                                             'generated_at': datetime.now().isoformat()},
                                            semester_config
                                        )
                                    except Exception as _fb_err:
                                        st.warning(f"⚠️ Could not save Section {_sec} to Firebase: {_fb_err}")

                            st.success(f"✅ Created combined timetables for {len(_gen_sections)} section(s): "
                                       f"{', '.join(f'Sec {s}' for s in _gen_sections)}. "
                                       f"See 'Generated Timetables' tab.")




# app.py - Part 5: Remaining Tabs, Faculty Portal, Student Portal
# Continuation from Part 4

        # ==================== TAB 3: GENERATED TIMETABLES ====================
        with tab3:
            st.markdown("### 📋 View Generated Timetables")

            # Tab 3 - UNIFIED SECTION TIMETABLE (v4):
            # One timetable per section. Theory classes are shared (one block).
            # Labs/Tutorials are stacked batch-by-batch in the same cell.
            # Dropdown shows only "Sem2 | Sec A", "Sem2 | Sec B" — no batch sub-options.

            _t3_sem_cfg = st.session_state.get('current_semester_config', None)
            _t3_prog_key = f"{selected_school}_{selected_program}"
            _t3_sem = selected_semester

            _t3_all_sbs = st.session_state.get('section_batch_schedules', {})
            # Try both with and without school prefix so the key always resolves
            _t3_sbs = (
                _t3_all_sbs.get(_t3_prog_key, {}).get(_t3_sem, {}) or
                next(
                    (v.get(_t3_sem, {}) for k, v in _t3_all_sbs.items()
                     if selected_program in k and v.get(_t3_sem)),
                    {}
                )
            )

            # ── FIREBASE FALLBACK: if session_state is empty, probe Firebase for
            #    SecA_Combined, SecB_Combined … documents directly.
            #    This handles a fresh browser session where the user hasn't clicked
            #    "Load" yet but timetables were previously generated and saved.
            if not _t3_sbs and firebase_manager:
                import copy as _t3_copy
                _t3_fb_sbs = {}
                _t3_alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                for _t3_si in range(6):  # probe up to 6 sections (A..F)
                    _t3_sl = _t3_alpha[_t3_si]
                    _t3_doc_key = f"{_t3_prog_key}_Sem{_t3_sem}_Sec{_t3_sl}_Combined"
                    try:
                        _t3_doc = firebase_manager.load_timetable(_t3_doc_key)
                        if _t3_doc and _t3_doc.get('schedule'):
                            _t3_inner = _t3_doc['schedule']
                            _t3_sk = next(iter(_t3_inner), None)
                            if _t3_sk:
                                _t3_combined_inner = _t3_inner[_t3_sk]
                                _t3_combined_sched = next(
                                    (v for v in _t3_combined_inner.values()
                                     if isinstance(v, dict) and
                                     any(d in v for d in ['Monday', 'Tuesday', 'Wednesday'])),
                                    None
                                )
                                if _t3_combined_sched:
                                    _t3_fb_sbs[_t3_sl] = {'_combined': _t3_combined_sched}
                        else:
                            break  # no more sections in Firebase
                    except Exception:
                        break
                if _t3_fb_sbs:
                    _t3_sbs = _t3_fb_sbs
                    # Also persist so subsequent tab switches don't re-query Firebase
                    st.session_state.section_batch_schedules \
                        .setdefault(_t3_prog_key, {})[_t3_sem] = {
                            _k: _t3_copy.deepcopy(_v) for _k, _v in _t3_fb_sbs.items()
                        }

            # Build one option per section using the '_combined' schedule
            _t3_options = []  # list of (label, combined_schedule_dict)

            if _t3_sbs:
                for _t3_sec, _t3_sec_data in sorted(_t3_sbs.items()):
                    _combined_sched = _t3_sec_data.get('_combined')
                    if _combined_sched:
                        _t3_options.append((f"Sem{_t3_sem} | Sec {_t3_sec}", _combined_sched))

            # Fallback: use raw current_schedule if no combined schedules built yet
            if not _t3_options:
                _t3_sched = st.session_state.get('current_schedule', {})
                if _t3_sched:
                    _t3_raw_sk = next(iter(_t3_sched), None)
                    if _t3_raw_sk:
                        _t3_raw_flat = _t3_sched[_t3_raw_sk]
                        _t3_day_names = {'Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'}
                        _t3_first_child = next(iter(_t3_raw_flat), None)
                        if _t3_first_child in _t3_day_names:
                            _t3_options.append((f"Sem{_t3_sem} | Full Schedule", _t3_raw_flat))
                        else:
                            for _fallback_bk, _fallback_sched in _t3_raw_flat.items():
                                _t3_options.append((f"Sem{_t3_sem} | {_fallback_bk}", _fallback_sched))

            if _t3_options:
                _t3_label_opts = [lbl for lbl, _ in _t3_options]
                _t3_sel_label = st.selectbox("Select Section", _t3_label_opts, key="t3_sec_batch_sel")

                # Retrieve the combined schedule for the selected section
                batch_schedule = next(sched for lbl, sched in _t3_options if lbl == _t3_sel_label)

                st.markdown(f"#### 📅 Timetable – **{_t3_sel_label}**")
                st.caption("🟩 Theory classes are shared for all batches.  🟧 Each colored block in lab slots = one batch.")

                # Show semester config info
                if _t3_sem_cfg:
                    cfg_c1, cfg_c2, cfg_c3 = st.columns(3)
                    with cfg_c1:
                        _lc = _t3_sem_cfg.get('lunch', {})
                        if _lc:
                            st.info(f"🍴 Lunch: {_lc.get('start','N/A')}-{_lc.get('end','N/A')} ({_lc.get('duration',50)} min)")
                    with cfg_c2:
                        _bc2 = _t3_sem_cfg.get('breaks', {})
                        if _bc2 and _bc2.get('enabled'):
                            st.info(f"☕ Breaks: {_bc2.get('duration')} min after {_bc2.get('placements')}")
                        else:
                            st.info("☕ No breaks configured")
                    with cfg_c3:
                        st.info(f"📊 {len(_t3_sem_cfg.get('time_slots', []))} time slots")

                # Clash detection
                _t3_full_sched = st.session_state.get('current_schedule', {})
                _cd = ClashDetector(firebase_manager)
                _clashes = _cd.detect_all_clashes(_t3_full_sched) if _t3_full_sched else []
                if _clashes:
                    st.error(f"⚠️ {len(_clashes)} clashes detected in full schedule")
                else:
                    st.success("✅ Clash Count: 0")

                # ── Unified pastel color palette (30 colors, light enough for black text) ──
                _CLR_PALETTE = [
                    "#aecbfa", "#fdd7b5", "#b5ead7", "#f4c2c2", "#c3b1e1",
                    "#fdffb6", "#a2d2ff", "#d0f4de", "#ffe5b4", "#e8c9fa",
                    "#b9fbc0", "#ffcbdb", "#c9e4de", "#ffc8a2", "#d4e6f1",
                    "#ffefd5", "#d5f5e3", "#f9e4b7", "#dce8f8", "#f5cba7",
                    "#d7bde2", "#a9cce3", "#a8e6cf", "#ffd3b6", "#ffaaa5",
                    "#c7ceea", "#b2f7ef", "#ecd6f7", "#ffd6e0", "#c8f7c5",
                ]
                _CLR_LUNCH = "#e8f5e9"  # very light green for lunch
                _CLR_BREAK = "#fff9c4"  # very light yellow for break

                # ── Build unified subject → color map (one color per unique subject name) ──
                # Scans all days/slots in batch_schedule to discover every unique subject
                _subj_colors = {}
                _c_idx = 0
                _days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
                if any("Saturday" in str(k) for k in batch_schedule.keys()):
                    _days.append("Saturday")
                _ts = ExportManager.get_time_slots_from_schedule(batch_schedule)
                # CHANGE PARALLEL-LAB-FIX-9: Handle both direct and {batch,ci} wrapper formats
                for _d in _days:
                    for _sl in _ts:
                        _slot_val = (batch_schedule or {}).get(_d, {}).get(_sl)
                        if not _slot_val:
                            continue
                        entries = _slot_val if isinstance(_slot_val, list) else [_slot_val]
                        for _e in entries:
                            # Extract the real class info dict – handles both direct
                            # entries ({subject,faculty,...}) and post-processed wrappers ({batch,ci:{...}})
                            _ci_inner = _e.get('ci', _e) if isinstance(_e, dict) else _e
                            _sn = _ci_inner.get('subject', '') if isinstance(_ci_inner, dict) else ''
                            _ct = str(_ci_inner.get('type', '') or '').strip().upper() if isinstance(_ci_inner, dict) else ''
                            if _sn and _sn not in _subj_colors and _ct not in ('LUNCH', 'BREAK'):
                                _subj_colors[_sn] = _CLR_PALETTE[_c_idx % len(_CLR_PALETTE)]
                                _c_idx += 1

                # ── Cell renderers ──
                # CHANGE PARALLEL-BATCH-FIX-5: Updated cell rendering to match official format and fix parallel height
                def _block(ci, batch_label=None):
                    """Single colored block for one class."""
                    ct = str(ci.get('type','') or '').strip().upper()
                    subj = ci.get('subject','N/A'); fac = ci.get('faculty','TBD'); room = ci.get('room','TBD')
                    bg = _subj_colors.get(subj, '#bde0fe')
                    
                    # Layout as per official format: "Batch X: Subject" for parallel labs
                    is_lab = any(k in ct for k in ['LAB', 'TUTORIAL', 'PRACTICAL'])
                    
                    # Don't show (Theory) for regular lectures. Keep (Lab) optional depending on name
                    type_display = f" ({ci.get('type', '')})" if ct not in ('THEORY', '') and not subj.upper().endswith(ct) else ""

                    if is_lab:
                        # Target format: [Subject Name] Lab Batch [0X] [Faculty Initial] [Room]
                        batch_str = f" Batch {batch_label}" if batch_label and str(batch_label).strip() else ""
                        lab_text = f"{subj}{type_display}{batch_str} {fac} {room}"
                        return (f'<div style="background:{bg};color:#000;border-bottom:1px solid rgba(0,0,0,0.2);'
                                f'padding:8px;text-align:center;box-sizing:border-box;'
                                f'display:flex;flex-direction:column;justify-content:center;flex-grow:1;min-height:0;">'
                                f'<div style="font-weight:bold;font-size:0.8rem;line-height:1.3;">{lab_text}</div>'
                                f'</div>')
                    else:
                        return (f'<div style="background:{bg};color:#000;border-bottom:1px solid rgba(0,0,0,0.2);'
                                f'padding:8px;text-align:center;box-sizing:border-box;'
                                f'display:flex;flex-direction:column;justify-content:center;flex-grow:1;min-height:0;">'
                                f'<div style="font-weight:bold;font-size:0.85rem;line-height:1.2;margin-bottom:2px;">{subj}{type_display}</div>'
                                f'<div style="font-size:0.75rem;">{fac} | {room}</div>'
                                f'</div>')

                def _td(slot_val, day_idx, slot_key, all_ts):
                    """Render a <td> for one timetable slot with vertical merging for lunch/break."""
                    base = 'vertical-align:middle;padding:0;min-width:140px;border:1px solid #000;height:100px;'
                    
                    if slot_val is None:
                        return f'<td style="{base}background:#fdfdfd;text-align:center;color:#888;font-size:0.8rem;">FREE</td>'
                    
                    # Check for Lunch/Break
                    is_list = isinstance(slot_val, list)
                    entries = slot_val if is_list else [slot_val]
                    ct = str(entries[0].get('type','') or '').strip().upper()
                    
                    if ct in ('LUNCH', 'BREAK'):
                        # Render on first day with rowspan = total number of days
                        if day_idx == 0:
                            label = "L<br>U<br>N<br>C<br>H" if ct == 'LUNCH' else "S<br>H<br>O<br>R<br>T<br><br>B<br>R<br>E<br>A<br>K"
                            bg = "#fff" # White background for high contrast vertical label
                            return (f'<td rowspan="{len(_days)}" style="{base}background:{bg};color:#000;text-align:center;'
                                    f'font-weight:bold;font-size:0.9rem;width:40px;letter-spacing:2px;border:1px solid #000;">{label}</td>')
                        else:
                            return "" # Skip for other days due to rowspan
                            
                    if is_list:
                        inner_html = ''
                        for item in slot_val:
                            ci_data = item.get('ci', item)
                            # CHANGE PARALLEL-LAB-FIX-10: Strip .0 from batch strings
                            batch_num = str(item.get('batch', item.get('_batch_label', '')))
                            if batch_num.endswith('.0'):
                                batch_num = batch_num[:-2]
                            inner_html += _block(ci_data, batch_label=batch_num)
                        # Parent container with stretch-to-fit
                        return f'<td style="{base}"><div style="display:flex;flex-direction:column;height:100%;align-items:stretch;">{inner_html}</div></td>'
                    
                    # CHANGE UNIVERSAL-BATCH-FIX-6: Only pass batch_label for lab/tutorial types, NEVER for theory
                    _sv_type = str(slot_val.get('type', '') or '').strip().upper()
                    _sv_is_lab = any(k in _sv_type for k in ['LAB', 'TUTORIAL', 'PRACTICAL'])
                    _sv_batch = ''
                    if _sv_is_lab:
                        _sv_batch = str(slot_val.get('batch', '')).strip()
                        if _sv_batch.endswith('.0'):
                            _sv_batch = _sv_batch[:-2]
                    return f'<td style="{base}"><div style="display:flex;flex-direction:column;height:100%;align-items:stretch;">{_block(slot_val, batch_label=_sv_batch)}</div></td>'

                # ── Build HTML table ──
                # Use "am to pm" format for headers
                def _fmt_ts(sk):
                    pts = sk.split('-')
                    if len(pts) == 2:
                        s_time = TimeSlotManager.format_time_12hr(pts[0].strip()).lower()
                        e_time = TimeSlotManager.format_time_12hr(pts[1].strip()).lower()
                        return f"{s_time} to {e_time}"
                    return sk

                _slot_labels = [_fmt_ts(sl) for sl in _ts]
                
                # Header formatting
                # Academic Year (e.g. 2025-26)
                curr_year = datetime.now().year
                ay_str = f"{curr_year}-{str(curr_year+1)[2:]}"
                
                # Use the program config to get proper program name if available
                prog_name = selected_program  # BTECH etc.
                class_info_str = f"{prog_name}. {selected_school} SEC {_t3_sel_label.replace('|', '').replace('Sem', 'SEM').replace('Sec ', '').strip()}"
                
                # Header row with rowspan handling for Lunch/Break
                # We need to know which indices are global lunch/break
                global_v_cols = []
                for i, sl in enumerate(_ts):
                    # Check if at least one day has LUNCH/BREAK for this slot
                    is_v = False
                    for d in _days:
                        v = (batch_schedule or {}).get(d, {}).get(sl)
                        if v and not isinstance(v, list) and str(v.get('type','')).upper() in ('LUNCH', 'BREAK'):
                            is_v = True; break
                    if is_v: global_v_cols.append(i)

                _hdr = ''.join(
                    f'<th style="background:#fff;color:#000;padding:10px 8px;font-size:0.85rem;border:1px solid #000;font-weight:bold;">{lbl}</th>'
                    for lbl in _slot_labels)
                
                _rows = ''
                for d_idx, _d in enumerate(_days):
                    _cells = ''
                    for s_idx, sl in enumerate(_ts):
                        _slot_val = (batch_schedule or {}).get(_d, {}).get(sl)
                        _cells += _td(_slot_val, d_idx, sl, _ts)
                    
                    _d_label = _d[:3] if "Thursday" not in _d else "Thurs"
                    _rows += (f'<tr><td style="background:#fff;color:#000;font-weight:bold;border:1px solid #000;'
                              f'padding:12px 10px;text-align:center;font-size:0.9rem;">{_d_label}</td>{_cells}</tr>')

                st.markdown(f'''
                <div style="overflow-x:auto; margin-top:20px;">
                <table style="border-collapse:collapse; width:100%; border:2px solid #000; font-family:Arial, sans-serif;">
                  <thead>
                    <!-- Academic Header -->
                    <tr>
                      <th colspan="{len(_ts) + 1}" style="background:#000; color:#fff; padding:8px; font-size:1.1rem; border:1px solid #000;">
                        Time Table AY {ay_str} wef {datetime.now().strftime("%d.%m.%Y")}
                      </th>
                    </tr>
                    <tr>
                      <th colspan="{len(_ts) + 1}" style="background:#ffc107; color:#000; padding:10px; font-size:1.3rem; border:1px solid #000; font-weight:900;">
                        {class_info_str}
                      </th>
                    </tr>
                    <tr>
                      <th style="background:#fff; color:#000; padding:10px 8px; font-size:0.85rem; border:1px solid #000; width:80px;"></th>
                      {_hdr}
                    </tr>
                  </thead>
                  <tbody>{_rows}</tbody>
                </table>
                </div>''', unsafe_allow_html=True)

                # ── Subject Color Legend ──
                if _subj_colors:
                    st.markdown("##### 🎨 Subject Color Legend")
                    _legend_items = ''.join(
                        f'<span style="display:inline-flex;align-items:center;margin:4px 6px;padding:5px 12px;'
                        f'background:{_col};border-radius:20px;border:1px solid rgba(0,0,0,0.15);'
                        f'font-size:0.78rem;font-weight:600;color:#111;white-space:nowrap;">'
                        f'<span style="display:inline-block;width:10px;height:10px;border-radius:50%;'
                        f'background:{_col};border:1px solid rgba(0,0,0,0.25);margin-right:6px;"></span>'
                        f'{_sn}</span>'
                        for _sn, _col in sorted(_subj_colors.items())
                    )
                    st.markdown(
                        f'<div style="display:flex;flex-wrap:wrap;gap:2px;padding:10px 0;'
                        f'border-top:1px solid #e0e0e0;margin-top:12px;">'
                        f'{_legend_items}</div>',
                        unsafe_allow_html=True
                    )

                # Plain-text DataFrame for CSV/XLSX export
                _tdata = []
                for _d in _days:
                    _row = {'Day': _d}
                    for _sl in _ts:
                        _dsl = ExportManager.format_slot_for_display(_sl)
                        _ci = (batch_schedule or {}).get(_d, {}).get(_sl)
                        if _ci is None:
                            _row[_dsl] = "FREE"
                        elif isinstance(_ci, list):
                            # CHANGE UNIVERSAL-BATCH-FIX-6: Format CSV with Batch prefix for labs only
                            _csv_parts = []
                            for e in _ci:
                                _e_ci = e.get('ci', e) if isinstance(e, dict) else e
                                _e_bl = str(e.get('batch', '')).strip() if isinstance(e, dict) else ''
                                if _e_bl.endswith('.0'): _e_bl = _e_bl[:-2]
                                _e_subj = _e_ci.get('subject','N/A') if isinstance(_e_ci, dict) else 'N/A'
                                _e_fac = _e_ci.get('faculty','TBD') if isinstance(_e_ci, dict) else 'TBD'
                                _e_room = _e_ci.get('room','TBD') if isinstance(_e_ci, dict) else 'TBD'
                                _csv_parts.append(f"{_e_subj} Batch {_e_bl} {_e_fac} {_e_room}")
                            _row[_dsl] = '\n'.join(_csv_parts)
                        else:
                            _cit = str(_ci.get('type','') or '').upper()
                            if _cit == 'LUNCH': _row[_dsl] = f"LUNCH ({_ci.get('duration',50)} min)"
                            elif _cit == 'BREAK': _row[_dsl] = f"BREAK ({_ci.get('duration',10)} min)"
                            # CHANGE UNIVERSAL-BATCH-FIX-6: No batch prefix for theory in CSV
                            else: _row[_dsl] = f"{_ci.get('subject','N/A')} | {_ci.get('faculty','TBD')} | {_ci.get('room','TBD')}"
                    _tdata.append(_row)
                _df = pd.DataFrame(_tdata)

                # Export options
                _safe_label = _t3_sel_label.replace(' ', '_').replace('|', '').replace('–', '-').strip()
                st.markdown("#### 📥 Export Options")
                _ec1, _ec2, _ec3 = st.columns(3)
                with _ec1:
                    _xlsx = ExportManager.export_to_excel_formatted(
                        batch_schedule, school_name=selected_school, batch_name=_t3_sel_label)
                    st.download_button("📥 Download Excel", _xlsx,
                        file_name=f"timetable_{_safe_label}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="export_excel_view")
                with _ec2:
                    st.download_button("📥 Download CSV", _df.to_csv(index=False),
                        file_name=f"timetable_{_safe_label}.csv",
                        mime="text/csv", key="export_csv_view")
                with _ec3:
                    _pdf = ExportManager.export_to_pdf_detailed(
                        batch_schedule, school_name=selected_school, batch_name=_t3_sel_label)
                    st.download_button("📥 Download PDF", _pdf,
                        file_name=f"timetable_{_safe_label}.pdf",
                        mime="application/pdf", key="export_pdf_view")
            else:
                st.info("No timetables generated yet. Please generate a timetable first, or configure Sections & Batches in the sidebar.")

            




        # ==================== TAB 4: EDIT & UPDATE TIMETABLE ====================
        with tab4:
            st.markdown("### ✏️ Edit & Update Timetable")
            
            if st.session_state.current_schedule:
                # Edit Mode Toggle
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    if st.button("🔧 Enable Edit Mode", type="primary", disabled=st.session_state.edit_mode):
                        st.session_state.edit_mode = True
                        st.session_state.edited_schedule = st.session_state.editor.enable_edit_mode(
                            st.session_state.current_schedule
                        )
                        st.rerun()
                
                with col2:
                    if st.button("💾 Save Changes", disabled=not st.session_state.edit_mode):
                        st.session_state.current_schedule = copy.deepcopy(st.session_state.edited_schedule)
                        st.session_state.edit_mode = False
                        
                        # Save to Firebase with semester config
                        if firebase_manager:
                            year_key = list(st.session_state.generated_schedules.keys())[0] if st.session_state.generated_schedules else "default"
                            success, msg = st.session_state.editor.save_to_firebase(
                                st.session_state.current_schedule,
                                year_key,
                                semester_config=st.session_state.get('current_semester_config')
                            )
                            if success:
                                st.success(f"✅ {msg}")
                            else:
                                st.warning(f"⚠️ {msg}")
                        else:
                            st.success("✅ Changes saved locally!")
                        st.rerun()
                
                with col3:
                    if st.button("↩️ Undo Last Change", disabled=not st.session_state.edit_mode):
                        if st.session_state.editor.edit_history:
                            st.session_state.edited_schedule = st.session_state.editor.undo_last_change(
                                st.session_state.edited_schedule
                            )
                            st.success("↩️ Last change undone")
                            st.rerun()
                
                with col4:
                    if st.button("🔄 Reset to Original", disabled=not st.session_state.edit_mode):
                        st.session_state.edited_schedule = st.session_state.editor.reset_to_original()
                        st.session_state.detected_clashes = []
                        st.info("🔄 Reset to original schedule")
                        st.rerun()
                
                # Edit Mode Indicator
                if st.session_state.edit_mode:
                    st.markdown('<div class="edit-mode">🔧 <b>EDIT MODE ACTIVE</b> - Make changes to the timetable below</div>', unsafe_allow_html=True)
                    
                    # Clash Detection
                    clash_detector = ClashDetector(firebase_manager)
                    clashes = clash_detector.detect_all_clashes(st.session_state.edited_schedule, save_to_firebase=True)
                    st.session_state.detected_clashes = clashes
                    
                    if clashes:
                        st.markdown(f'<div class="clash-detected">⚠️ <b>{len(clashes)} CLASHES DETECTED after editing</b></div>', unsafe_allow_html=True)
                        
                        with st.expander("🔍 View Clash Details", expanded=True):
                            for i, clash in enumerate(clashes, 1):
                                st.error(f"**Clash {i}:** {clash['type']} - {clash['details']}")
                                if 'time' in clash:
                                    st.write(f"   📅 Time: {clash['time']}")
                                if 'faculty' in clash:
                                    st.write(f"   👨‍🏫 Faculty: {clash['faculty']}")
                                if 'room' in clash:
                                    st.write(f"   🏢 Room: {clash['room']}")
                                st.markdown("---")
                    else:
                        st.markdown('<div class="no-clash">✅ <b>No clashes - Schedule is valid!</b></div>', unsafe_allow_html=True)
                    
                    # Edit interface
                    if st.session_state.edited_schedule:
                        schedule = st.session_state.edited_schedule
                        
                        school_list = list(schedule.keys())
                        if school_list:
                            edit_school = st.selectbox("Select School/Program to Edit", school_list, key="edit_school")
                            
                            if edit_school in schedule:
                                batch_list = list(schedule[edit_school].keys())
                                if batch_list:
                                    edit_batch = st.selectbox("Select Batch to Edit", batch_list, key="edit_batch")
                                    
                                    if edit_batch in schedule[edit_school]:
                                        st.markdown(f"#### Editing: {edit_school} - {edit_batch}")
                                        
                                        batch_schedule = schedule[edit_school][edit_batch]
                                        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
                                        
                                        # CHANGE 1: Get dynamic time slots
                                        all_slots = ExportManager.get_time_slots_from_schedule(batch_schedule)
                                        
                                        # Edit operations
                                        st.markdown("##### Edit Operations")
                                        
                                        edit_op = st.selectbox("Select Operation", 
                                            ["Swap Slots", "Update Class", "Remove Class", "Add Class"],
                                            key="edit_operation"
                                        )
                                        
                                        if edit_op == "Swap Slots":
                                            col1, col2 = st.columns(2)
                                            with col1:
                                                st.markdown("**Slot 1:**")
                                                day1 = st.selectbox("Day 1", days, key="swap_day1")
                                                slot1 = st.selectbox("Time 1", all_slots, 
                                                    format_func=lambda x: ExportManager.format_slot_for_display(x),
                                                    key="swap_slot1")
                                            with col2:
                                                st.markdown("**Slot 2:**")
                                                day2 = st.selectbox("Day 2", days, key="swap_day2")
                                                slot2 = st.selectbox("Time 2", all_slots,
                                                    format_func=lambda x: ExportManager.format_slot_for_display(x),
                                                    key="swap_slot2")
                                            
                                            if st.button("🔄 Swap", key="do_swap"):
                                                st.session_state.edited_schedule, success, msg = st.session_state.editor.swap_slots(
                                                    st.session_state.edited_schedule, edit_school, edit_batch, day1, slot1, day2, slot2
                                                )
                                                if success:
                                                    st.success(msg)
                                                else:
                                                    st.error(msg)
                                                st.rerun()
                                        
                                        elif edit_op == "Remove Class":
                                            day = st.selectbox("Day", days, key="remove_day")
                                            slot = st.selectbox("Time Slot", all_slots,
                                                format_func=lambda x: ExportManager.format_slot_for_display(x),
                                                key="remove_slot")
                                            
                                            if st.button("🗑️ Remove", key="do_remove"):
                                                st.session_state.edited_schedule, success, msg = st.session_state.editor.remove_class(
                                                    st.session_state.edited_schedule, edit_school, edit_batch, day, slot
                                                )
                                                if success:
                                                    st.success(msg)
                                                st.rerun()
                                        
                                        elif edit_op == "Add Class":
                                            day = st.selectbox("Day", days, key="add_day")
                                            slot = st.selectbox("Time Slot", all_slots,
                                                format_func=lambda x: ExportManager.format_slot_for_display(x),
                                                key="add_slot")
                                            
                                            subject = st.text_input("Subject Name", key="add_subject")
                                            faculty = st.text_input("Faculty Name", key="add_faculty")
                                            room = st.text_input("Room", key="add_room")
                                            class_type = st.selectbox("Type", ["Theory", "Lab", "Tutorial"], key="add_type")
                                            
                                            if st.button("➕ Add Class", key="do_add"):
                                                # Get slot info for duration
                                                slot_parts = slot.split('-')
                                                start_time = slot_parts[0] if len(slot_parts) > 0 else ''
                                                end_time = slot_parts[1] if len(slot_parts) > 1 else ''
                                                
                                                duration = DEFAULT_LECTURE_DURATION
                                                if start_time and end_time:
                                                    duration = TimeSlotManager.time_to_minutes(end_time) - TimeSlotManager.time_to_minutes(start_time)
                                                
                                                class_info = {
                                                    'subject': subject,
                                                    'faculty': faculty,
                                                    'room': room,
                                                    'type': class_type,
                                                    'duration': duration,
                                                    'start': start_time,
                                                    'end': end_time
                                                }
                                                st.session_state.edited_schedule, success, msg = st.session_state.editor.add_class(
                                                    st.session_state.edited_schedule, edit_school, edit_batch, day, slot, class_info
                                                )
                                                if success:
                                                    st.success(msg)
                                                else:
                                                    st.error(msg)
                                                st.rerun()
                                        
                                        elif edit_op == "Update Class":
                                            day = st.selectbox("Day", days, key="update_day")
                                            slot = st.selectbox("Time Slot", all_slots,
                                                format_func=lambda x: ExportManager.format_slot_for_display(x),
                                                key="update_slot")
                                            
                                            current_val = batch_schedule.get(day, {}).get(slot, {})
                                            # Handle combined list - take first entry for pre-filling edit fields
                                            current_class = current_val[0] if isinstance(current_val, list) and current_val else current_val
                                            
                                            subject = st.text_input("Subject Name", 
                                                value=current_class.get('subject', '') if isinstance(current_class, dict) else '', 
                                                key="update_subject")
                                            faculty = st.text_input("Faculty Name", 
                                                value=current_class.get('faculty', '') if isinstance(current_class, dict) else '', 
                                                key="update_faculty")
                                            room = st.text_input("Room", 
                                                value=current_class.get('room', '') if isinstance(current_class, dict) else '', 
                                                key="update_room")
                                            
                                            type_options = ["Theory", "Lab", "Tutorial", "LUNCH", "BREAK"]
                                            current_type = current_class.get('type', 'Theory') if current_class else 'Theory'
                                            type_index = type_options.index(current_type) if current_type in type_options else 0
                                            
                                            class_type = st.selectbox("Type", type_options, index=type_index, key="update_type")
                                            
                                            if st.button("✏️ Update", key="do_update"):
                                                slot_parts = slot.split('-')
                                                start_time = slot_parts[0] if len(slot_parts) > 0 else ''
                                                end_time = slot_parts[1] if len(slot_parts) > 1 else ''
                                                
                                                new_info = {
                                                    'subject': subject,
                                                    'faculty': faculty,
                                                    'room': room,
                                                    'type': class_type,
                                                    'duration': current_class.get('duration', DEFAULT_LECTURE_DURATION) if current_class else DEFAULT_LECTURE_DURATION,
                                                    'start': start_time,
                                                    'end': end_time
                                                }
                                                st.session_state.edited_schedule, success, msg = st.session_state.editor.update_class_info(
                                                    st.session_state.edited_schedule, edit_school, edit_batch, day, slot, new_info
                                                )
                                                if success:
                                                    st.success(msg)
                                                else:
                                                    st.error(msg)
                                                st.rerun()
                                        
                                        # Display current timetable
                                        st.markdown("##### Current Timetable")
                                        timetable_data = []
                                        for day in days:
                                            row = {'Day': day}
                                            for slot in all_slots:
                                                display_slot = ExportManager.format_slot_for_display(slot)
                                                slot_val = batch_schedule.get(day, {}).get(slot)
                                                if slot_val:
                                                    # Handle combined list
                                                    slot_items = slot_val if isinstance(slot_val, list) else [slot_val]
                                                    item_labels = []
                                                    for class_info in slot_items:
                                                        if isinstance(class_info, dict):
                                                            if class_info.get('type') == 'LUNCH':
                                                                item_labels.append("🍴 LUNCH")
                                                            elif class_info.get('type') == 'BREAK':
                                                                item_labels.append("☕ BREAK")
                                                            else:
                                                                item_labels.append(f"{class_info.get('subject', 'N/A')[:15]}")
                                                    row[display_slot] = " / ".join(list(dict.fromkeys(item_labels))) # unique labels
                                                else:
                                                    row[display_slot] = "FREE"
                                            timetable_data.append(row)
                                        
                                        df = pd.DataFrame(timetable_data)
                                        st.dataframe(df, use_container_width=True)
            else:
                st.info("📝 No timetables generated yet. Please generate a timetable first to enable editing.")
        
        # ==================== TAB 5: FIREBASE MANAGEMENT ====================
        with tab5:
            st.markdown("### 🔥 Firebase Database Management")
            
            if firebase_manager:
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("#### 📊 Database Statistics")
                    
                    timetables_count = len(firebase_manager.get_all_timetables())
                    info_datasets = firebase_manager.get_info_dataset()
                    room_datasets = firebase_manager.get_room_dataset()
                    lunch_configs = firebase_manager.get_all_lunch_configs()
                    break_configs = firebase_manager.get_all_break_configs()
                    
                    st.metric("Timetables in Database", timetables_count)
                    st.metric("Info Datasets", len(info_datasets) if info_datasets else 0)
                    st.metric("Room Datasets", len(room_datasets) if room_datasets else 0)
                    
                    # CHANGE 1, 2: Show config counts
                    st.metric("Lunch Configurations", len(lunch_configs) if lunch_configs else 0)
                    st.metric("Break Configurations", len(break_configs) if break_configs else 0)
                    
                    # Show existing datasets
                    if info_datasets:
                        with st.expander("📚 Info Datasets in Firebase"):
                            for dataset in info_datasets:
                                st.write(f"• {dataset.get('program', 'N/A')} Sem {dataset.get('semester', 'N/A')} - {dataset.get('record_count', 0)} records")
                    
                    if room_datasets:
                        with st.expander("🏢 Room Datasets in Firebase"):
                            for dataset in room_datasets:
                                st.write(f"• {dataset.get('program', 'N/A')} Sem {dataset.get('semester', 'N/A')} - {dataset.get('record_count', 0)} mappings")
                    
                    # CHANGE 1, 2: Show configs
                    if lunch_configs:
                        with st.expander("🍴 Lunch Configurations"):
                            for config in lunch_configs:
                                status = "🔒 Locked" if config.get('locked') else "🔓 Unlocked"
                                custom = "Custom" if config.get('custom') else "Default"
                                st.write(f"• {config.get('program', 'N/A')} Sem {config.get('semester', 'N/A')} - {config.get('start', 'N/A')}-{config.get('end', 'N/A')} ({custom}) {status}")
                    
                    if break_configs:
                        with st.expander("☕ Break Configurations"):
                            for config in break_configs:
                                if config.get('enabled'):
                                    st.write(f"• {config.get('program', 'N/A')} Sem {config.get('semester', 'N/A')} - {config.get('duration', 0)} min after {config.get('placements', [])}")
                
                with col2:
                    st.markdown("#### 🔄 Sync Operations")
                    
                    if st.button("📥 Pull All Data from Firebase", use_container_width=True):
                        with st.spinner("Pulling data from Firebase..."):
                            # Pull info datasets and extract data
                            all_info = firebase_manager.get_info_dataset()
                            if all_info:
                                all_records = []
                                for dataset in all_info:
                                    if 'data' in dataset:
                                        all_records.extend(dataset['data'])
                                
                                if all_records:
                                    upload_manager = DatasetUploadManager(firebase_manager)
                                    st.session_state.info_dataset = all_records
                                    st.session_state.subjects = upload_manager.convert_info_to_subjects(all_records)
                                    st.session_state.faculties = upload_manager.extract_faculty_from_info(all_records)
                            
                            # Pull rooms
                            st.session_state.rooms = firebase_manager.get_rooms_list()
                            
                            st.success("✅ Data pulled from Firebase")
                    
                    if st.button("📤 Push Current Data to Firebase", use_container_width=True):
                        with st.spinner("Pushing data to Firebase..."):
                            current_program = st.session_state.get('selected_program', 'BTECH')
                            current_semester = st.session_state.get('selected_semester', 1)
                            
                            if st.session_state.info_dataset:
                                firebase_manager.save_info_dataset(
                                    current_program, current_semester, st.session_state.info_dataset
                                )
                            
                            if st.session_state.room_dataset:
                                firebase_manager.save_room_dataset(
                                    current_program, current_semester, st.session_state.room_dataset
                                )
                            
                            st.success("✅ Data pushed to Firebase")
                    
                    if st.button("🗑️ Clear Local Cache", use_container_width=True):
                        st.session_state.faculties = []
                        st.session_state.subjects = []
                        st.session_state.rooms = []
                        st.session_state.info_dataset = []
                        st.session_state.room_dataset = []
                        st.session_state.room_allocations = {}
                        st.session_state.generated_schedules = {}
                        st.session_state.current_schedule = None
                        st.session_state.current_semester_config = None
                        st.success("✅ Local cache cleared")
                        st.rerun()
                    
                    st.markdown("---")
                    with st.expander("🗑️ Delete Timetable from Firebase", expanded=False):
                        st.warning("⚠️ This action will permanently wipe the timetable and all associated datasets (Info, Room, Lunch/Break Configs).")
                        
                        existing_tt = firebase_manager.get_all_timetables()
                        if existing_tt:
                            tt_options = {f"{t['year']} (Generated: {t.get('created_at', 'N/A')})": t['year'] for t in existing_tt}
                            selected_tt_label = st.selectbox("Select Timetable to Wipe", list(tt_options.keys()), key="delete_tt_select")
                            selected_tt_key = tt_options[selected_tt_label]
                            
                            if st.button("Delete Selected Timetable", type="primary", use_container_width=True):
                                with st.spinner("Wiping timetable data..."):
                                    success, msg = firebase_manager.delete_timetable(selected_tt_key, archive=True)
                                    if success:
                                        st.success(f"✅ {msg}")
                                        time_module.sleep(1)
                                        st.rerun()
                                    else:
                                        st.error(f"❌ {msg}")
                        else:
                            st.info("No timetables found in database to delete.")

                    with st.expander("🚨 Global Reset: Empty Entire Database", expanded=False):
                        st.error("❗ DANGER ZONE: This will delete EVERYTHING from the Firebase database across all semesters and programs.")
                        st.info("Associated datasets, configurations, archives, and generated schedules will be permanently lost.")
                        
                        confirm_text = st.text_input("Type 'DELETE' to unlock the reset button:", key="confirm_wipe_input")
                        
                        if st.button("EXECUTE PERMANENT GLOBAL RESET", type="primary", use_container_width=True, disabled=(confirm_text != "DELETE")):
                            with st.spinner("Deleting all Firebase data..."):
                                success, msg = firebase_manager.wipe_entire_database()
                                if success:
                                    # Clear Session State to reflect reality immediately
                                    keys_to_reset = [
                                        'info_dataset', 'room_dataset', 'subjects', 'faculties', 'rooms',
                                        'room_allocations', 'generated_schedules', 'current_schedule',
                                        'edited_schedule', 'detected_clashes', 'section_batch_schedules',
                                        'schools_data', 'sections'
                                    ]
                                    for k in keys_to_reset:
                                        if k in st.session_state:
                                            if isinstance(st.session_state[k], list): st.session_state[k] = []
                                            elif isinstance(st.session_state[k], dict): st.session_state[k] = {}
                                            else: st.session_state[k] = None
                                    
                                    st.success(f"✅ {msg}")
                                    time_module.sleep(2)
                                    st.rerun()
                                else:
                                    st.error(f"❌ {msg}")
                
                # Unresolved clashes
                st.markdown("---")
                st.markdown("#### ⚠️ Unresolved Clashes")
                
                clashes = firebase_manager.get_unresolved_clashes()
                if clashes:
                    for clash in clashes[:5]:
                        st.warning(f"{clash.get('type', 'Unknown')} - {clash.get('details', 'No details')}")
                else:
                    st.success("No unresolved clashes")
                
                # CHANGE 3: Faculty morning constraints
                st.markdown("---")
                st.markdown("#### 🌅 Faculty Morning Constraints")
                
                morning_counts = firebase_manager.get_faculty_morning_counts()
                if morning_counts:
                    at_limit = [(f, c) for f, c in morning_counts.items() if c >= FACULTY_MORNING_LIMIT]
                    under_limit = [(f, c) for f, c in morning_counts.items() if c < FACULTY_MORNING_LIMIT]
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**At Limit:**")
                        if at_limit:
                            for f, c in at_limit[:10]:
                                st.write(f"⚠️ {f}: {c}/{FACULTY_MORNING_LIMIT}")
                        else:
                            st.write("None")
                    
                    with col2:
                        st.markdown("**Under Limit:**")
                        if under_limit:
                            for f, c in under_limit[:10]:
                                st.write(f"✅ {f}: {c}/{FACULTY_MORNING_LIMIT}")
                        else:
                            st.write("None")
                else:
                    st.info("No morning constraint data available")
            else:
                st.error("Firebase not connected. Please check your configuration.")
        
        # ==================== TAB 6: AI AGENT ====================
        with tab6:
            from agent.agent_ui import render_agent_tab
            from genetic_algorithm import GeneticAlgorithm
            render_agent_tab(
                firebase_manager,
                ClashDetector,
                genetic_algorithm=GeneticAlgorithm(),
            )
        
        # ==================== TAB 7: REPORTS & ANALYTICS ====================
        with tab7:
            st.markdown("### 📈 Reports & Analytics")
            st.markdown("Generate comprehensive reports for administration and analysis.")
            
            if firebase_manager:
                report_generator = ReportGenerator(firebase_manager)
                
                report_type = st.selectbox(
                    "Select Report Type",
                    [
                        "📊 Faculty Workload Analysis",
                        "🏢 Room Utilization Report",
                        "📚 Program Summary Report",
                        "⚠️ Clash History Report",
                        "🤖 Agent Repair History Report",
                        "📊 Research Metrics & Paper Exports",
                        "⚙️ Semester Configurations Report",  # CHANGE 1, 2: New report
                        "📋 Comprehensive Report (All)"
                    ],
                    key="report_type_selector"
                )
                
                st.markdown("---")
                
                if report_type == "📊 Faculty Workload Analysis":
                    st.markdown("#### 👨‍🏫 Faculty Workload Analysis")
                    
                    if st.button("🔄 Generate Report", key="gen_faculty_report"):
                        with st.spinner("Generating faculty workload report..."):
                            df = report_generator.generate_faculty_workload_report()
                            
                            if not df.empty:
                                # Display metrics
                                col1, col2, col3, col4 = st.columns(4)
                                
                                with col1:
                                    st.metric("Total Faculty", len(df))
                                with col2:
                                    overloaded = len(df[df['Workload Status'] == 'Overloaded'])
                                    st.metric("Overloaded", overloaded)
                                with col3:
                                    optimal = len(df[df['Workload Status'] == 'Optimal'])
                                    st.metric("Optimal Load", optimal)
                                with col4:
                                    # CHANGE 3: Show morning limit issues
                                    at_limit = len(df[df['Morning Limit Status'] == '⚠️ At Limit'])
                                    st.metric("At Morning Limit", at_limit)
                                
                                st.dataframe(df, use_container_width=True, height=400)
                                
                                # Export button
                                excel_data = report_generator.export_faculty_report_to_excel()
                                st.download_button(
                                    label="📥 Download Faculty Report (Excel)",
                                    data=excel_data,
                                    file_name=f"faculty_workload_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                            else:
                                st.info("No data available.")
                
                elif report_type == "🏢 Room Utilization Report":
                    st.markdown("#### 🏢 Room Utilization Report")
                    
                    if st.button("🔄 Generate Report", key="gen_room_report"):
                        with st.spinner("Generating room utilization report..."):
                            df = report_generator.generate_room_utilization_report()
                            
                            if not df.empty:
                                st.dataframe(df, use_container_width=True, height=400)
                                
                                excel_data = report_generator.export_room_report_to_excel()
                                st.download_button(
                                    label="📥 Download Room Report (Excel)",
                                    data=excel_data,
                                    file_name=f"room_utilization_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                            else:
                                st.info("No data available.")
                
                elif report_type == "📚 Program Summary Report":
                    st.markdown("#### 📚 Program Summary Report")
                    
                    if st.button("🔄 Generate Report", key="gen_program_report"):
                        with st.spinner("Generating program summary report..."):
                            df = report_generator.generate_program_summary_report()
                            
                            if not df.empty:
                                st.dataframe(df, use_container_width=True, height=300)
                            else:
                                st.info("No data available.")
                
                elif report_type == "⚠️ Clash History Report":
                    st.markdown("#### ⚠️ Clash History Report")
                    
                    if st.button("🔄 Generate Report", key="gen_clash_report"):
                        with st.spinner("Generating clash history report..."):
                            df = report_generator.generate_clash_history_report()
                            st.dataframe(df, use_container_width=True, height=400)

                elif report_type == "🤖 Agent Repair History Report":
                    st.markdown("#### 🤖 Agent Repair History Report")
                    st.markdown(
                        "View all agentic repair actions saved to Firebase `/repair_history`."
                    )

                    if st.button("🔄 Generate Report", key="gen_agent_repair_report"):
                        with st.spinner("Loading agent repair history from Firebase..."):
                            df = report_generator.generate_agent_repair_history_report()
                            st.dataframe(df, use_container_width=True, height=400)

                            sessions = firebase_manager.get_agent_sessions(limit=20)
                            if sessions:
                                st.markdown("##### Recent Agent Sessions")
                                session_rows = []
                                for session in sessions:
                                    session_rows.append({
                                        'Session ID': session.get('session_id', session.get('id', '')),
                                        'Timetable': session.get('timetable_key', ''),
                                        'Status': session.get('status', ''),
                                        'Clashes Found': session.get('clashes_found', 0),
                                        'Clashes Fixed': session.get('clashes_fixed', 0),
                                        'Turns Used': session.get('turns_used', 0),
                                    })
                                st.dataframe(
                                    pd.DataFrame(session_rows),
                                    use_container_width=True,
                                    hide_index=True,
                                )

                elif report_type == "📊 Research Metrics & Paper Exports":
                    st.markdown("#### 📊 Research Metrics & Paper Exports")
                    st.markdown(
                        "Run Phase 3 research benchmarks and export CSV metrics, "
                        "before/after schedules, conversation logs, and paper figures."
                    )

                    if st.button("🔄 Run Benchmarks & Export Bundle", key="gen_research_bundle"):
                        with st.spinner("Running legacy vs agentic benchmarks..."):
                            from agent.metrics_collector import MetricsCollector
                            from agent.mock_repair_client import AutoRepairMockClient
                            from agent.research_export import export_research_bundle
                            from agent.scenarios import RESEARCH_SCENARIOS
                            from genetic_algorithm import GeneticAlgorithm

                            collector = MetricsCollector(
                                firebase_manager=firebase_manager,
                                genetic_algorithm=GeneticAlgorithm(),
                                clash_detector_cls=ClashDetector,
                            )
                            client_factory = lambda sched, cons: AutoRepairMockClient(
                                sched, cons, max_moves=30
                            )
                            results = collector.run_all_research_scenarios(
                                llm_client_factory=client_factory
                            )

                            comparison_df = pd.DataFrame(
                                collector.build_comparison_table(results)
                            )
                            time_df = pd.DataFrame(
                                collector.build_time_complexity_table(results)
                            )
                            st.markdown("##### Table 1: Clash Resolution Comparison")
                            st.dataframe(comparison_df, use_container_width=True, hide_index=True)
                            st.markdown("##### Table 2: Time Complexity")
                            st.dataframe(time_df, use_container_width=True, hide_index=True)

                            conversation_logs = {}
                            schedules = {}
                            for key, builder in RESEARCH_SCENARIOS.items():
                                before, constraints, _, scenario_name = builder()
                                semester = 4 if "scenario_b" in key else 2
                                after = copy.deepcopy(before)
                                agent_result = collector.run_agentic_repair(
                                    after,
                                    constraints,
                                    scenario_name,
                                    llm_client_factory=client_factory,
                                    semester=semester,
                                )
                                session_id = agent_result.session_id or f"session_{key}"
                                conversation_logs[session_id] = [
                                    {"scenario": scenario_name, "metrics": agent_result.to_dict()}
                                ]
                                schedules[scenario_name] = {"before": before, "after": after}

                            bundle = export_research_bundle(
                                results,
                                conversation_logs=conversation_logs,
                                before_after_schedules=schedules,
                            )
                            st.session_state["last_research_bundle"] = bundle
                            st.success("✅ Research bundle exported to research_output/")

                    if st.session_state.get("last_research_bundle"):
                        bundle = st.session_state["last_research_bundle"]
                        st.markdown("##### Latest Export Manifest")
                        st.json({
                            "manifest": bundle.get("manifest"),
                            "metrics_csv": bundle.get("metrics_csv"),
                            "figures": bundle.get("figures"),
                            "screenshot_figures": bundle.get("screenshot_figures"),
                        })
                
                # CHANGE 1, 2: Semester Configurations Report
                elif report_type == "⚙️ Semester Configurations Report":
                    st.markdown("#### ⚙️ Semester Configurations Report")
                    st.markdown("View all custom lunch and break configurations per semester.")
                    
                    if st.button("🔄 Generate Report", key="gen_config_report"):
                        with st.spinner("Generating configurations report..."):
                            df = report_generator.generate_semester_config_report()
                            
                            if not df.empty and df['Program/Semester'].iloc[0] != 'No configs':
                                # Summary metrics
                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    custom_lunch = len(df[df['Custom Lunch'] == 'Yes'])
                                    st.metric("Custom Lunch Configs", custom_lunch)
                                with col2:
                                    breaks_enabled = len(df[df['Breaks Enabled'] == 'Yes'])
                                    st.metric("Break Configs", breaks_enabled)
                                with col3:
                                    locked = len(df[df['Lunch Locked'] == '🔒'])
                                    st.metric("Locked Configs", locked)
                                
                                st.dataframe(df, use_container_width=True, height=400)
                            else:
                                st.info("No semester configurations found. Configure lunch/breaks in the sidebar.")
                
                elif report_type == "📋 Comprehensive Report (All)":
                    st.markdown("#### 📋 Comprehensive Report")
                    
                    if st.button("📥 Generate & Download", type="primary", key="gen_comprehensive"):
                        with st.spinner("Generating comprehensive report..."):
                            try:
                                excel_data = report_generator.export_comprehensive_report_to_excel()
                                
                                st.success("✅ Report generated successfully!")
                                
                                st.download_button(
                                    label="📥 Download Comprehensive Report",
                                    data=excel_data,
                                    file_name=f"comprehensive_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="download_comprehensive"
                                )
                            except Exception as e:
                                st.error(f"Error generating report: {str(e)}")
            else:
                st.error("❌ Firebase not connected. Reports require Firebase connection.")
    
    # ==================== FACULTY PORTAL ====================
    elif st.session_state.portal == 'faculty':
        st.markdown('<h1 class="main-header">👨‍🏫 Faculty Portal</h1>', unsafe_allow_html=True)
        
        if st.button("← Back to Portal Selection"):
            st.session_state.portal = None
            st.rerun()
        
        st.markdown("### 📅 Faculty Timetables")
        st.markdown("View faculty schedules organized by school and program")
        
        if firebase_manager:
            # AUTO-LOAD: Check if data is already loaded
            if 'all_faculty_schedules' not in st.session_state or not st.session_state.all_faculty_schedules:
                
                with st.spinner("🔄 Loading faculty data from Firebase..."):
                    all_faculty_schedules = defaultdict(lambda: defaultdict(dict))
                    faculty_info = {}
                    faculty_by_school = defaultdict(set)
                    
                    all_timetables = firebase_manager.get_all_timetables()
                    
                    if all_timetables:
                        for timetable in all_timetables:
                            schedule = timetable.get('schedule', {})
                            timetable_id = timetable.get('year', 'Unknown')
                            
                            for school in schedule:
                                for batch in schedule[school]:
                                    for day in schedule[school][batch]:
                                        for slot, slot_val in schedule[school][batch][day].items():
                                            slot_items = slot_val if isinstance(slot_val, list) else [slot_val]
                                            for class_info in slot_items:
                                                if (class_info and isinstance(class_info, dict) 
                                                        and class_info.get('faculty') 
                                                        and class_info.get('type') not in ['LUNCH', 'BREAK']):
                                                    faculty_name = class_info['faculty']
                                                    
                                                    if faculty_name and faculty_name != 'TBD':
                                                        school_type = 'STME' if 'STME' in school else ('SOC' if 'SOC' in school else 'SOL')
                                                        faculty_by_school[school_type].add(faculty_name)
                                                        
                                                        if faculty_name not in faculty_info:
                                                            faculty_info[faculty_name] = {
                                                                'schools': set(),
                                                                'subjects': set(),
                                                                'total_hours': 0,
                                                                'morning_slots': 0  # CHANGE 3
                                                            }
                                                        
                                                        faculty_info[faculty_name]['schools'].add(school)
                                                        faculty_info[faculty_name]['subjects'].add(class_info.get('subject', ''))
                                                        faculty_info[faculty_name]['total_hours'] += 1
                                                        
                                                        # CHANGE 3: Track morning slots
                                                        if '09:00' in slot:
                                                            faculty_info[faculty_name]['morning_slots'] += 1
                                                        
                                                        # Use a list to store multiple classes in the same slot (clashes or combined batches)
                                                        if slot not in all_faculty_schedules[faculty_name][day]:
                                                            all_faculty_schedules[faculty_name][day][slot] = []
                                                        
                                                        all_faculty_schedules[faculty_name][day][slot].append({
                                                            'subject': class_info.get('subject', 'N/A'),
                                                            'room': class_info.get('room', 'TBD'),
                                                            'school': school,
                                                            'batch': batch,
                                                            'type': class_info.get('type', 'Theory'),
                                                            'timetable': timetable_id,
                                                            'duration': class_info.get('duration', DEFAULT_LECTURE_DURATION),
                                                            'start': class_info.get('start', ''),
                                                            'end': class_info.get('end', '')
                                                        })
                        
                        st.session_state.all_faculty_schedules = dict(all_faculty_schedules)
                        st.session_state.faculty_info = faculty_info
                        st.session_state.faculty_by_school = dict(faculty_by_school)
                        st.session_state.all_timetables_count = len(all_timetables)
                        st.session_state.all_timetables_list = all_timetables
                        st.session_state.faculty_data_loaded_at = datetime.now().strftime("%H:%M:%S")
                    else:
                        st.session_state.all_faculty_schedules = {}
                        st.session_state.faculty_info = {}
                        st.session_state.faculty_by_school = {}
                        st.session_state.all_timetables_count = 0
                        st.session_state.all_timetables_list = []
            
            # Display stats
            st.markdown("---")
            st.markdown("### 📊 Faculty Statistics")
            
            all_timetables_count = st.session_state.get('all_timetables_count', 0)
            faculty_info = st.session_state.get('faculty_info', {})
            faculty_by_school = st.session_state.get('faculty_by_school', {})
            
            stat_col1, stat_col2, stat_col3, stat_col4, stat_col5 = st.columns(5)
            
            with stat_col1:
                st.metric("📅 Total Timetables", all_timetables_count)
            with stat_col2:
                st.metric("👨‍🏫 Total Faculty", len(faculty_info))
            with stat_col3:
                st.metric("🔧 STME Faculty", len(faculty_by_school.get('STME', set())))
            with stat_col4:
                st.metric("💼 SOC Faculty", len(faculty_by_school.get('SOC', set())))
            with stat_col5:
                st.metric("⚖️ SOL Faculty", len(faculty_by_school.get('SOL', set())))
            
            # Refresh button
            col1, col2 = st.columns([1, 5])
            with col1:
                if st.button("🔄 Refresh", key="refresh_faculty_data"):
                    for key in ['all_faculty_schedules', 'faculty_info', 'faculty_by_school', 
                               'all_timetables_count', 'all_timetables_list']:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()
            with col2:
                st.caption(f"Last updated: {st.session_state.get('faculty_data_loaded_at', 'N/A')}")
            
            st.markdown("---")
            
            if st.session_state.get('all_faculty_schedules'):
                all_faculty_schedules = st.session_state.all_faculty_schedules
                faculty_info = st.session_state.faculty_info
                faculty_by_school = st.session_state.faculty_by_school
                
                # School-wise tabs
                school_tabs = st.tabs(["🏫 All Schools", "🔧 STME", "💼 SOC", "⚖️ SOL"])
                
                with school_tabs[0]:
                    st.markdown("### 📋 All Faculty Members")
                    
                    search_query = st.text_input("🔍 Search Faculty", placeholder="Enter faculty name...", key="search_all")
                    
                    filtered_faculties = list(faculty_info.keys())
                    if search_query:
                        filtered_faculties = [f for f in filtered_faculties if search_query.lower() in f.lower()]
                    
                    if filtered_faculties:
                        selected_faculty = st.selectbox(
                            "Select Faculty Member",
                            sorted(filtered_faculties),
                            key="faculty_selector_all"
                        )
                        
                        if selected_faculty and selected_faculty in all_faculty_schedules:
                            display_faculty_timetable(
                                selected_faculty, 
                                all_faculty_schedules[selected_faculty],
                                faculty_info[selected_faculty],
                                tab_id="all"
                            )
                    else:
                        st.info("No faculty members found.")
                
                with school_tabs[1]:
                    st.markdown("### 🔧 STME Faculty")
                    
                    stme_faculties = sorted(list(faculty_by_school.get('STME', set())))
                    
                    if stme_faculties:
                        stme_search = st.text_input("🔍 Search STME Faculty", key="search_stme")
                        
                        filtered_stme = stme_faculties
                        if stme_search:
                            filtered_stme = [f for f in stme_faculties if stme_search.lower() in f.lower()]
                        
                        if filtered_stme:
                            selected_stme_faculty = st.selectbox("Select Faculty", filtered_stme, key="faculty_selector_stme")
                            
                            if selected_stme_faculty and selected_stme_faculty in all_faculty_schedules:
                                display_faculty_timetable(
                                    selected_stme_faculty,
                                    all_faculty_schedules[selected_stme_faculty],
                                    faculty_info[selected_stme_faculty],
                                    tab_id="stme"
                                )
                    else:
                        st.info("No STME faculty data available.")
                
                with school_tabs[2]:
                    st.markdown("### 💼 SOC Faculty")
                    
                    soc_faculties = sorted(list(faculty_by_school.get('SOC', set())))
                    
                    if soc_faculties:
                        soc_search = st.text_input("🔍 Search SOC Faculty", key="search_soc")
                        
                        filtered_soc = soc_faculties
                        if soc_search:
                            filtered_soc = [f for f in soc_faculties if soc_search.lower() in f.lower()]
                        
                        if filtered_soc:
                            selected_soc_faculty = st.selectbox("Select Faculty", filtered_soc, key="faculty_selector_soc")
                            
                            if selected_soc_faculty and selected_soc_faculty in all_faculty_schedules:
                                display_faculty_timetable(
                                    selected_soc_faculty,
                                    all_faculty_schedules[selected_soc_faculty],
                                    faculty_info[selected_soc_faculty],
                                    tab_id="soc"
                                )
                    else:
                        st.info("No SOC faculty data available.")
                
                with school_tabs[3]:
                    st.markdown("### ⚖️ SOL Faculty")
                    
                    sol_faculties = sorted(list(faculty_by_school.get('SOL', set())))
                    
                    if sol_faculties:
                        sol_search = st.text_input("🔍 Search SOL Faculty", key="search_sol")
                        
                        filtered_sol = sol_faculties
                        if sol_search:
                            filtered_sol = [f for f in sol_faculties if sol_search.lower() in f.lower()]
                        
                        if filtered_sol:
                            selected_sol_faculty = st.selectbox("Select Faculty", filtered_sol, key="faculty_selector_sol")
                            
                            if selected_sol_faculty and selected_sol_faculty in all_faculty_schedules:
                                display_faculty_timetable(
                                    selected_sol_faculty,
                                    all_faculty_schedules[selected_sol_faculty],
                                    faculty_info[selected_sol_faculty],
                                    tab_id="sol"
                                )
                    else:
                        st.info("No SOL faculty data available.")
            else:
                st.warning("⚠️ No faculty data found. Please generate timetables from Admin Portal first.")
        else:
            st.error("❌ Firebase not connected.")
    
    # ==================== STUDENT PORTAL ====================
    elif st.session_state.portal == 'student':
        st.markdown('<h1 class="main-header">👨‍🎓 Student Portal</h1>', unsafe_allow_html=True)
        
        if st.button("← Back to Portal Selection"):
            st.session_state.portal = None
            st.rerun()
        
        st.markdown("### 📚 View Your Timetable")
        
        # Load from Firebase
        if firebase_manager:
            if st.button("📥 Load Latest from Firebase"):
                timetables = firebase_manager.get_all_timetables()
                if timetables:
                    st.session_state.current_schedule = timetables[0].get('schedule')
                    st.session_state.current_semester_config = timetables[0].get('semester_config')
                    st.success("✅ Loaded from Firebase")
        
        # CHANGE SECTION-BATCH-4: Student Portal section+batch selectors
        # Uses section_batch_schedules when available; falls back to raw schedule view.
        _stud_sbs = st.session_state.get('section_batch_schedules', {})

        if _stud_sbs:
            # --- Section/Batch-aware path ---
            _stud_prog_keys = []
            for _pk, _sem_dict in _stud_sbs.items():
                for _sm in _sem_dict:
                    _stud_prog_keys.append(f"{_pk} – Sem {_sm}")

            if _stud_prog_keys:
                _sc1, _sc2, _sc3, _sc4 = st.columns(4)
                with _sc1:
                    _stud_pk_sel = st.selectbox("Your Program / Semester", _stud_prog_keys, key="stud_prog_sem")
                _stud_vpk, _stud_sem_str = _stud_pk_sel.split(" – Sem ", 1)
                try:
                    _stud_vsem = int(_stud_sem_str)
                except ValueError:
                    _stud_vsem = _stud_sem_str

                _stud_sec_dict = _stud_sbs.get(_stud_vpk, {}).get(_stud_vsem, {})
                _stud_secs = [s for s in _stud_sec_dict.keys() if s != '_section']

                with _sc2:
                    _stud_sec = st.selectbox("Your Section", _stud_secs if _stud_secs else ['A'], key="stud_sec")
                with _sc3:
                    _stud_batch_nums = [b for b in _stud_sec_dict.get(_stud_sec, {}).keys() if b != '_section']
                    _stud_batch_sel = st.selectbox("Your Batch", _stud_batch_nums if _stud_batch_nums else [1], key="stud_batch_sel")
                with _sc4:
                    _stud_view_btn = st.button("📅 View My Timetable", type="primary", key="stud_view_btn")

                if _stud_view_btn:
                    # CHANGE PARALLEL-BATCH-3: Student Portal uses combined section view if batch selection is 'All' or specific
                    # For simplicity in this display logic, we enable the combined view for students too
                    _sec_data = _stud_sec_dict.get(_stud_sec, {})
                    batch_schedule = _sec_data.get('_combined')
                    if not batch_schedule:
                        batch_schedule = _sec_data.get(_stud_batch_sel, {})
                    
                    school = _stud_vpk
                    batch = f"Sec{_stud_sec}_Batch{_stud_batch_sel}"

                    st.markdown(f"### 📅 Timetable – {_stud_vpk} | Section {_stud_sec} (Batch {_stud_batch_sel})")
        else:
            # --- Fallback: original raw-schedule path (existing behaviour) ---
            if st.session_state.current_schedule:
                schedule = st.session_state.current_schedule

                col1, col2, col3 = st.columns(3)
                with col1:
                    if schedule:
                        school = st.selectbox("Select Your School", list(schedule.keys()), key="student_school")
                with col2:
                    if school and school in schedule:
                        batches = list(schedule[school].keys())
                        batch = st.selectbox("Select Your Batch", batches, key="student_batch")
                with col3:
                    view_btn = st.button("📅 View My Timetable", type="primary")

                if view_btn and school and batch:
                    batch_schedule = schedule[school][batch]
                    st.markdown(f"### 📅 Timetable for {school} - {batch}")
            else:
                st.info("No timetables available. Please contact your administrator.")

        # --- Shared display block for both paths (batch_schedule must be defined above) ---
        if st.session_state.current_schedule or _stud_sbs:
            if 'batch_schedule' in dir() and batch_schedule:
                
                # CHANGE 1: Get dynamic time slots
                time_slots = ExportManager.get_time_slots_from_schedule(batch_schedule)
                days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
                
                timetable_data = []
                total_classes = 0
                
                for day in days:
                    row = {'Day': day}
                    for slot in time_slots:
                        display_slot = ExportManager.format_slot_for_display(slot)
                        
                        if day in batch_schedule and slot in batch_schedule[day]:
                            class_info = batch_schedule[day][slot]
                            if class_info:
                                slot_items = (
                                    class_info
                                    if isinstance(class_info, list)
                                    else [class_info]
                                )
                                cell_parts = []
                                for item in slot_items:
                                    if not isinstance(item, dict):
                                        continue
                                    if item.get('type') == 'LUNCH':
                                        duration = item.get('duration', 50)
                                        cell_parts.append(f"🍴 LUNCH ({duration} min)")
                                    elif item.get('type') == 'BREAK':
                                        duration = item.get('duration', 10)
                                        cell_parts.append(f"☕ BREAK ({duration} min)")
                                    else:
                                        cell_text = f"📚 {item.get('subject', 'N/A')}\n"
                                        cell_text += f"👨‍🏫 {item.get('faculty', 'TBD')}\n"
                                        cell_text += f"📍 {item.get('room', 'TBD')}"
                                        cell_parts.append(cell_text)
                                        total_classes += 1
                                row[display_slot] = (
                                    "\n\n".join(cell_parts) if cell_parts else "FREE"
                                )
                            else:
                                row[display_slot] = "FREE"
                        else:
                            row[display_slot] = "FREE"
                    timetable_data.append(row)
                
                df = pd.DataFrame(timetable_data)
                st.dataframe(df, use_container_width=True, height=400)
                
                st.markdown("### 📊 Your Schedule Statistics")
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Classes/Week", total_classes)
                with col2:
                    avg_daily = total_classes / 5 if total_classes > 0 else 0
                    st.metric("Avg Classes/Day", f"{avg_daily:.1f}")
                with col3:
                    lecture_slots = len([s for s in time_slots if 'LUNCH' not in str(batch_schedule.get('Monday', {}).get(s, {})) 
                                        and 'BREAK' not in str(batch_schedule.get('Monday', {}).get(s, {}))])
                    free_periods = (5 * lecture_slots) - total_classes
                    st.metric("Free Periods", max(0, free_periods))
                with col4:
                    st.metric("Days/Week", "5")
                
                # Export options
                st.markdown("---")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    csv = df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download CSV",
                        data=csv,
                        file_name=f"my_timetable_{school}_{batch}.csv",
                        mime="text/csv"
                    )
                
                with col2:
                    excel_data = ExportManager.export_to_excel_formatted(
                        batch_schedule, school_name=school, batch_name=batch
                    )
                    st.download_button(
                        label="📥 Download Excel",
                        data=excel_data,
                        file_name=f"my_timetable_{school}_{batch}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with col3:
                    pdf_buffer = ExportManager.export_to_pdf_detailed(
                        batch_schedule, school_name=school, batch_name=batch
                    )
                    st.download_button(
                        label="📥 Download PDF",
                        data=pdf_buffer,
                        file_name=f"my_timetable_{school}_{batch}.pdf",
                        mime="application/pdf"
                    )
        else:
            st.info("No timetables available. Please contact your administrator.")


if __name__ == "__main__":
    main()
